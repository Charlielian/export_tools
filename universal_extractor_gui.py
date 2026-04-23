# -*- coding: utf-8 -*-
"""
通用数据提取工具 - GUI版本
支持多种数据表：4G干扰小区、5G干扰小区、5G小区容量报表等

使用方法：
    python universal_extractor_gui.py

特点：所有依赖模块已内嵌，可独立运行
"""
import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox
import threading
import os
import sys
import time
import struct
from datetime import datetime, timedelta
import calendar
import queue
import logging
import io

# ==================== 全局日志系统 ====================
class TeeLogger:
    """同时输出到控制台和日志文件的日志系统"""
    _instance = None
    _log_file = None
    
    @classmethod
    def get_instance(cls):
        if cls._instance is None:
            cls._instance = cls()
        return cls._instance
    
    def __init__(self):
        self.handlers = []
        
    def add_file_handler(self, filepath):
        """添加文件日志处理器"""
        self._log_file = filepath
        # 创建文件handler
        handler = logging.FileHandler(filepath, encoding='utf-8')
        handler.setLevel(logging.DEBUG)
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s',
                                      datefmt='%Y-%m-%d %H:%M:%S')
        handler.setFormatter(formatter)
        logging.root.addHandler(handler)
        self.handlers.append(handler)
        
    def write(self, message):
        """同时输出到控制台和日志文件"""
        # 输出到控制台
        sys.__stdout__.write(message)
        # 输出到日志文件
        if self._log_file:
            with open(self._log_file, 'a', encoding='utf-8') as f:
                f.write(message)

# 全局print替换函数
_original_print = print
_log_file_path = None

def set_log_file(filepath):
    """设置日志文件路径并启用日志记录"""
    global _log_file_path
    _log_file_path = filepath
    TeeLogger.get_instance().add_file_handler(filepath)

def debug_print(*args, **kwargs):
    """增强的print函数，同时输出到控制台和日志文件"""
    # 获取调用栈信息来确定日志级别标签
    import traceback
    stack = traceback.extract_stack()
    caller = stack[-2] if len(stack) >= 2 else None
    
    # 构建输出字符串
    output = io.StringIO()
    kwargs['file'] = output
    kwargs['end'] = kwargs.get('end', '\n')
    _original_print(*args, **kwargs)
    message = output.getvalue()
    output.close()
    
    # 添加标签前缀
    prefix = ""
    if caller:
        filename = os.path.basename(caller.filename)
        # 根据print内容判断日志级别
        msg_str = str(args[0]) if args else ""
        if '[ERROR' in msg_str or '失败' in msg_str or '异常' in msg_str:
            prefix = ""
        elif '[WARNING' in msg_str or '警告' in msg_str:
            prefix = ""
        elif '[SUCCESS' in msg_str:
            prefix = ""
        elif '[DEBUG' in msg_str:
            prefix = ""
    
    # 输出到控制台（重置file参数，使用默认的sys.stdout）
    _print_kwargs = {k: v for k, v in kwargs.items() if k != 'file'}
    _original_print(*args, **_print_kwargs)
    
    # 同时写入日志文件
    if _log_file_path:
        with open(_log_file_path, 'a', encoding='utf-8') as f:
            f.write(message)

# 替换全局print函数
print = debug_print

# ==================== 第三方依赖 ====================
import requests
import pandas as pd
from lxml import etree
from openpyxl import load_workbook, Workbook
try:
    from Crypto.PublicKey import RSA
    from Crypto.Cipher import PKCS1_v1_5
    from Crypto.Cipher import AES as AES_Cipher
    from Crypto.Util.Padding import pad, unpad
except ModuleNotFoundError:
    from Cryptodome.PublicKey import RSA
    from Cryptodome.Cipher import PKCS1_v1_5
    from Cryptodome.Cipher import AES as AES_Cipher
    from Cryptodome.Util.Padding import pad, unpad
import base64
from urllib.parse import quote
import random
import pickle
import json
import yaml


# ==================== 授权配置 ====================
# 到期日期（格式：YYYY-MM-DD）
# 修改这里即可设置软件到期日期
EXPIRY_DATE = "2026-06-30"

# license.dat 文件位置（与脚本同目录）
LICENSE_FILE = "license.dat"


# ==================== 硬件信息获取（用于生成机器码） ====================
import hashlib
import platform
import subprocess


def get_macos_hw_info():
    """获取macOS硬件信息"""
    hw_info = {"cpu_id": "", "board_sn": "", "disk_sn": "", "mac": ""}
    try:
        cmd = ["ioreg", "-l", "-w0", "-r", "-c", "IOPlatformExpertDevice", "-d", "2"]
        output = subprocess.check_output(cmd, stderr=subprocess.DEVNULL).decode("utf-8")
        hw_info["board_sn"] = output.split('"IOPlatformSerialNumber" = ')[1].split('"')[1].strip()
    except:
        hw_info["board_sn"] = "unknown_board"
    try:
        cmd = ["sysctl", "-n", "machdep.cpu.brand_string"]
        hw_info["cpu_id"] = subprocess.check_output(cmd).decode("utf-8").strip()
    except:
        hw_info["cpu_id"] = "unknown_cpu"
    try:
        cmd = ["diskutil", "info", "/"]
        output = subprocess.check_output(cmd).decode("utf-8")
        hw_info["disk_sn"] = output.split("Volume UUID:")[1].split("\n")[0].strip()
    except:
        hw_info["disk_sn"] = "unknown_disk"
    try:
        cmd = ["ifconfig", "en0"]
        output = subprocess.check_output(cmd, stderr=subprocess.DEVNULL).decode("utf-8")
        hw_info["mac"] = output.split("ether")[1].split(" ")[1].strip().replace(":", "")
    except:
        hw_info["mac"] = "unknown_mac"
    return hw_info


def get_windows_hw_info():
    """获取Windows硬件信息"""
    try:
        import wmi
        c = wmi.WMI()
        hw_info = {"cpu_id": "", "board_sn": "", "disk_sn": "", "mac": ""}
        cpu_list = c.Win32_Processor()
        hw_info["cpu_id"] = cpu_list[0].ProcessorId.strip() if (cpu_list and cpu_list[0].ProcessorId) else "unknown_cpu"
        board_list = c.Win32_BaseBoard()
        hw_info["board_sn"] = board_list[0].SerialNumber.strip() if (board_list and board_list[0].SerialNumber) else "unknown_board"
        disk_list = c.Win32_DiskDrive()
        hw_info["disk_sn"] = disk_list[0].SerialNumber.strip() if (disk_list and disk_list[0].SerialNumber) else "unknown_disk"
        nic_list = c.Win32_NetworkAdapterConfiguration(IPEnabled=True)
        hw_info["mac"] = nic_list[0].MACAddress.strip().replace(":", "") if (nic_list and nic_list[0].MACAddress) else "unknown_mac"
        return hw_info
    except:
        return {"cpu_id": "unknown", "board_sn": "unknown", "disk_sn": "unknown", "mac": "unknown"}


def get_linux_hw_info():
    """获取Linux硬件信息"""
    hw_info = {"cpu_id": "", "board_sn": "", "disk_sn": "", "mac": ""}
    try:
        with open("/proc/cpuinfo", "r") as f:
            for line in f.readlines():
                if "serial" in line.lower():
                    hw_info["cpu_id"] = line.split(":")[1].strip()
                    break
        hw_info["cpu_id"] = hw_info["cpu_id"] if hw_info["cpu_id"] else "unknown_cpu"
    except:
        hw_info["cpu_id"] = "unknown_cpu"
    try:
        with open("/sys/devices/virtual/dmi/id/board_serial", "r") as f:
            hw_info["board_sn"] = f.read().strip()
    except:
        hw_info["board_sn"] = "unknown_board"
    try:
        cmd = ["lsblk", "-o", "SERIAL", "-n", "/dev/sda"]
        hw_info["disk_sn"] = subprocess.check_output(cmd, stderr=subprocess.DEVNULL).decode("utf-8").strip()
    except:
        hw_info["disk_sn"] = "unknown_disk"
    try:
        for path in ["/sys/class/net/eth0/address", "/sys/class/net/ens33/address"]:
            if os.path.exists(path):
                with open(path, "r") as f:
                    hw_info["mac"] = f.read().strip().replace(":", "")
                break
        hw_info["mac"] = hw_info.get("mac", "unknown_mac")
    except:
        hw_info["mac"] = "unknown_mac"
    return hw_info


def get_hw_info():
    """跨平台获取硬件信息"""
    system = platform.system()
    if system == "Windows":
        return get_windows_hw_info()
    elif system == "Darwin":
        return get_macos_hw_info()
    elif system == "Linux":
        return get_linux_hw_info()
    else:
        return {"cpu_id": "unknown", "board_sn": "unknown", "disk_sn": "unknown", "mac": "unknown"}


def generate_machine_code(hw_info):
    """生成机器码（基于硬件信息）"""
    raw_str = f"{hw_info['cpu_id']}-{hw_info['board_sn']}-{hw_info['disk_sn']}-{hw_info['mac']}"
    return hashlib.sha256(raw_str.encode("utf-8")).hexdigest()


# ==================== 配置文件加载 ====================
def load_config():
    """从 YAML 文件加载配置"""
    config = {
        'auth': {
            'username': 'XXXXX',
            'password': 'XXXX'
        },
        'paths': {
            'output_dir': './data_output',
            'cookie_dir': './cookies',
            'captcha_dir': './captcha_images',
            'log_dir': './logs'
        },
        'server': {
            'base_url': 'https://nqi.gmcc.net:20443'
        }
    }

    config_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'config.yaml')
    if os.path.exists(config_file):
        try:
            with open(config_file, 'r', encoding='utf-8') as f:
                yaml_config = yaml.safe_load(f)
                if yaml_config:
                    config.update(yaml_config)
        except Exception as e:
            print(f"警告：加载配置文件失败 ({config_file}): {e}")

    return config


# ==================== 配置区域 ====================
_config = load_config()

DEFAULT_USERNAME = _config['auth']['username']
DEFAULT_PASSWORD = _config['auth']['password']
OUTPUT_DIR = _config['paths']['output_dir']
COOKIE_DIR = _config['paths']['cookie_dir']
CAPTCHA_DIR = _config['paths']['captcha_dir']
LOG_DIR = _config['paths']['log_dir']

BASE_URL = _config['server']['base_url']
LOGIN_URL = f'{BASE_URL}/cas/login?service={BASE_URL}/pro-portal/'
CAPTCHA_URL = f'{BASE_URL}/cas/captcha.jpg'
GET_CONFIG_URL = f'{BASE_URL}/cas/getConfig'
SEND_CODE_URL = f'{BASE_URL}/cas/sendCode1'
JXCX_URL = f'{BASE_URL}/pro-adhoc/adhocquery/getTable'
JXCX_COUNT_URL = f'{BASE_URL}/pro-adhoc/adhocquery/getTableCount'
JXCX_SEARCH_URL = f'{BASE_URL}/pro-adhoc/adhocquery/search'
JXCX_TABLE_URL = f'{BASE_URL}/pro-adhoc/adhocquery/getSelectTable'

# 分批查询配置
MAX_SINGLE_QUERY = 500000  # 超过此数量自动分批

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/102.0.0.0 Safari/537.36',
    'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8'
}

HEADERS_JSON = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/102.0.0.0 Safari/537.36',
    'Content-Type': 'application/json'
}


# ==================== 工具函数 ====================
def ensure_dirs():
    """确保必要的目录存在"""
    for dir_path in [OUTPUT_DIR, COOKIE_DIR, CAPTCHA_DIR, LOG_DIR]:
        os.makedirs(dir_path, exist_ok=True)


def save_cookie(cookie, username):
    """保存cookie到文件"""
    ensure_dirs()
    filepath = os.path.join(COOKIE_DIR, f'{username}.pkl')
    with open(filepath, 'wb') as f:
        pickle.dump(cookie, f)


def load_cookie(username):
    """从文件加载cookie"""
    filepath = os.path.join(COOKIE_DIR, f'{username}.pkl')
    if os.path.exists(filepath):
        with open(filepath, 'rb') as f:
            return pickle.load(f)
    return None


def rsa_encrypt(data, public_key):
    """RSA加密"""
    public_key = '-----BEGIN PUBLIC KEY-----\n' + public_key + '\n-----END PUBLIC KEY-----'
    rsa_key = RSA.importKey(public_key)
    cipher = PKCS1_v1_5.new(rsa_key)
    encrypted_data = base64.b64encode(cipher.encrypt(data.encode(encoding="utf-8")))
    return encrypted_data.decode('utf-8')


def captcha_handle(img_content, attempt=1):
    """验证码处理（OCR识别）"""
    try:
        from PIL import Image, ImageFilter
        import pytesseract
        from io import BytesIO
        
        bytes_stream = BytesIO(img_content)
        img = Image.open(bytes_stream)
        img_gray = img.convert('L')
        img_black_white = img_gray.point(lambda x: 255 if x > 85 else 0)
        img_qucao = img_black_white.filter(ImageFilter.SMOOTH_MORE)
        img = img_qucao.convert('RGB')
        
        config = '--psm 7 --oem 3 -c tessedit_char_whitelist=0123456789abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ'
        result = pytesseract.image_to_string(img, config=config)[0:4].replace('\n', '')
        return result
    except Exception as e:
        print(f"验证码OCR识别失败: {e}")
        ensure_dirs()
        img_path = os.path.join(CAPTCHA_DIR, f'captcha_{attempt}.jpg')
        with open(img_path, 'wb') as f:
            f.write(img_content)
        return None


# ==================== 登录模块 ====================
class LoginDialog:
    """登录对话框 - 处理验证码和短信验证"""

    def __init__(self, parent, username, password, session):
        self.username = username
        self.password = password
        self.sess = session
        self.result = False
        self.msg_code = None

        self.dialog = tk.Toplevel(parent)
        self.dialog.title("登录验证")
        self.dialog.geometry("400x480")
        self.dialog.resizable(False, False)
        self.dialog.transient(parent)
        self.dialog.grab_set()

        # 居中显示
        self.dialog.update_idletasks()
        x = (self.dialog.winfo_screenwidth() // 2) - (400 // 2)
        y = (self.dialog.winfo_screenheight() // 2) - (480 // 2)
        self.dialog.geometry(f'400x480+{x}+{y}')

        self._create_widgets()
        self._fetch_captcha()

        self.dialog.protocol("WM_DELETE_WINDOW", self._on_close)

    def _create_widgets(self):
        """创建对话框组件"""
        main_frame = ttk.Frame(self.dialog, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # 标题
        ttk.Label(main_frame, text="请完成安全验证",
                 font=("Arial", 12, "bold")).pack(pady=(0, 15))

        # 验证码图片区域
        captcha_frame = ttk.LabelFrame(main_frame, text="图形验证码", padding="10")
        captcha_frame.pack(fill=tk.X, pady=10)

        self.captcha_label = ttk.Label(captcha_frame)
        self.captcha_label.pack()

        captcha_input_frame = ttk.Frame(captcha_frame)
        captcha_input_frame.pack(fill=tk.X, pady=10)

        ttk.Label(captcha_input_frame, text="验证码:").pack(side=tk.LEFT)
        self.captcha_var = tk.StringVar()
        self.captcha_entry = ttk.Entry(captcha_input_frame, textvariable=self.captcha_var, width=10)
        self.captcha_entry.pack(side=tk.LEFT, padx=10)
        self.captcha_entry.focus()

        ttk.Button(captcha_input_frame, text="刷新",
                  command=self._fetch_captcha).pack(side=tk.LEFT, padx=5)
        ttk.Button(captcha_input_frame, text="验证图形码",
                  command=self._verify_captcha).pack(side=tk.LEFT)

        self.captcha_msg = ttk.Label(captcha_frame, text="", foreground="blue")
        self.captcha_msg.pack(pady=5)

        # 短信验证码区域
        sms_frame = ttk.LabelFrame(main_frame, text="短信验证码", padding="10")
        sms_frame.pack(fill=tk.X, pady=10)

        self.sms_var = tk.StringVar()
        sms_input_frame = ttk.Frame(sms_frame)
        sms_input_frame.pack(fill=tk.X, pady=5)

        ttk.Label(sms_input_frame, text="短信码:").pack(side=tk.LEFT)
        self.sms_entry = ttk.Entry(sms_input_frame, textvariable=self.sms_var, width=15)
        self.sms_entry.pack(side=tk.LEFT, padx=10)
        self.sms_entry.bind('<Return>', lambda e: self._submit())

        self.send_sms_btn = ttk.Button(sms_input_frame, text="发送短信",
                                       command=self._send_sms, state=tk.DISABLED)
        self.send_sms_btn.pack(side=tk.LEFT, padx=5)

        self.sms_msg = ttk.Label(sms_frame, text="请先验证图形验证码", foreground="gray")
        self.sms_msg.pack(pady=5)

        self.countdown_label = ttk.Label(sms_frame, text="", foreground="red")
        self.countdown_label.pack()

        # 按钮区域
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(pady=20)

        self.submit_btn = ttk.Button(btn_frame, text="确认登录",
                                    command=self._submit, width=15)
        self.submit_btn.pack(side=tk.LEFT, padx=10)

        ttk.Button(btn_frame, text="取消",
                  command=self._on_close, width=10).pack(side=tk.LEFT)

        # 状态显示
        self.status_var = tk.StringVar(value="请先输入图形验证码，点击【验证图形码】")
        ttk.Label(main_frame, textvariable=self.status_var,
                 foreground="green").pack(pady=10)

    def _verify_captcha(self):
        """验证图形验证码"""
        captcha_code = self.captcha_var.get().strip()

        if not captcha_code:
            self.captcha_msg.config(text="请输入图形验证码", foreground="red")
            self.status_var.set("请先输入图形验证码")
            return

        self.status_var.set("正在验证图形验证码...")
        self.captcha_msg.config(text="验证中...")

        try:
            public_key = None
            html_res = self.sess.get(LOGIN_URL, headers=HEADERS)
            html_res.encoding = 'utf-8'
            html = etree.HTML(html_res.text)
            public_key = html.xpath('//*[@type="text/javascript"]/text()')[0].split('setPublicKey("')[1].split('")')[0]

            username_e = rsa_encrypt(self.username, public_key)
            password_e = rsa_encrypt(self.password, public_key)

            data = {
                'password': password_e,
                'loginId': username_e,
                'captcha': captcha_code,
            }
            res = self.sess.post(GET_CONFIG_URL, data=json.dumps(data), headers=HEADERS_JSON)

            if res.status_code == 200:
                result = json.loads(res.text)
                if result.get('code') == '1':
                    self.captcha_msg.config(text="图形验证码正确 ✓", foreground="green")
                    self.captcha_entry.config(state=tk.DISABLED)
                    self.send_sms_btn.config(state=tk.NORMAL)
                    self.status_var.set("图形验证码正确！请点击【发送短信】")
                    # 保存加密后的密码供后续使用
                    self._encrypted_username = username_e
                    self._encrypted_password = password_e
                else:
                    self.captcha_msg.config(text="图形验证码错误，请重试", foreground="red")
                    self.status_var.set("图形验证码错误，请重新输入")
                    self.captcha_entry.select_clear()
                    self.captcha_entry.focus()
            else:
                self.status_var.set("验证请求失败")
        except Exception as e:
            self.status_var.set(f"验证失败: {e}")
            self.captcha_msg.config(text=f"错误: {e}", foreground="red")

    def _fetch_captcha(self):
        """获取图形验证码"""
        try:
            captcha_res = self.sess.get(CAPTCHA_URL)
            if captcha_res.status_code == 200:
                from PIL import Image, ImageTk
                from io import BytesIO

                bytes_stream = BytesIO(captcha_res.content)
                img = Image.open(bytes_stream)
                img = img.resize((200, 80), Image.Resampling.LANCZOS)
                self.captcha_photo = ImageTk.PhotoImage(img)
                self.captcha_label.config(image=self.captcha_photo)

                self.captcha_img_data = captcha_res.content
                self.captcha_var.set('')  # 清空验证码输入
                self.captcha_entry.config(state=tk.NORMAL)  # 启用输入框
                self.send_sms_btn.config(state=tk.DISABLED)
                self.status_var.set("请输入图形验证码，点击【验证图形码】")
                self.captcha_msg.config(text="请输入图片中的验证码")
                self.sms_msg.config(text="请先验证图形验证码")
                self.captcha_entry.focus()
            else:
                self.status_var.set("获取验证码失败")
        except Exception as e:
            self.status_var.set(f"获取验证码失败: {e}")

    def _send_sms(self):
        """发送短信验证码"""
        if not hasattr(self, '_encrypted_username') or not hasattr(self, '_encrypted_password'):
            self.captcha_msg.config(text="请先验证图形验证码", foreground="red")
            self.status_var.set("请先完成图形验证码验证")
            return

        self.status_var.set("正在发送短信验证码...")
        self.send_sms_btn.config(state=tk.DISABLED)

        try:
            # 发送短信
            data_sms = {'loginId': self._encrypted_username, 'password': self._encrypted_password}
            res_sms = self.sess.post(SEND_CODE_URL, data=json.dumps(data_sms), headers=HEADERS_JSON)
            result_sms = json.loads(res_sms.text)

            if result_sms.get('msg') == 'success':
                self.sms_msg.config(text="短信验证码已发送，请查收 ✓", foreground="green")
                self.status_var.set("短信已发送，请在下方输入短信验证码后按回车或点击【确认登录】")
                self._start_countdown(60)
                self.sms_entry.focus()
            else:
                self.sms_msg.config(text="短信发送失败，请重试", foreground="red")
                self.status_var.set("短信发送失败")
                self.send_sms_btn.config(state=tk.NORMAL)

        except Exception as e:
            self.sms_msg.config(text=f"发送失败: {e}", foreground="red")
            self.status_var.set(f"短信发送失败: {e}")
            self.send_sms_btn.config(state=tk.NORMAL)

    def _start_countdown(self, seconds):
        """倒计时"""
        if seconds > 0:
            self.countdown_label.config(text=f"请在 {seconds} 秒内输入")
            self.dialog.after(1000, lambda: self._start_countdown(seconds - 1))
        else:
            self.countdown_label.config(text="验证码可能已过期")

    def _submit(self):
        """提交验证"""
        msg_code = self.sms_var.get().strip()

        if not msg_code:
            self.sms_msg.config(text="请输入短信验证码", foreground="red")
            self.status_var.set("请输入短信验证码")
            return

        if not hasattr(self, '_encrypted_username') or not hasattr(self, '_encrypted_password'):
            self.sms_msg.config(text="请先完成图形验证码验证", foreground="red")
            self.status_var.set("请先完成图形验证码验证")
            return

        self.status_var.set("正在验证短信验证码...")
        self.submit_btn.config(state=tk.DISABLED)

        try:
            html_res = self.sess.get(LOGIN_URL, headers=HEADERS)
            html_res.encoding = 'utf-8'
            html = etree.HTML(html_res.text)
            execution = html.xpath('//*[@id="fm1"]/div[4]/input[1]')[0].attrib.get('value')

            login_data = {
                'password': self._encrypted_password,
                'username': self._encrypted_username,
                'msgCode': msg_code,
                'captcha': self.captcha_var.get().strip(),
                'uuid': '',
                'execution': execution,
                '_eventId': 'submit',
                'geolocation': ''
            }

            res_login = self.sess.post(LOGIN_URL, data=login_data, headers=HEADERS)

            if self.sess.cookies.get('CASTGC'):
                self.msg_code = msg_code
                self.result = True
                self.status_var.set("登录成功！")
                self.sms_msg.config(text="登录成功 ✓", foreground="green")
                self.dialog.after(500, self.dialog.destroy)
            else:
                self.sms_msg.config(text="短信验证码错误", foreground="red")
                self.status_var.set("短信验证码错误，请重试")
                self.submit_btn.config(state=tk.NORMAL)

        except Exception as e:
            self.status_var.set(f"验证失败: {e}")
            self.sms_msg.config(text=f"错误: {e}", foreground="red")
            self.submit_btn.config(state=tk.NORMAL)

    def _on_close(self):
        """关闭窗口"""
        self.result = False
        self.dialog.destroy()

    def show(self):
        """显示对话框并返回结果"""
        self.dialog.wait_window()
        return self.result


class LoginManager:
    """登录管理器"""

    def __init__(self, username=None, password=None, parent=None):
        self.username = username or DEFAULT_USERNAME
        self.password = password or DEFAULT_PASSWORD
        self.parent = parent  # GUI 父窗口
        self.sess = requests.Session()

    def login(self, try_times=3):
        """执行登录"""
        print("=" * 60)
        print("开始登录大数据平台...")
        print(f"账号: {self.username}")
        print("=" * 60)

        saved_cookie = load_cookie(self.username)
        if saved_cookie:
            self.sess.cookies = saved_cookie
            if self._check_session():
                print("✓ 使用保存的Cookie登录成功！")
                return True
            else:
                print("⚠ 保存的Cookie已失效，重新登录...")
                self.sess = requests.Session()

        for i in range(try_times):
            if self._login_once(i):
                print(f"✓ 登录成功！（尝试次数: {i+1}）")
                save_cookie(self.sess.cookies, self.username)
                return True
            else:
                print(f"⚠ 登录失败，尝试次数: {i+1}/{try_times}")

        print("✗ 登录失败，已达到最大尝试次数")
        return False

    def _check_session(self):
        """检查session是否有效"""
        try:
            url = f'{BASE_URL}/pro-wfm-biz-server/cas/login/info'
            res = self.sess.get(url, headers=HEADERS, timeout=10)
            if res.status_code == 200:
                data = json.loads(res.text)
                return data.get('data', {}).get('loginId') == self.username
        except:
            pass
        return False

    def _login_once(self, attempt=0):
        """执行一次完整的登录流程"""
        try:
            # 如果有父窗口，使用 GUI 对话框
            if self.parent:
                return self._login_with_gui()

            # 否则使用命令行模式
            return self._login_with_input(attempt)

        except Exception as e:
            print(f"✗ 登录过程出错: {e}")
            import traceback
            traceback.print_exc()
            return False

    def _login_with_gui(self):
        """使用 GUI 对话框进行登录验证"""
        dialog = LoginDialog(self.parent, self.username, self.password, self.sess)
        result = dialog.show()
        return result

    def _login_with_input(self, attempt=0):
        """使用命令行输入进行登录验证"""
        try:
            res = self.sess.get(LOGIN_URL, headers=HEADERS)
            res.encoding = 'utf-8'
            html = etree.HTML(res.text)

            execution = html.xpath('//*[@id="fm1"]/div[4]/input[1]')[0].attrib.get('value')
            public_key = html.xpath('//*[@type="text/javascript"]/text()')[0].split('setPublicKey("')[1].split('")')[0]

            username_e = rsa_encrypt(self.username, public_key)
            password_e = rsa_encrypt(self.password, public_key)

            captcha_code = None
            for i in range(10):
                captcha_res = self.sess.get(CAPTCHA_URL)

                if i < 5:
                    captcha_code = captcha_handle(captcha_res.content, attempt * 10 + i)

                if not captcha_code:
                    ensure_dirs()
                    img_path = os.path.join(CAPTCHA_DIR, f'captcha_{attempt}_{i}.jpg')
                    with open(img_path, 'wb') as f:
                        f.write(captcha_res.content)
                    captcha_code = input(f'请查看图片 {img_path} 并输入验证码: ')

                data = {
                    'password': password_e,
                    'loginId': username_e,
                    'captcha': captcha_code,
                }
                res = self.sess.post(GET_CONFIG_URL, data=json.dumps(data), headers=HEADERS_JSON)

                if res.status_code == 200:
                    result = json.loads(res.text)
                    if result.get('code') == '1':
                        print(f"✓ 图形验证码验证通过（尝试次数: {i+1}）")
                        break
                    else:
                        print(f"⚠ 图形验证码错误，重试...")
                        captcha_code = None
            else:
                print("✗ 图形验证码验证失败次数过多")
                return False

            if attempt == 0:
                data_sms = {'loginId': username_e, 'password': password_e}
                res_sms = self.sess.post(SEND_CODE_URL, data=json.dumps(data_sms), headers=HEADERS_JSON)
                result_sms = json.loads(res_sms.text)
                if result_sms.get('msg') == 'success':
                    print("✓ 短信验证码已发送")
                else:
                    print("⚠ 短信验证码发送失败")

            msg_code = input('请输入短信验证码: ')

            login_data = {
                'password': password_e,
                'username': username_e,
                'msgCode': msg_code,
                'captcha': captcha_code,
                'uuid': '',
                'execution': execution,
                '_eventId': 'submit',
                'geolocation': ''
            }

            res_login = self.sess.post(LOGIN_URL, data=login_data, headers=HEADERS)

            if self.sess.cookies.get('CASTGC'):
                return True
            else:
                print("⚠ 短信验证码错误")
                return False

        except Exception as e:
            print(f"✗ 登录过程出错: {e}")
            import traceback
            traceback.print_exc()
            return False


# ==================== 即席查询模块 ====================
class JXCXQuery:
    """即席查询类"""
    
    def __init__(self, session):
        self.sess = session
        self.enabled = False
        self._field_config_cache = {}
    
    def get_field_config(self, table_key, fieldtype, api_type='search'):
        """动态获取表字段配置（从API获取，而非硬编码）
        
        Args:
            table_key: API查询关键字
            fieldtype: 字段类型过滤条件
            api_type: API类型，'search'使用adhocquery/search接口，'table'使用adhocquery/getSelectTable接口
        """
        cache_key = f"{table_key}_{fieldtype}_{api_type}"
        if cache_key in self._field_config_cache:
            print(f"[DEBUG-FIELD] 使用缓存的字段配置: {cache_key}")
            return self._field_config_cache[cache_key]
        
        print(f"[DEBUG-FIELD] 动态获取字段配置: table_key={table_key}, fieldtype={fieldtype}, api_type={api_type}")
        
        try:
            if api_type == 'table':
                # 5G/4G KPI报表使用getSelectTable接口
                data = {'tablename': table_key}
                res = self.sess.post(JXCX_TABLE_URL, data=data, headers=HEADERS, timeout=30)
                
                if res.status_code == 200:
                    result = json.loads(res.content)
                    configs = result.get('CFG_ADHOC_CONF_TABLE', [])
                    print(f"[DEBUG-FIELD] 从getSelectTable接口获取到 {len(configs)} 个字段配置")
                    
                    self._field_config_cache[cache_key] = configs
                    return configs
                else:
                    print(f"[ERROR-FIELD] 获取字段配置失败: {res.status_code}")
                    return None
            else:
                # VoLTE/EPSFB/VONR使用search接口
                data = {
                    'key': table_key,
                    'field': 'columnname_cn',
                    'field': 'columnname',
                    'field': 'fieldtype',
                    'field': 'datatype',
                    'field': 'tablename',
                    'field': 'tablename_cn',
                    'field': 'columntype',
                    'field': 'sort'
                }
                res = self.sess.post(JXCX_SEARCH_URL, data=data, headers=HEADERS, timeout=30)
                
                if res.status_code == 200:
                    result = json.loads(res.content)
                    configs = result.get('CFG_ADHOC_CONF_SEARCH', [])
                    print(f"[DEBUG-FIELD] 从search接口获取到 {len(configs)} 个字段配置")
                    
                    self._field_config_cache[cache_key] = configs
                    return configs
                else:
                    print(f"[ERROR-FIELD] 获取字段配置失败: {res.status_code}")
                    return None
        except Exception as e:
            print(f"[ERROR-FIELD] 获取字段配置异常: {e}")
            return None
    
    def build_payload_from_config(self, table_key, fieldtype, where_conditions, api_type='search'):
        """从动态获取的字段配置构建payload
        
        Args:
            table_key: API查询关键字
            fieldtype: 字段类型过滤条件
            where_conditions: 查询条件列表
            api_type: API类型，'search'使用adhocquery/search接口，'table'使用adhocquery/getSelectTable接口
        """
        configs = self.get_field_config(table_key, fieldtype, api_type)
        if not configs:
            return None
        
        print(f"[DEBUG-PAYLOAD] API返回的字段配置数量: {len(configs)}")
        print(f"[DEBUG-PAYLOAD] API返回的字段名(前10个): {[c.get('columnname', '') for c in configs[:10]]}")
        print(f"[DEBUG-PAYLOAD] API返回的fieldtype(前3个): {list(set(c.get('fieldtype', '') for c in configs[:3]))}")
        
        # 按sort排序
        sorted_configs = sorted(configs, key=lambda x: x.get('sort', 0))
        
        # 从配置中获取维度参数（第一个配置项包含这些信息）
        first_config = sorted_configs[0]
        geographicdimension = first_config.get('geographicdimension', '小区')
        timedimension = first_config.get('timedimension', '天')
        enodeb_field = first_config.get('enodeb_field', 'enodeb_id')
        cgi_field = first_config.get('cgi_field', 'cgi')
        time_field = first_config.get('time_field', 'starttime')
        cell_field = first_config.get('cell_field', 'cell')
        city_field = first_config.get('city_field', 'city')
        
        print(f"[DEBUG-PAYLOAD] 从API获取维度参数:")
        print(f"  geographicdimension: {geographicdimension}")
        print(f"  timedimension: {timedimension}")
        print(f"  enodebField: {enodeb_field}")
        print(f"  cgiField: {cgi_field}")
        print(f"  timeField: {time_field}")
        print(f"  cellField: {cell_field}")
        print(f"  cityField: {city_field}")
        
        # 构建字段列表
        field_list = [c['columnname'] for c in sorted_configs]
        
        # 构建columns参数
        columns = []
        for field in field_list:
            columns.append({
                'data': field,
                'name': '',
                'searchable': True,
                'orderable': True,
                'search': {'value': '', 'regex': False}
            })
        
        # 构建result参数
        table_name = first_config.get('tablename', '')
        table_name_cn = first_config.get('tablename_cn', '')
        supporteddimension = first_config.get('supporteddimension')
        supportedtimedimension = first_config.get('supportedtimedimension', '')
        
        result_list = []
        for c in sorted_configs:
            result_list.append({
                'feildtype': c.get('fieldtype', ''),
                'table': c.get('tablename', ''),
                'tableName': c.get('tablename_cn', ''),
                'datatype': c.get('datatype', 'character varying'),
                'columntype': c.get('columntype', 1),
                'feildName': c.get('columnname_cn', ''),
                'feild': c.get('columnname', ''),
                'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'
            })
        
        result = {
            'result': result_list,
            'tableParams': {
                'supporteddimension': supporteddimension,
                'supportedtimedimension': supportedtimedimension
            },
            'columnname': ''
        }
        
        payload = {
            'draw': 1,
            'start': 0,
            'length': 200,
            'total': 0,
            'geographicdimension': geographicdimension,
            'timedimension': timedimension,
            'enodebField': enodeb_field,
            'cgiField': cgi_field,
            'timeField': time_field,
            'cellField': cell_field,
            'cityField': city_field,
            'columns': columns,
            'order': [{'column': 0, 'dir': 'desc'}],
            'search': {'value': '', 'regex': False},
            'result': result,
            'where': where_conditions,
            'indexcount': 0
        }
        
        print(f"[DEBUG-PAYLOAD] 构建的payload包含 {len(columns)} 个字段")
        return payload
    
    def enter_jxcx(self, retry_times=3, timeout=60):
        """进入即席查询模块"""
        print("\n[DEBUG-JXCX] ========== 进入即席查询模块 ==========")

        for attempt in range(retry_times):
            if attempt > 0:
                print(f"\n[DEBUG-JXCX] 重试第 {attempt} 次...")

            try:
                castgc = self.sess.cookies.get('CASTGC', domain='nqi.gmcc.net')
                if not castgc:
                    castgc = self.sess.cookies.get('CASTGC')

                if not castgc:
                    print("[ERROR-JXCX] 未找到CASTGC cookie")
                    continue

                print(f"[DEBUG-JXCX] CASTGC获取成功: {castgc[:20]}...")

                url = f'{BASE_URL}/pro-portal/pure/urlAction.action'
                params = {
                    'url': 'pro-adhoc/index',
                    'random': random.random(),
                    '__PID': 'JXCX',
                    'token': castgc
                }

                url_with_params = f"{url}?url={params['url']}&__PID={params['__PID']}&random={params['random']}&token={params['token']}"
                print(f"[DEBUG-JXCX] 请求URL: {url_with_params[:200]}...")

                start_time = time.time()
                res = self.sess.get(url_with_params, headers=HEADERS, timeout=timeout)
                elapsed_time = time.time() - start_time

                print(f"[DEBUG-JXCX] 响应状态码: {res.status_code}, 耗时: {elapsed_time:.2f}秒")

                if res.status_code == 200:
                    self.enabled = True
                    print("[SUCCESS-JXCX] 即席查询模块初始化成功！")
                    return True
                else:
                    print(f"[ERROR-JXCX] 进入即席查询失败，状态码: {res.status_code}")
                    continue

            except requests.exceptions.Timeout:
                print(f"[ERROR-JXCX] 请求超时 (timeout={timeout}s)")
                continue
            except requests.exceptions.ConnectionError as e:
                print(f"[ERROR-JXCX] 网络连接错误: {e}")
                continue
            except Exception as e:
                print(f"[ERROR-JXCX] 未知错误: {e}")
                continue

        print(f"[ERROR-JXCX] 进入即席查询失败，已尝试 {retry_times} 次")
        return False
    
    def get_table_count(self, payload):
        """获取查询结果行数"""
        if not self.enabled:
            self.enter_jxcx()

        key_list = ['geographicdimension', 'timedimension', 'enodebField', 'cgiField',
                    'timeField', 'cellField', 'cityField', 'result', 'where', 'indexcount',
                    'columns', 'order', 'search']
        payload_count = {key: value for key, value in payload.items() if key in key_list}
        payload_encoded = self._encode_payload(payload_count)

        print(f"[DEBUG-COUNT] 查询总数 URL: {JXCX_COUNT_URL}")
        print(f"[DEBUG-COUNT] 查询参数 (前500字符): {payload_encoded[:500]}...")

        try:
            res = self.sess.post(JXCX_COUNT_URL, data=payload_encoded, headers=HEADERS, timeout=180)
            print(f"[DEBUG-COUNT] 响应状态码: {res.status_code}")

            if res.status_code == 200:
                # 检查响应内容是否为空或无效
                if not res.content or len(res.content.strip()) == 0:
                    print(f"[ERROR-COUNT] 响应内容为空，可能是Session过期")
                    self.enabled = False  # 标记Session可能已过期
                    return 0

                try:
                    result = json.loads(res.content)
                except json.JSONDecodeError as e:
                    print(f"[ERROR-COUNT] JSON解析失败: {e}")
                    print(f"[ERROR-COUNT] 响应内容 (前500字符): {res.text[:500]}")
                    self.enabled = False  # 标记Session可能已过期
                    return 0

                print(f"[DEBUG-COUNT] 响应内容: {result}")

                # 检查是否有错误消息
                if 'message' in result and result['message']:
                    msg = str(result['message'])
                    print(f"[WARNING-COUNT] 服务器返回消息: {msg}")
                    if '不存在' in msg:
                        print(f"[WARNING-COUNT] 数据不存在，返回0")
                        return 0

                count = result.get('count', result.get('data', {}).get('total', 1000000))
                print(f"[DEBUG-COUNT] 查询到的数据行数: {count}")
                return count
        except Exception as e:
            print(f"[ERROR-COUNT] 查询异常: {e}")
            import traceback
            traceback.print_exc()
            pass
        # 超时时返回MAX_SINGLE_QUERY
        print(f"[WARNING-COUNT] 查询超时，返回MAX_SINGLE_QUERY({MAX_SINGLE_QUERY})")
        return MAX_SINGLE_QUERY

    
    def get_4g_voice_table(self, volte_payload, epsfb_payload, to_df=True):
        """获取4G语音小区报表数据（VoLTE+EPSFB联合，需分别查询后合并）

        Args:
            volte_payload: VoLTE表的payload
            epsfb_payload: EPSFB表的payload
            to_df: 是否返回DataFrame

        Returns:
            DataFrame或dict: 合并后的数据（向后兼容）
        """
        result = self._get_4g_voice_table_internal(volte_payload, epsfb_payload)
        if to_df:
            return result['merged']
        else:
            return {'data': result['merged'].to_dict('records')}

    def _get_4g_voice_table_internal(self, volte_payload, epsfb_payload):
        """获取4G语音小区报表数据（内部方法，返回原始数据和合并数据）

        Args:
            volte_payload: VoLTE表的payload
            epsfb_payload: EPSFB表的payload

        Returns:
            dict: {'volte': DataFrame, 'epsfb': DataFrame, 'merged': DataFrame}
        """
        result = {'volte': pd.DataFrame(), 'epsfb': pd.DataFrame(), 'merged': pd.DataFrame()}

        if not self.enabled:
            print("[DEBUG-4G-VOICE] JXCX 未启用，尝试进入...")
            if not self.enter_jxcx():
                print("[ERROR-4G-VOICE] 无法进入即席查询模块")
                return result

        print(f"[DEBUG-4G-VOICE] ========== 开始获取4G语音小区数据 ==========")

        # 1. 获取VoLTE数据
        print("[DEBUG-4G-VOICE] 正在获取VoLTE数据...")
        volte_df = self.get_table(volte_payload, to_df=True)
        print(f"[DEBUG-4G-VOICE] VoLTE数据: {len(volte_df)} 行")

        # 打印VoLTE列名
        print(f"[DEBUG-4G-VOICE] VoLTE列名: {list(volte_df.columns) if not volte_df.empty else 'N/A'}")

        result['volte'] = volte_df

        # 2. 获取EPSFB数据
        print("[DEBUG-4G-VOICE] 正在获取EPSFB数据...")
        epsfb_df = self.get_table(epsfb_payload, to_df=True)
        print(f"[DEBUG-4G-VOICE] EPSFB数据: {len(epsfb_df)} 行")

        # 打印EPSFB列名
        print(f"[DEBUG-4G-VOICE] EPSFB列名: {list(epsfb_df.columns) if not epsfb_df.empty else 'N/A'}")

        result['epsfb'] = epsfb_df

        if volte_df.empty and epsfb_df.empty:
            print("[WARNING-4G-VOICE] VoLTE和EPSFB数据均为空")
            return result

        # 3. 统一列名 - API返回的是英文列名(starttime, cgi)，转换为中文便于合并
        # 列名映射：英文 -> 中文
        col_name_map = {
            'starttime': '时间', 'city': '地市', 'cgi': '小区', 'grid': '责任网格',
            'area': '区县', 'nrcell_name': '小区名称'
        }

        # VoLTE列名转换
        volte_rename = {}
        for en_col in volte_df.columns:
            if en_col in col_name_map:
                volte_rename[en_col] = col_name_map[en_col]
        volte_df = volte_df.rename(columns=volte_rename)

        # EPSFB列名转换
        epsfb_rename = {}
        for en_col in epsfb_df.columns:
            if en_col in col_name_map:
                epsfb_rename[en_col] = col_name_map[en_col]
        epsfb_df = epsfb_df.rename(columns=epsfb_rename)

        print(f"[DEBUG-4G-VOICE] VoLTE转换后列名: {list(volte_df.columns)}")
        print(f"[DEBUG-4G-VOICE] EPSFB转换后列名: {list(epsfb_df.columns)}")

        # 4. 确定合并键
        merge_keys = []
        for key in ['时间', '小区']:
            if key in volte_df.columns:
                merge_keys.append(key)

        print(f"[DEBUG-4G-VOICE] 合并键: {merge_keys}")

        if not merge_keys:
            print("[WARNING-4G-VOICE] 无法确定合并键，使用简单concat")
            merged_df = pd.concat([volte_df, epsfb_df], ignore_index=True)
            result['merged'] = merged_df
            return result

        # 5. 确定VoLTE和EPSFB各自的特有字段
        common_cols = set(['时间', '小区', '地市', '责任网格', '区县', '小区名称', 'starttime', 'city', 'cgi', 'grid', 'area', 'nrcell_name'])

        # VoLTE特有字段（英文 volta_ 或中文 VoLTE 开头，但排除公共字段）
        volte_cols = [c for c in volte_df.columns if (c.startswith('volte_') or 'VoLTE' in c) and c not in common_cols]
        # EPSFB特有字段（英文 epsfb_ 或中文 EPSFB 开头，但排除公共字段）
        epsfb_cols = [c for c in epsfb_df.columns if (c.startswith('epsfb_') or 'EPSFB' in c) and c not in common_cols]

        print(f"[DEBUG-4G-VOICE] VoLTE特有字段: {volte_cols}")
        print(f"[DEBUG-4G-VOICE] EPSFB特有字段: {epsfb_cols}")

        # 6. 准备合并数据
        # VoLTE: 合并键 + VoLTE特有字段
        volte_merge_cols = [c for c in merge_keys if c in volte_df.columns] + [c for c in volte_cols if c in volte_df.columns]
        volte_for_merge = volte_df[volte_merge_cols].copy()

        # EPSFB: 合并键 + EPSFB特有字段
        epsfb_merge_cols = [c for c in merge_keys if c in epsfb_df.columns] + [c for c in epsfb_cols if c in epsfb_df.columns]
        epsfb_for_merge = epsfb_df[epsfb_merge_cols].copy()

        # 7. 执行合并（outer join，保留两边的数据）
        if not merge_keys:
            merged_df = pd.concat([volte_for_merge, epsfb_for_merge], axis=1)
        else:
            merged_df = pd.merge(volte_for_merge, epsfb_for_merge, on=merge_keys, how='outer')

        # 清理可能的空列
        merged_df = merged_df.dropna(axis=1, how='all')

        print(f"[SUCCESS-4G-VOICE] 合并完成，最终数据: {len(merged_df)} 行, {len(merged_df.columns)} 列")
        print(f"[DEBUG-4G-VOICE] 合并后列名: {list(merged_df.columns)}")

        result['merged'] = merged_df
        return result

    def _extract_cities(self, payload):
        """从payload中提取地市列表"""
        cities = []
        if 'where' in payload:
            for condition in payload['where']:
                if condition.get('feild', '').lower() == 'city':
                    val = condition.get('val', '')
                    # 支持逗号分隔的多地市，如 "广州,深圳,东莞"
                    cities = [c.strip() for c in str(val).split(',') if c.strip()]
                    break
        return cities

    def _set_single_city(self, payload, city):
        """设置payload为单个地市"""
        if 'where' in payload:
            for condition in payload['where']:
                if condition.get('feild', '').lower() == 'city':
                    condition['val'] = city
                    break
        return payload

    def _fetch_by_loop(self, payload, total_count):
        """循环获取数据，每批5000条，超过100万行才用分批策略"""
        import copy
        MAX_ROWS = 1000000
        if total_count <= MAX_ROWS:
            # 循环获取
            data_list = []
            start = 0
            page_size = 5000
            retry_count = 0
            max_retries = 3

            while start < total_count:
                # 检查 session 是否还有效
                castgc = self.sess.cookies.get('CASTGC', domain='nqi.gmcc.net')
                if not castgc:
                    castgc = self.sess.cookies.get('CASTGC')

                if not castgc and retry_count < max_retries:
                    retry_count += 1
                    print(f"[WARNING-TABLE] Session可能已失效，尝试重新进入即席查询 (重试 {retry_count}/{max_retries})...")
                    if self.enter_jxcx(retry_times=2, timeout=60):
                        print("[SUCCESS-TABLE] 重新进入即席查询成功！")
                        continue
                    else:
                        print("[ERROR-TABLE] 重新进入即席查询失败")
                        break

                p = copy.deepcopy(payload)
                p['start'] = start
                p['length'] = page_size
                print(f"[DEBUG-TABLE] 查询: start={start}, length={page_size}")
                batch = self._fetch_data(p)

                if not batch:
                    # 如果返回空数据且session可能失效，尝试重连
                    castgc = self.sess.cookies.get('CASTGC', domain='nqi.gmcc.net')
                    if not castgc:
                        castgc = self.sess.cookies.get('CASTGC')

                    if not castgc and retry_count < max_retries:
                        retry_count += 1
                        print(f"[WARNING-TABLE] 空数据+Session失效，尝试重新进入即席查询 (重试 {retry_count}/{max_retries})...")
                        if self.enter_jxcx(retry_times=2, timeout=60):
                            print("[SUCCESS-TABLE] 重新进入即席查询成功！")
                            continue
                        else:
                            print("[ERROR-TABLE] 重新进入即席查询失败")
                            break

                    print(f"[WARNING-TABLE] start={start} 返回空数据，停止查询")
                    break

                data_list.extend(batch)
                print(f"[DEBUG-TABLE] 已获取 {len(data_list)}/{total_count} 条")

                if len(batch) < page_size:
                    print(f"[DEBUG-TABLE] 返回数据少于请求数量，已到最后一页")
                    break

                start += page_size
                if start >= total_count:
                    break

            return data_list
        else:
            # 超过100万行：使用分批查询策略
            print(f"[DEBUG-TABLE] 数据量 {total_count} > {MAX_ROWS}，使用分批查询策略")
            return self._fetch_data_batch(payload, total_count)

    def get_table(self, payload, to_df=True):
        """获取查询数据（100万行以内直接循环获取，超过100万行才分批）"""
        import copy

        if not self.enabled:
            print("[DEBUG-TABLE] JXCX 未启用，尝试进入...")
            if not self.enter_jxcx():
                print("[ERROR-TABLE] 无法进入即席查询模块")
                return pd.DataFrame() if to_df else {}

        print(f"[DEBUG-TABLE] ========== 开始查询数据 ==========")
        print(f"[DEBUG-TABLE] 请求URL: {JXCX_URL}")

        # 打印表名
        if 'result' in payload and 'result' in payload['result']:
            table_names = [r.get('table', '') for r in payload['result']['result']]
            print(f"[DEBUG-TABLE] 查询表名: {set(table_names)}")

        # 打印 where 条件
        if 'where' in payload:
            print(f"[DEBUG-TABLE] 查询条件 (where): {json.dumps(payload['where'], ensure_ascii=False)}")

        # 验证 session 有效性
        castgc = self.sess.cookies.get('CASTGC', domain='nqi.gmcc.net')
        if not castgc:
            castgc = self.sess.cookies.get('CASTGC')
        print(f"[DEBUG-TABLE] CASTGC cookie 状态: {'有效' if castgc else '无效或过期'}")

        # 如果 session 失效，尝试重新进入即席查询
        if not castgc:
            print("[WARNING-TABLE] Session 可能已过期，尝试重新进入即席查询...")
            if not self.enter_jxcx(retry_times=2, timeout=60):
                print("[ERROR-TABLE] 重新进入即席查询失败，请尝试重新登录")
                return pd.DataFrame() if to_df else {}
        
        # 提取地市列表（用于判断是否需要分地市提取）
        city_list = self._extract_cities(payload)
        multiple_cities = len(city_list) > 1
        print(f"[DEBUG-TABLE] 检测到地市: {city_list} (共{len(city_list)}个，{'多地市' if multiple_cities else '单地市'})")

        # 多地市：分地市提取再合并
        if multiple_cities:
            print(f"[DEBUG-TABLE] 多地市模式：分地市提取再合并")
            all_data = []
            for city in city_list:
                print(f"[DEBUG-TABLE] ===== 开始提取地市: {city} =====")
                city_payload = self._set_single_city(copy.deepcopy(payload), city)
                city_count = self.get_table_count(copy.deepcopy(city_payload))
                print(f"[DEBUG-TABLE] {city} 数据量: {city_count}")

                city_data = self._fetch_by_loop(city_payload, city_count)
                all_data.extend(city_data)
                print(f"[DEBUG-TABLE] {city} 提取完成，已获取 {len(city_data)} 条")

            data_list = all_data
        else:
            # 单地市或无地市条件：直接获取
            total_count = self.get_table_count(copy.deepcopy(payload))
            print(f"[DEBUG-TABLE] 调用 get_table_count 获取数据总数: {total_count}")
            data_list = self._fetch_by_loop(payload, total_count)

        if to_df:
            en_zh_df = self._get_field_mapping(payload)
            res_df = pd.DataFrame(data_list)

            if res_df.empty:
                print(f"[WARNING-TABLE] DataFrame为空")
                return pd.DataFrame()

            print(f"[DEBUG-TABLE] 字段映射: {en_zh_df.to_dict()}")
            res_df = pd.concat([en_zh_df, res_df], ignore_index=True)
            index_first = res_df.index.tolist()[0]
            to_colname = list(res_df.loc[index_first])
            res_df.columns = to_colname
            res_df.drop(index=index_first, inplace=True)

            print(f"[SUCCESS-TABLE] 最终返回 DataFrame, shape: {res_df.shape}")
            return res_df
        else:
            return {'data': data_list}

    def _fetch_data(self, payload, timeout=None):
        """发送请求获取数据"""
        if timeout is None:
            timeout = getattr(self, '_current_batch_timeout', 300)
            
        payload_encoded = self._encode_payload(payload)
        print(f"[DEBUG-TABLE] 编码后的参数 (前800字符): {payload_encoded[:800]}...")
        print(f"[DEBUG-TABLE] start={payload.get('start', 0)}, length={payload.get('length', 'N/A')}")

        try:
            print(f"[DEBUG-TABLE] 开始发送请求 (timeout={timeout}s)...")
            start_request_time = time.time()
            res = self.sess.post(JXCX_URL, data=payload_encoded, headers=HEADERS, timeout=timeout)
            elapsed = time.time() - start_request_time
            print(f"[DEBUG-TABLE] 请求完成，耗时: {elapsed:.2f}秒，状态码: {res.status_code}")

            if res.status_code != 200:
                print(f"[ERROR-TABLE] 请求失败，状态码: {res.status_code}")
                print(f"[ERROR-TABLE] 响应内容: {res.text[:500]}")
                self.enabled = False  # 标记Session可能已过期
                return []

            # 检查响应内容是否为空
            if not res.content or len(res.content.strip()) == 0:
                print(f"[ERROR-TABLE] 响应内容为空，可能是Session过期或服务器错误")
                self.enabled = False
                return []

            try:
                result = json.loads(res.content)
            except json.JSONDecodeError as e:
                print(f"[ERROR-TABLE] JSON解析失败: {e}")
                print(f"[ERROR-TABLE] 响应内容 (前500字符): {res.text[:500]}")
                self.enabled = False  # 标记Session可能已过期
                return []

            print(f"[DEBUG-TABLE] 响应JSON keys: {result.keys()}")

            # 打印 recordsFiltered 和 recordsTotal（DataTables 分页信息）
            records_filtered = result.get('recordsFiltered', 'N/A')
            records_total = result.get('recordsTotal', 'N/A')
            print(f"[DEBUG-TABLE] recordsFiltered: {records_filtered}, recordsTotal: {records_total}")

            # 打印可能的错误信息
            if 'msg' in result:
                print(f"[DEBUG-TABLE] 响应消息: {result['msg']}")
            if 'message' in result and result['message']:
                print(f"[DEBUG-TABLE] 响应消息: {result['message']}")

            # 检查是否有错误
            if 'message' in result and result['message']:
                if '不存在' in str(result['message']):
                    print(f"[WARNING-TABLE] 服务器返回: 数据不存在")
                    return []

            data_list = result.get('data') or []
            print(f"[DEBUG-TABLE] 返回数据条数: {len(data_list)}")
            if data_list:
                print(f"[DEBUG-TABLE] 数据样例 (第一条): {data_list[0]}")
            else:
                # data 为空时，也打印完整响应以便排查问题
                print(f"[DEBUG-TABLE] data为空，打印完整响应以便分析: {json.dumps(result, ensure_ascii=False)[:3000]}")

            return data_list

        except Exception as e:
            print(f"[ERROR-TABLE] 请求异常: {e}")
            import traceback
            traceback.print_exc()
            return []

    def _fetch_data_batch(self, original_payload, total_count):
        """分批查询数据（支持自动调整批次大小）"""
        import copy
        
        # 尝试确定服务器支持的批次大小
        BATCH_SIZES = [50000, 10000, 5000, 2000, 1000, 500, 200]  # 从大到小尝试
        
        # 不同批次大小对应的超时时间（秒）
        BATCH_TIMEOUTS = {
            50000: 300,
            10000: 180,
            5000: 120,
            2000: 90,
            1000: 60,
            500: 45,
            200: 30
        }
        
        # 存储测试时的超时设置
        self._current_batch_timeout = 300
        all_data = []
        start = 0
        
        print(f"[DEBUG-BATCH] 开始分批查询，总数: {total_count}")
        
        # 先测试服务器支持的批次大小
        for test_size in BATCH_SIZES:
            test_timeout = BATCH_TIMEOUTS.get(test_size, 300)
            self._current_batch_timeout = test_timeout
            test_payload = copy.deepcopy(original_payload)
            test_payload['start'] = 0
            test_payload['length'] = test_size
            print(f"[DEBUG-BATCH] 测试批次大小 {test_size} (timeout={test_timeout}s)...")
            test_data = self._fetch_data(test_payload)
            
            if test_data:
                BATCH_SIZE = test_size
                self._current_batch_timeout = BATCH_TIMEOUTS.get(BATCH_SIZE, 300)
                all_data.extend(test_data)
                start = test_size
                print(f"[DEBUG-BATCH] 服务器支持批次大小: {BATCH_SIZE}，已获取 {len(all_data)} 条数据")
                break
            else:
                print(f"[DEBUG-BATCH] 批次大小 {test_size} 返回空数据，尝试更小的批次...")
        
        if not all_data:
            print(f"[WARNING-BATCH] 所有批次大小测试均返回空数据")
            return []
        
        # 继续分批获取剩余数据
        while start < total_count:
            payload = copy.deepcopy(original_payload)
            payload['start'] = start
            payload['length'] = BATCH_SIZE
            
            print(f"[DEBUG-BATCH] 查询批次: start={start}, length={BATCH_SIZE}")
            batch_data = self._fetch_data(payload)
            
            if not batch_data:
                print(f"[WARNING-BATCH] 批次 {start}~{start+BATCH_SIZE} 返回空数据，停止分批查询")
                break
            
            all_data.extend(batch_data)
            print(f"[DEBUG-BATCH] 已获取 {len(all_data)}/{total_count} 条数据")
            
            # 如果返回数据少于请求数量，说明已经是最后一批
            if len(batch_data) < BATCH_SIZE:
                print(f"[DEBUG-BATCH] 返回数据 {len(batch_data)} < 请求数量 {BATCH_SIZE}，视为最后一页")
                break
                
            start += BATCH_SIZE
            
            # 防止无限循环
            if start >= total_count:
                break
        
        print(f"[SUCCESS-BATCH] 分批查询完成，共获取 {len(all_data)} 条数据")
        return all_data
    
    def _encode_payload(self, payload):
        """编码payload为URL格式"""
        out_list = []
        for key in payload:
            # columns, order, search 使用DataTables标准的扁平URL编码格式（不是JSON）
            # result, where 使用JSON序列化
            if key == 'columns':
                # 调试：检查columns类型
                col_val = payload[key]
                if isinstance(col_val, str):
                    print(f"[DEBUG-ENCODE] columns是字符串（可能是JSON序列化后的），长度: {len(col_val)}")
                    # 如果columns是字符串（JSON序列化后的），直接使用
                    out_list.append(quote(key) + '=' + quote(col_val))
                    continue
                elif not isinstance(col_val, list):
                    print(f"[DEBUG-ENCODE] columns类型异常: {type(col_val)}")
                    out_list.append(quote(key) + '=' + quote(str(col_val)))
                    continue
                    
                # DataTables标准格式: columns[0][data]=field&columns[0][name]=&...
                col_parts = []
                for i, col in enumerate(payload[key]):
                    # 调试：检查col类型
                    if isinstance(col, str):
                        col_parts.append(f'columns[{i}]={quote(col)}')
                        continue
                    try:
                        for sub_key, sub_val in col.items():
                            if isinstance(sub_val, dict):
                                # 处理嵌套对象如 search: {'value': '', 'regex': False}
                                for ss_key, ss_val in sub_val.items():
                                    col_parts.append(f'columns[{i}][{sub_key}][{ss_key}]={quote(str(ss_val))}')
                            else:
                                col_parts.append(f'columns[{i}][{sub_key}]={quote(str(sub_val))}')
                    except AttributeError as e:
                        print(f"[DEBUG-ENCODE] columns[{i}]类型异常: {type(col)}, 值: {str(col)[:100]}")
                        continue
                out_list.append('&'.join(col_parts))
            elif key == 'order':
                # DataTables标准格式: order[0][column]=0&order[0][dir]=desc
                order_parts = []
                for i, ord_item in enumerate(payload[key]):
                    for sub_key, sub_val in ord_item.items():
                        order_parts.append(f'order[{i}][{sub_key}]={quote(str(sub_val))}')
                out_list.append('&'.join(order_parts))
            elif key == 'search':
                # DataTables标准格式: search[value]=&search[regex]=false
                search_parts = []
                for sub_key, sub_val in payload[key].items():
                    search_parts.append(f'search[{sub_key}]={quote(str(sub_val))}')
                out_list.append('&'.join(search_parts))
            elif key in ['result', 'where']:
                right = quote(json.dumps(payload[key]))
                out_list.append(quote(key) + '=' + right)
            elif type(payload[key]) is int:
                right = str(payload[key])
                out_list.append(quote(key) + '=' + right)
            else:
                right = quote(str(payload[key]))
                out_list.append(quote(key) + '=' + right)
        return '&'.join(out_list)
    
    def _get_field_mapping(self, payload):
        """获取字段中英文映射"""
        result_list = payload['result']['result']
        result_df = pd.DataFrame(result_list)
        zn = list(result_df['feildName'])
        en = list(result_df['feild'])
        en_zh_dict = dict(zip(en, zn))
        return pd.DataFrame([en_zh_dict])


# ==================== Payload模板 ====================
def get_5g_interference_payload():
    """获取5G干扰小区查询payload"""
    return {
        'draw': 1, 'start': 0, 'length': 200, 'total': 0,
        'geographicdimension': '小区', 'timedimension': '天',
        'enodebField': 'gnodeb_id', 'cgiField': 'cgi', 'timeField': 'starttime',
        'cellField': 'cell', 'cityField': 'city',
        'result': {'result': [
            {'feildtype': '5G干扰报表（忙时）', 'table': 'appdbv3.a_interfere_nr_cell_zb2_d',
             'tableName': '5G干扰报表（忙时）', 'datatype': 'character varying', 'columntype': '1',
             'feildName': '数据时间', 'feild': 'starttime', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G干扰报表（忙时）', 'table': 'appdbv3.a_interfere_nr_cell_zb2_d',
             'tableName': '5G干扰报表（忙时）', 'datatype': 'character varying', 'columntype': '1',
             'feildName': '结束时间', 'feild': 'endtime', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G干扰报表（忙时）', 'table': 'appdbv3.a_interfere_nr_cell_zb2_d',
             'tableName': '5G干扰报表（忙时）', 'datatype': 'character varying', 'columntype': '1',
             'feildName': 'CGI', 'feild': 'cgi', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G干扰报表（忙时）', 'table': 'appdbv3.a_interfere_nr_cell_zb2_d',
             'tableName': '5G干扰报表（忙时）', 'datatype': 'character varying', 'columntype': '1',
             'feildName': '小区名', 'feild': 'cell_name', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G干扰报表（忙时）', 'table': 'appdbv3.a_interfere_nr_cell_zb2_d',
             'tableName': '5G干扰报表（忙时）', 'datatype': 'character varying', 'columntype': '1',
             'feildName': '频段', 'feild': 'freq', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G干扰报表（忙时）', 'table': 'appdbv3.a_interfere_nr_cell_zb2_d',
             'tableName': '5G干扰报表（忙时）', 'datatype': 'character varying', 'columntype': '1',
             'feildName': '微网格标识', 'feild': 'micro_grid', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G干扰报表（忙时）', 'table': 'appdbv3.a_interfere_nr_cell_zb2_d',
             'tableName': '5G干扰报表（忙时）', 'datatype': 'character varying', 'columntype': '1',
             'feildName': '全频段均值', 'feild': 'averagevalue', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G干扰报表（忙时）', 'table': 'appdbv3.a_interfere_nr_cell_zb2_d',
             'tableName': '5G干扰报表（忙时）', 'datatype': 'character varying', 'columntype': '1',
             'feildName': 'D1均值', 'feild': 'averagevalued1', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G干扰报表（忙时）', 'table': 'appdbv3.a_interfere_nr_cell_zb2_d',
             'tableName': '5G干扰报表（忙时）', 'datatype': 'character varying', 'columntype': '1',
             'feildName': 'D2均值', 'feild': 'averagevalued2', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G干扰报表（忙时）', 'table': 'appdbv3.a_interfere_nr_cell_zb2_d',
             'tableName': '5G干扰报表（忙时）', 'datatype': 'character varying', 'columntype': '1',
             'feildName': '是否干扰小区', 'feild': 'is_interfere_5g', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'}
        ], 'tableParams': {'supporteddimension': None, 'supportedtimedimension': ''}, 'columnname': ''},
        'where': [
            {'datatype': 'timestamp', 'feild': 'starttime', 'feildName': '', 'symbol': '>=',
             'val': '2025-01-13 00:00:00', 'whereCon': 'and', 'query': True},
            {'datatype': 'timestamp', 'feild': 'starttime', 'feildName': '', 'symbol': '<',
             'val': '2025-01-19 23:59:59', 'whereCon': 'and', 'query': True},
            {'datatype': 'character', 'feild': 'city', 'feildName': '', 'symbol': 'in',
             'val': '阳江', 'whereCon': 'and', 'query': True}
        ],
        'indexcount': 0
    }


def get_4g_interference_payload():
    """获取4G干扰小区查询payload"""
    return {
        'draw': 1, 'start': 0, 'length': 200, 'total': 0,
        'geographicdimension': '小区', 'timedimension': '天',
        'enodebField': 'enodeb_id', 'cgiField': 'cgi', 'timeField': 'starttime',
        'cellField': 'cell', 'cityField': 'city',
        'result': {'result': [
            {'feildtype': '4G干扰报表（忙时）', 'table': 'appdbv3.a_interfere_lte_cell_zb2_d',
             'tableName': '4G干扰报表（忙时）', 'datatype': '1', 'feildName': '数据时间',
             'feild': 'starttime', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '4G干扰报表（忙时）', 'table': 'appdbv3.a_interfere_lte_cell_zb2_d',
             'tableName': '4G干扰报表（忙时）', 'datatype': '1', 'feildName': '结束时间',
             'feild': 'endtime', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '4G干扰报表（忙时）', 'table': 'appdbv3.a_interfere_lte_cell_zb2_d',
             'tableName': '4G干扰报表（忙时）', 'datatype': '1', 'feildName': 'CGI',
             'feild': 'cgi', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '4G干扰报表（忙时）', 'table': 'appdbv3.a_interfere_lte_cell_zb2_d',
             'tableName': '4G干扰报表（忙时）', 'datatype': 'character varying', 'columntype': '1',
             'feildName': '小区名', 'feild': 'cell_name', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '4G干扰报表（忙时）', 'table': 'appdbv3.a_interfere_lte_cell_zb2_d',
             'tableName': '4G干扰报表（忙时）', 'datatype': 'character varying', 'columntype': '1',
             'feildName': '频段', 'feild': 'freq', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '4G干扰报表（忙时）', 'table': 'appdbv3.a_interfere_lte_cell_zb2_d',
             'tableName': '4G干扰报表（忙时）', 'datatype': 'character varying', 'columntype': '1',
             'feildName': '微网格标识', 'feild': 'micro_grid', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '4G干扰报表（忙时）', 'table': 'appdbv3.a_interfere_lte_cell_zb2_d',
             'tableName': '4G干扰报表（忙时）', 'datatype': 'character varying', 'columntype': '1',
             'feildName': '系统带宽', 'feild': 'bandwidth', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '4G干扰报表（忙时）', 'table': 'appdbv3.a_interfere_lte_cell_zb2_d',
             'tableName': '4G干扰报表（忙时）', 'datatype': 'character varying', 'columntype': '1',
             'feildName': '平均干扰电平', 'feild': 'averagevalue', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '4G干扰报表（忙时）', 'table': 'appdbv3.a_interfere_lte_cell_zb2_d',
             'tableName': '4G干扰报表（忙时）', 'datatype': 'character varying', 'columntype': '1',
             'feildName': '是否干扰小区', 'feild': 'is_interfere', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'}
        ], 'tableParams': {'supporteddimension': None, 'supportedtimedimension': ''}, 'columnname': ''},
        'where': [
            {'datatype': 'timestamp', 'feild': 'starttime', 'feildName': '', 'symbol': '>=',
             'val': '2025-01-13 00:00:00', 'whereCon': 'and', 'query': True},
            {'datatype': 'timestamp', 'feild': 'starttime', 'feildName': '', 'symbol': '<',
             'val': '2025-01-19 23:59:59', 'whereCon': 'and', 'query': True},
            {'datatype': 'character', 'feild': 'city', 'feildName': '', 'symbol': 'in',
             'val': '阳江', 'whereCon': 'and', 'query': True}
        ],
        'indexcount': 0
    }


def get_5g_capacity_payload():
    """获取5G小区容量报表payload - 基于浏览器HAR抓包的实际请求"""
    print("[DEBUG-PAYLOAD] 生成 5G小区容量报表 payload")
    payload = {
        'draw': 1, 'start': 0, 'length': 200, 'total': 0,
        'geographicdimension': '小区', 'timedimension': '天',
        'enodebField': 'enodeb_id', 'cgiField': 'ncgi',
        'timeField': 'starttime', 'cellField': 'nrcell', 'cityField': 'city',
        'result': {'result': [
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': '2', 'feildName': '记录开始时间',
             'feild': 'starttime', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': '2', 'feildName': '记录结束时间',
             'feild': 'endtime', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': '1', 'feildName': '地市',
             'feild': 'city', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': '1', 'feildName': '责任网格',
             'feild': 'grid', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'character varying', 'columntype': '2',
             'feildName': '小区名称', 'feild': 'nrcell_name', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'character varying', 'columntype': '2',
             'feildName': 'NCGI', 'feild': 'ncgi', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'timestamp', 'columntype': '1',
             'feildName': '忙时', 'feild': 'busy_hour', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'character varying', 'columntype': '1',
             'feildName': '网元状态', 'feild': 'state', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'character varying', 'columntype': '1',
             'feildName': '厂家', 'feild': 'vendor', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'character varying', 'columntype': '1',
             'feildName': '详细使用频段', 'feild': 'frequency_band_detail', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'character varying', 'columntype': '1',
             'feildName': '使用频段', 'feild': 'freq', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'character varying', 'columntype': '1',
             'feildName': '覆盖类型', 'feild': 'cover_type', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'character varying', 'columntype': '1',
             'feildName': '覆盖场景', 'feild': 'cover_scene', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'boolean', 'columntype': '1',
             'feildName': '是否拉远', 'feild': 'is_remote', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'character varying', 'columntype': '1',
             'feildName': '所属基站', 'feild': 'station_name', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'decimal', 'columntype': '1',
             'feildName': '经度', 'feild': 'longitude', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'decimal', 'columntype': '1',
             'feildName': '纬度', 'feild': 'latitude', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'bigint', 'columntype': '1',
             'feildName': '忙时PDCCH信道CCE占用个数', 'feild': 'bh_rru_pdcchcceutil', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'character varying', 'columntype': '1',
             'feildName': '共站同覆盖小区名称', 'feild': 'sectors_name', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'bigint', 'columntype': '1',
             'feildName': '忙时PDCCH信道CCE可用个数', 'feild': 'bh_rru_pdcchcceavail', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'decimal', 'columntype': '1',
             'feildName': '忙时PDCCH信道CCE占用率(%)', 'feild': 'bh_pdcchcceoccupancyrate', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'decimal', 'columntype': '1',
             'feildName': '忙时上行PUSCH PRB占用个数', 'feild': 'bh_rru_puschprbassn', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'decimal', 'columntype': '1',
             'feildName': '忙时上行PUSCH PRB可用个数', 'feild': 'bh_rru_puschprbtot', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'decimal', 'columntype': '1',
             'feildName': '忙时上行PRB平均利用率(%)', 'feild': 'bh_prbassnrateul', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'decimal', 'columntype': '1',
             'feildName': '忙时下行PDSCH PRB占用个数', 'feild': 'bh_rru_pdschprbassn', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'decimal', 'columntype': '1',
             'feildName': '忙时下行PDSCH PRB可用个数', 'feild': 'bh_rru_pdschprbtot', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'decimal', 'columntype': '1',
             'feildName': '忙时下行PRB平均利用率(%)', 'feild': 'bh_prbassnratedl', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'decimal', 'columntype': '1',
             'feildName': '忙时RLC层上行业务字节数(G)', 'feild': 'bh_rlc_upoctul', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'decimal', 'columntype': '1',
             'feildName': '忙时RLC层下行业务字节数(G)', 'feild': 'bh_rlc_upoctdl', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'decimal', 'columntype': '1',
             'feildName': '忙时MAC层上行业务流量(G)', 'feild': 'bh_mac_cpoctul', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'decimal', 'columntype': '1',
             'feildName': '忙时MAC层下行业务流量(G)', 'feild': 'bh_mac_cpoctdl', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'decimal', 'columntype': '1',
             'feildName': '忙时PDCP上行业务字节数(G)', 'feild': 'bh_pdcp_upoctul', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'decimal', 'columntype': '1',
             'feildName': '忙时PDCP下行业务字节数(G)', 'feild': 'bh_pdcp_upoctdl', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'decimal', 'columntype': '1',
             'feildName': '日RLC层上行业务字节数(G)', 'feild': 'rlc_upoctul', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'decimal', 'columntype': '1',
             'feildName': '日RLC层下行业务字节数(G)', 'feild': 'rlc_upoctdl', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'decimal', 'columntype': '1',
             'feildName': '日RLC层上下行总流量(G)', 'feild': 'rlc_upoctudl', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'decimal', 'columntype': '1',
             'feildName': '日MAC层上下行总流量(G)', 'feild': 'mac_cpoctudl', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'decimal', 'columntype': '1',
             'feildName': '日PDCP层上下行总流量(G)', 'feild': 'pdcp_upoctudl', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'character varying', 'columntype': '2',
             'feildName': 'RRC连接平均数-忙时', 'feild': 'bh_rrc_connmean', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'boolean', 'columntype': '1',
             'feildName': '是否高负荷待扩容小区', 'feild': 'is_highload', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'character varying', 'columntype': '2',
             'feildName': 'RRC连接最大数-忙时', 'feild': 'bh_rrc_connmax', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'character varying', 'columntype': '2',
             'feildName': 'Flow建立请求数-忙时', 'feild': 'bh_flow_nbrattestab', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'character varying', 'columntype': '2',
             'feildName': 'Flow建立成功数-忙时', 'feild': 'bh_flow_nbrsuccestab', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'character varying', 'columntype': '2',
             'feildName': 'QoS Flow建立成功率-忙时', 'feild': 'bh_kpi_flowsuccconnrate', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'character varying', 'columntype': '1',
             'feildName': '路测网格', 'feild': 'grid_road', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'character varying', 'columntype': '1',
             'feildName': '二级场景细化', 'feild': 'second_scene_detail', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'character varying', 'columntype': '1',
             'feildName': '忙时上行PRB可用空分层数', 'feild': 'bh_puschprbtot_reuse', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'character varying', 'columntype': '1',
             'feildName': '忙时上行PRB占用空分层数', 'feild': 'bh_puschprbassn_reuse', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'character varying', 'columntype': '1',
             'feildName': '忙时上行业务信道平均空分层数', 'feild': 'bh_avgdtchmimolayerul', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'character varying', 'columntype': '1',
             'feildName': '忙时下行PRB可用空分层数', 'feild': 'bh_pdschprbtot_reuse', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'character varying', 'columntype': '1',
             'feildName': '忙时下行PRB占用空分层数', 'feild': 'bh_pdschprbassn_reuse', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'character varying', 'columntype': '1',
             'feildName': '忙时下行业务信道平均空分层数', 'feild': 'bh_avgdtchmimolayerdl', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'character varying', 'columntype': '1',
             'feildName': '上行业务信道最大空分层数', 'feild': 'maxdtchmimolayerul', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'character varying', 'columntype': '1',
             'feildName': '下行业务信道最大空分层数', 'feild': 'maxdtchmimolayerdl', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'character varying', 'columntype': '1',
             'feildName': '忙时上行业务信道空分PRB占用率', 'feild': 'bh_dtchmimoprbassnrateul', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'character varying', 'columntype': '1',
             'feildName': '忙时下行业务信道空分PRB占用率', 'feild': 'bh_dtchmimoprbassnratedl', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'character varying', 'columntype': '1',
             'feildName': '忙时小区PRB利用率', 'feild': 'bh_cellprbrate', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'character varying', 'columntype': '1',
             'feildName': '忙时切换进入Flow数', 'feild': 'bh_flow_nbrhoinc', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'character varying', 'columntype': '1',
             'feildName': '忙时每Flow流量', 'feild': 'bh_upoctudl_perflow', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'character varying', 'columntype': '1',
             'feildName': '小区带宽', 'feild': 'band_width', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'character varying', 'columntype': '1',
             'feildName': '中心载频号', 'feild': 'ssbfrequenc', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'character varying', 'columntype': '1',
             'feildName': '射频通道数', 'feild': 'txrxmode', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'character varying', 'columntype': '1',
             'feildName': '区域类型', 'feild': 'area_type', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'character varying', 'columntype': '1',
             'feildName': '微网格标识', 'feild': 'micro_grid', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'character varying', 'columntype': '1',
             'feildName': '覆盖场景1', 'feild': 'cover_scene1', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'character varying', 'columntype': '1',
             'feildName': '覆盖场景2', 'feild': 'cover_scene2', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'character varying', 'columntype': '1',
             'feildName': '覆盖场景3', 'feild': 'cover_scene3', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '5G小区容量报表 - 天粒度', 'table': 'appdbv3.a_adhoc_capacity_nr_nrcell_d',
             'tableName': '5G小区容量报表 - 天粒度', 'datatype': 'character varying', 'columntype': '1',
             'feildName': '覆盖场景4', 'feild': 'cover_scene4', 'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
        ], 'tableParams': {'supporteddimension': None, 'supportedtimedimension': ''}, 'columnname': ''},
        'where': [
            {'datatype': 'timestamp', 'feild': 'starttime', 'feildName': '', 'symbol': '>=',
             'val': '2026-04-19 00:00:00', 'whereCon': 'and', 'query': True},
            {'datatype': 'timestamp', 'feild': 'starttime', 'feildName': '', 'symbol': '<',
             'val': '2026-04-19 23:59:59', 'whereCon': 'and', 'query': True},
            {'datatype': 'character', 'feild': 'city', 'feildName': '', 'symbol': 'in',
             'val': '阳江', 'whereCon': 'and', 'query': True}
        ],
        'indexcount': 1
    }
    print(f"[DEBUG-PAYLOAD] 5G小区容量报表 payload 生成完成，表名: appdbv3.a_adhoc_capacity_nr_nrcell_d，字段数: {len(payload['result']['result'])}")
    return payload


def get_important_scene_payload():
    """获取重要场景-天报表payload - 基于浏览器HAR抓包的实际请求"""
    print("[DEBUG-PAYLOAD] 生成 重要场景-天 payload")
    payload = {
        'draw': 1, 'start': 0, 'length': 200, 'total': 0,
        'geographicdimension': '小区', 'timedimension': '天',
        'enodebField': 'enodeb_id', 'cgiField': 'cgi',
        'timeField': 'starttime', 'cellField': 'cell', 'cityField': 'city',
        'result': {'result': [
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': '1',
             'feildName': '记录开始时间', 'feild': 'starttime',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': '1',
             'feildName': '记录结束时间', 'feild': 'endtime',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': '1',
             'feildName': '所属地市', 'feild': 'city',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': '1',
             'feildName': 'CGI', 'feild': 'cgi',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'timestamp', 'columntype': 1,
             'feildName': '自忙时', 'feild': 'busy_hour',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'character varying', 'columntype': 1,
             'feildName': '详细频段', 'feild': 'freq_name',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'character varying', 'columntype': 1,
             'feildName': '小区名称', 'feild': 'cell_name',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'character varying', 'columntype': 1,
             'feildName': '网元状态', 'feild': 'state',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'character varying', 'columntype': 1,
             'feildName': '场景', 'feild': 'scene',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'character varying', 'columntype': 1,
             'feildName': '场景具体名称', 'feild': 'scene_name',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'decimal', 'columntype': 1,
             'feildName': '日4G流量（GB）', 'feild': 'upoctudl',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'decimal', 'columntype': 1,
             'feildName': '日峰值上行PRB平均利用率', 'feild': 'ul_prbuse_rate_max',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'decimal', 'columntype': 1,
             'feildName': '日峰值下行PRB平均利用率', 'feild': 'dl_prbuse_rate_max',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'decimal', 'columntype': 1,
             'feildName': '日峰值PDCCH信道CCE占用率', 'feild': 'pdcchcceutilratio_max',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'integer', 'columntype': 1,
             'feildName': '自忙时有效RRC连接最大数', 'feild': 'bh_effectiveconnmax',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'integer', 'columntype': 1,
             'feildName': '自忙时RRC连接最大数', 'feild': 'bh_connmax',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'decimal', 'columntype': 1,
             'feildName': '自忙时有效RRC连接平均数', 'feild': 'bh_effectiveconnmean',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'decimal', 'columntype': 1,
             'feildName': '自忙时空口上行业务字节数', 'feild': 'bh_upoctul',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'decimal', 'columntype': 1,
             'feildName': '自忙时空口下行业务字节数', 'feild': 'bh_upoctdl',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'decimal', 'columntype': 1,
             'feildName': '自忙时上行PRB平均利用率', 'feild': 'bh_ul_prbuse_rate',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'decimal', 'columntype': 1,
             'feildName': '自忙时下行PRB平均利用率', 'feild': 'bh_dl_prbuse_rate',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'decimal', 'columntype': 1,
             'feildName': '自忙时PDCCH信道CCE占用率', 'feild': 'bh_pdcchcceutilratio',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'decimal', 'columntype': 1,
             'feildName': '无线接通率', 'feild': 'radio_succ_rate',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'decimal', 'columntype': 1,
             'feildName': '无线掉线率(小区级)', 'feild': 'radio_drop_rate_cell',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'decimal', 'columntype': 1,
             'feildName': '呼叫接通率(MTC+MOC)', 'feild': 'call_connect_rate',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'decimal', 'columntype': 1,
             'feildName': 'VOLTE掉话率', 'feild': 'volte_drop_rate',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'decimal', 'columntype': 1,
             'feildName': 'ESRVCC切换成功率', 'feild': 'esrvcc_ho_succ_rate',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'decimal', 'columntype': 1,
             'feildName': 'VOLTE语音话务量', 'feild': 'volte_voice_traffic',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'decimal', 'columntype': 1,
             'feildName': '小区自忙时平均E-RAB流量', 'feild': 'bh_avg_erab',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'decimal', 'columntype': 1,
             'feildName': '自忙时峰值利用率', 'feild': 'bh_peak_use_rate',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'integer', 'columntype': 1,
             'feildName': '自忙时E-RAB建立成功数', 'feild': 'bh_nbrsuccestab',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'character varying', 'columntype': 1,
             'feildName': '是否高流量预警小区（集团高负荷预警）', 'feild': 'is_highflow',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'character varying', 'columntype': 1,
             'feildName': '是否高负荷待扩容小区', 'feild': 'is_highload',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'integer', 'columntype': 1,
             'feildName': '自忙时PDCCH信道CCE占用个数', 'feild': 'bh_pdcchcceutil',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'integer', 'columntype': 1,
             'feildName': '自忙时PDCCH信道CCE可用个数', 'feild': 'bh_pdcchcceavail',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'integer', 'columntype': 1,
             'feildName': '自忙时下行PDSCH_PRB占用数', 'feild': 'bh_pdschprbassn',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'integer', 'columntype': 1,
             'feildName': '自忙时下行PDSCH_PRB可用数', 'feild': 'bh_pdschprbtot',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'integer', 'columntype': 1,
             'feildName': '自忙时上行PUSCH_PRB占用数', 'feild': 'bh_puschprbassn',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'integer', 'columntype': 1,
             'feildName': '自忙时上行PUSCH_PRB可用数', 'feild': 'bh_puschprbtot',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'integer', 'columntype': 1,
             'feildName': '日峰值PDCCH信道CCE占用个数', 'feild': 'pdcchcceutil_max',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'integer', 'columntype': 1,
             'feildName': '日峰值PDCCH信道CCE可用个数', 'feild': 'pdcchcceavail_max',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'integer', 'columntype': 1,
             'feildName': '日峰值下行PDSCH_PRB占用数', 'feild': 'pdschprbassn_max',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'integer', 'columntype': 1,
             'feildName': '日峰值下行PDSCH_PRB可用数', 'feild': 'pdschprbtot_max',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'integer', 'columntype': 1,
             'feildName': '日峰值上行PUSCH_PRB占用数', 'feild': 'puschprbassn_max',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'integer', 'columntype': 1,
             'feildName': '日峰值上行PUSCH_PRB可用数', 'feild': 'puschprbtot_max',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'character varying', 'columntype': 1,
             'feildName': '设备厂家', 'feild': 'vendor',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'character varying', 'columntype': 1,
             'feildName': '使用频段', 'feild': 'freq',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'decimal', 'columntype': 1,
             'feildName': '流量系数', 'feild': 'flow_coefficient',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'character varying', 'columntype': 1,
             'feildName': '小区带宽', 'feild': 'bandwidth',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'character varying', 'columntype': 1,
             'feildName': '是否拉远', 'feild': 'is_remote',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'character varying', 'columntype': 1,
             'feildName': '所属站点名称', 'feild': 'station_name',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'decimal', 'columntype': 1,
             'feildName': 'TCP二三次握手时延（ms)', 'feild': 'lte_soc_tcpsetup_c_007',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'decimal', 'columntype': 1,
             'feildName': '大包速率(>500KB)', 'feild': 'lte_soc_http_cell_c_055',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'decimal', 'columntype': 1,
             'feildName': 'TCP二三次握手成功率', 'feild': 'lte_soc_tcpsetup_c_003',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'character varying', 'columntype': 1,
             'feildName': '是否高流量感知问题小区', 'feild': 'is_highflow_perceive',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'decimal', 'columntype': 1,
             'feildName': '等效TDD_20M载波数', 'feild': 'equivalent_20m_carrier',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'integer', 'columntype': 1,
             'feildName': '三次握手成功次数(HTTP)', 'feild': 'lte_soc_tcpsetup_021',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'integer', 'columntype': 1,
             'feildName': '三次握手成功次数(S1U)', 'feild': 'lte_soc_tcpsetup_023',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'integer', 'columntype': 1,
             'feildName': 'TCP一、二次握手成功次数(HTTP)', 'feild': 'lte_soc_tcpsetup_027',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'integer', 'columntype': 1,
             'feildName': 'TCP一、二次握手成功次数(S1U)', 'feild': 'lte_soc_tcpsetup_029',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'decimal', 'columntype': 1,
             'feildName': 'TCP建立总延时（HTTP)', 'feild': 'lte_soc_tcpsetup_030',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'decimal', 'columntype': 1,
             'feildName': 'TCP建立总延时(S1U)', 'feild': 'lte_soc_tcpsetup_032',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'decimal', 'columntype': 1,
             'feildName': 'TCP一、二次握手总延时(HTTP)', 'feild': 'lte_soc_tcpsetup_033',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'decimal', 'columntype': 1,
             'feildName': 'TCP一、二次握手总延时(S1U)', 'feild': 'lte_soc_tcpsetup_035',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'decimal', 'columntype': 1,
             'feildName': '大包流量(>500KB)', 'feild': 'lte_soc_http_cell_078',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'decimal', 'columntype': 1,
             'feildName': '大包总时延(>500KB)', 'feild': 'lte_soc_http_cell_079',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'character varying', 'columntype': 1,
             'feildName': '所属共站同覆盖区域编号', 'feild': 'sectors_no',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'character varying', 'columntype': 1,
             'feildName': '所属共站同覆盖区域名', 'feild': 'sectors_name',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'integer', 'columntype': 1,
             'feildName': '小区自忙时5M感知需求能力-用户数', 'feild': 'aim_user',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'decimal', 'columntype': 1,
             'feildName': '小区自忙时5M感知需求能力-流量（GB）', 'feild': 'aim_flow',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'decimal', 'columntype': 1,
             'feildName': '小区自忙时5M感知需求能力-利用率（%）', 'feild': 'aim_utilization',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'character varying', 'columntype': 1,
             'feildName': '覆盖类型', 'feild': 'cover',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'character varying', 'columntype': 1,
             'feildName': '覆盖层标识', 'feild': 'coverlayer',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'character varying', 'columntype': 1,
             'feildName': '容量层标识', 'feild': 'capacitylayer',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'character varying', 'columntype': 1,
             'feildName': '扇区宽度', 'feild': 'sectors_width',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'character varying', 'columntype': 1,
             'feildName': '小区所属区域', 'feild': 'cell_scene_name',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'character varying', 'columntype': 1,
             'feildName': '经度', 'feild': 'longitude',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'character varying', 'columntype': 1,
             'feildName': '维度', 'feild': 'latitude',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'character varying', 'columntype': 1,
             'feildName': '天线通道数', 'feild': 'channel_numbers',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'character varying', 'columntype': 1,
             'feildName': '是否扩容小区+', 'feild': 'is_dilatation',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'character varying', 'columntype': 1,
             'feildName': '是否纳入载波调度', 'feild': 'is_carry',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'character varying', 'columntype': 1,
             'feildName': '网络结构属性', 'feild': 'network_structure',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'character varying', 'columntype': 1,
             'feildName': '小区所属区域类型', 'feild': 'cell_scene_type',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'character varying', 'columntype': 1,
             'feildName': '是否本地市尾部小区', 'feild': 'is_city_tail',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
            {'feildtype': '重要场景-小区天', 'table': 'appdbv3.a_overview_ispm_lte_cell_d',
             'tableName': '[管理视图]重要场景-小区天粒度', 'datatype': 'character varying', 'columntype': 1,
             'feildName': '是否全省尾部小区', 'feild': 'is_province_tail',
             'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'},
        ], 'tableParams': {'supporteddimension': None, 'supportedtimedimension': ''}, 'columnname': ''},
        'where': [
            {'datatype': 'timestamp', 'feild': 'starttime', 'feildName': '', 'symbol': '>=',
             'val': '2026-04-19 00:00:00', 'whereCon': 'and', 'query': True},
            {'datatype': 'timestamp', 'feild': 'starttime', 'feildName': '', 'symbol': '<',
             'val': '2026-04-19 23:59:59', 'whereCon': 'and', 'query': True},
            {'datatype': 'character', 'feild': 'city', 'feildName': '', 'symbol': 'in',
             'val': '阳江', 'whereCon': 'and', 'query': True}
        ],
        'indexcount': 0
    }
    print(f"[DEBUG-PAYLOAD] 重要场景-天 payload 生成完成，表名: appdbv3.a_overview_ispm_lte_cell_d")
    return payload


def _build_columns_param(field_list):
    """构建DataTables格式的columns参数"""
    columns = []
    for field in field_list:
        columns.append({
            'data': field,
            'name': '',
            'searchable': True,
            'orderable': True,
            'search': {'value': '', 'regex': False}
        })
    return columns


def get_volte_warning_payload():
    """获取VoLTE小区监控预警payload"""
    print("[DEBUG-PAYLOAD] 生成 VoLTE小区监控预警 payload")
    
    # VoLTE表完整字段列表（89个字段）
    volte_fields = [
        'starttime', 'city', 'cgi', 'grid', 'area', 'nrcell_name',
        'cs_reg1_suss_rate', 'cs_reg_suss_rate', 'cs_reg_sbc_suss_rate',
        'cs_moc_sbc_suss_rate', 'cs_moc_sbc_180_suss_rate', 'cs_mtc_sbc_suss_rate', 'cs_mtc_sbc_180_suss_rate',
        'cs_sbc_suss_rate', 'cs_sbc_180_suss_rate', 'cs_moc_sbc_net_suss_rate', 'cs_sbc_net_suss_rate',
        'cs_sbc_drops_rate', 'cs_xsrvcc_ho_suss_rate', 'cs_ho_len_avg', 'cs_ho_rtp_delay_avg',
        'cs_alert_delay_vf_avg', 'cs_alert_delay_vv_avg', 'cs_alert_delay_vall_rate', 'cs_mt_alert_delay_avg',
        'mconv_rtp_ul_mos_avg', 'mconv_rtp_dl_mos_avg', 'mconv_mos_300_nok_rate',
        'mconv_rtcp_ul_mos_avg', 'mconv_rtcp_dl_mos0', 'mconv_rtp_ul_pkts_lost_rate',
        'mconv_rtp_dl_pkts_lost_rate', 'mconv_rtcp_ul_pkts_lost_rate', 'mconv_rtcp_dl_pkts_lost_rate',
        'mconv_single_voice_call_rate', 'mconv_dx_call_rate', 'mconv_rtp_ul_delay_avg', 'mconv_rtp_dl_delay_avg',
        'mconv_rtcp_ul_delay_avg', 'mconv_rtcp_dl_delay_avg', 'mconv_dl_mos_300_nok_rate',
        'msli_ul_tunzi_len_rate', 'msli_ul_duanxu_len_rate', 'msli_ul_dantong_len_rate', 'msli_ul_mos_poor_len_rate',
        'msli_dl_tunzi_len_rate', 'msli_dl_duanxu_len_rate', 'msli_dl_dantong_len_rate', 'msli_dl_mos_poor_len_rate',
        'msli_ul_mos_v2v_avg', 'msli_ul_mos_v2f_avg', 'msli_ul_mos_v2n_avg', 'msli_ul_mos_v2cs_avg', 'msli_ul_mos_v2all_avg',
        'msli_dl_mos_v2v_avg', 'msli_dl_mos_v2f_avg', 'msli_dl_mos_v2n_avg', 'msli_dl_mos_v2cs_avg', 'msli_dl_mos_v2all_avg',
        'msli_ul_rtp_lost_rate', 'msli_dl_rtp_lost_rate',
        'volte_sbc_net_suss', 'volte_sbc_net_sums', 'volte_sbc_drops', 'volte_sbc_ans',
        'volte_local_radio_single_voice_call', 'volte_local_radio_dx_call', 'volte_ans_voice_call',
        'volte_local_radio_dtdx_rate', 'volte_ul_tunzi_len', 'volte_ul_dantong_len', 'volte_ul_duanxu_len', 'volte_ul_voice_sum_len',
        'volte_dl_tunzi_len', 'volte_dl_dantong_len', 'volte_dl_duanxu_len', 'volte_dl_voice_sum_len',
        'micro_grid', 'cover_scene1', 'cover_scene2', 'cover_scene3', 'cover_scene4',
        'grid_road', 'marketduty', 'vendor', 'state', 'coverage_type', 'network_type', 'freq'
    ]
    
    # result中的字段定义（关键字段）
    result_fields = [
        {'feild': 'starttime', 'feildName': '时间', 'datatype': '1'},
        {'feild': 'city', 'feildName': '地市', 'datatype': 'character varying'},
        {'feild': 'cgi', 'feildName': '小区', 'datatype': 'character varying'},
        {'feild': 'grid', 'feildName': '责任网格', 'datatype': 'character varying'},
        {'feild': 'area', 'feildName': '区县', 'datatype': 'character varying'},
        {'feild': 'nrcell_name', 'feildName': '小区名称', 'datatype': 'character varying'},
        {'feild': 'cs_reg1_suss_rate', 'feildName': '初始注册成功率（控制面）', 'datatype': 'numeric'},
        {'feild': 'cs_reg_suss_rate', 'feildName': 'VoLTE注册成功率（控制面）', 'datatype': 'numeric'},
        {'feild': 'cs_reg_sbc_suss_rate', 'feildName': 'SBC注册成功率（控制面）', 'datatype': 'numeric'},
        {'feild': 'cs_moc_sbc_suss_rate', 'feildName': '始呼接通率（控制面）', 'datatype': 'numeric'},
        {'feild': 'cs_moc_sbc_180_suss_rate', 'feildName': '始呼接通率(180)（控制面）', 'datatype': 'numeric'},
        {'feild': 'cs_mtc_sbc_suss_rate', 'feildName': '终呼接通率（控制面）', 'datatype': 'numeric'},
        {'feild': 'cs_mtc_sbc_180_suss_rate', 'feildName': '终呼接通率(180)（控制面）', 'datatype': 'numeric'},
        {'feild': 'cs_sbc_suss_rate', 'feildName': '呼叫接通率(MOC+MTC)（控制面）', 'datatype': 'numeric'},
        {'feild': 'cs_sbc_180_suss_rate', 'feildName': '呼叫接通率(180_MOC+MTC)（控制面）', 'datatype': 'numeric'},
        {'feild': 'cs_moc_sbc_net_suss_rate', 'feildName': '始呼网络接通率（控制面）', 'datatype': 'numeric'},
        {'feild': 'cs_sbc_net_suss_rate', 'feildName': '网络接通率(MOC+MTC)（控制面）', 'datatype': 'numeric'},
        {'feild': 'cs_sbc_drops_rate', 'feildName': 'VOLTE+掉话率（控制面）', 'datatype': 'numeric'},
        {'feild': 'cs_xsrvcc_ho_suss_rate', 'feildName': 'xSRVCC切换成功率（控制面）', 'datatype': 'numeric'},
        {'feild': 'cs_ho_len_avg', 'feildName': 'SRVCC平均切换时长(ms)（控制面）', 'datatype': 'numeric'},
        {'feild': 'cs_ho_rtp_delay_avg', 'feildName': 'SRVCC平均媒体切换时长(ms)（控制面）', 'datatype': 'numeric'},
        {'feild': 'cs_alert_delay_vf_avg', 'feildName': '呼叫建立平均时长(V-固网IMS)（控制面）', 'datatype': 'numeric'},
        {'feild': 'cs_alert_delay_vv_avg', 'feildName': '呼叫建立平均时长(V2V)（控制面）', 'datatype': 'numeric'},
        {'feild': 'cs_alert_delay_vall_rate', 'feildName': '呼叫建立平均时长(V2ALL)（控制面）', 'datatype': 'numeric'},
        {'feild': 'cs_mt_alert_delay_avg', 'feildName': '终呼平均接续时长(ms)（控制面）', 'datatype': 'numeric'},
        {'feild': 'mconv_rtp_ul_mos_avg', 'feildName': 'RTP上行平均MOS（会话）', 'datatype': 'numeric'},
        {'feild': 'mconv_rtp_dl_mos_avg', 'feildName': 'RTP下行平均MOS（会话）', 'datatype': 'numeric'},
        {'feild': 'mconv_mos_300_nok_rate', 'feildName': 'RTP上行MOS 3.0 差占比率（会话）', 'datatype': 'numeric'},
        {'feild': 'mconv_rtcp_ul_mos_avg', 'feildName': 'RTCP上行平均MOS（会话）', 'datatype': 'numeric'},
        {'feild': 'mconv_rtcp_dl_mos0', 'feildName': 'RTCP下行平均MOS（会话）', 'datatype': 'numeric'},
        {'feild': 'mconv_rtp_ul_pkts_lost_rate', 'feildName': 'RTP上行丢包率（会话）', 'datatype': 'numeric'},
        {'feild': 'mconv_rtp_dl_pkts_lost_rate', 'feildName': 'RTP下行丢包率（会话）', 'datatype': 'numeric'},
        {'feild': 'mconv_rtcp_ul_pkts_lost_rate', 'feildName': 'RTCP上行丢包率（会话）', 'datatype': 'numeric'},
        {'feild': 'mconv_rtcp_dl_pkts_lost_rate', 'feildName': 'RTCP下行丢包率（会话）', 'datatype': 'numeric'},
        {'feild': 'mconv_single_voice_call_rate', 'feildName': 'VoLTE语音单通率（会话）', 'datatype': 'numeric'},
        {'feild': 'mconv_dx_call_rate', 'feildName': 'VoLTE语音断续/掉话率（会话）', 'datatype': 'numeric'},
        {'feild': 'mconv_rtp_ul_delay_avg', 'feildName': 'RTP上行平均时延(us)（会话）', 'datatype': 'numeric'},
        {'feild': 'mconv_rtp_dl_delay_avg', 'feildName': 'RTP下行平均时延(us)（会话）', 'datatype': 'numeric'},
        {'feild': 'mconv_rtcp_ul_delay_avg', 'feildName': 'RTCP上行平均时延(us)（会话）', 'datatype': 'numeric'},
        {'feild': 'mconv_rtcp_dl_delay_avg', 'feildName': 'RTCP下行平均时延(us)（会话）', 'datatype': 'numeric'},
        {'feild': 'mconv_dl_mos_300_nok_rate', 'feildName': 'RTP下行MOS 3.0 差占比率（会话）', 'datatype': 'numeric'},
        {'feild': 'msli_ul_tunzi_len_rate', 'feildName': 'VoLTE语音上行质差率（片段）', 'datatype': 'numeric'},
        {'feild': 'msli_ul_duanxu_len_rate', 'feildName': 'VoLTE语音上行断续率（片段）', 'datatype': 'numeric'},
        {'feild': 'msli_ul_dantong_len_rate', 'feildName': 'VoLTE语音上行单通率（片段）', 'datatype': 'numeric'},
        {'feild': 'msli_ul_mos_poor_len_rate', 'feildName': 'VoLTE上行MOS质差率（片段）', 'datatype': 'numeric'},
        {'feild': 'msli_dl_tunzi_len_rate', 'feildName': 'VoLTE语音下行质差率（片段）', 'datatype': 'numeric'},
        {'feild': 'msli_dl_duanxu_len_rate', 'feildName': 'VoLTE语音下行断续率（片段）', 'datatype': 'numeric'},
        {'feild': 'msli_dl_dantong_len_rate', 'feildName': 'VoLTE语音下行单通率（片段）', 'datatype': 'numeric'},
        {'feild': 'msli_dl_mos_poor_len_rate', 'feildName': 'VoLTE下行MOS质差率（片段）', 'datatype': 'numeric'},
        {'feild': 'msli_ul_mos_v2v_avg', 'feildName': 'VoLTE上行平均MOS（对端VoLTE）（片段）', 'datatype': 'numeric'},
        {'feild': 'msli_ul_mos_v2f_avg', 'feildName': 'VoLTE上行平均MOS（对端EPS FB）（片段）', 'datatype': 'numeric'},
        {'feild': 'msli_ul_mos_v2n_avg', 'feildName': 'VoLTE上行平均MOS（对端VoNR）（片段）', 'datatype': 'numeric'},
        {'feild': 'msli_ul_mos_v2cs_avg', 'feildName': 'VoLTE上行平均MOS（对端CS）（片段）', 'datatype': 'numeric'},
        {'feild': 'msli_ul_mos_v2all_avg', 'feildName': 'VoLTE上行平均MOS(对端ALL)（片段）', 'datatype': 'numeric'},
        {'feild': 'msli_dl_mos_v2v_avg', 'feildName': 'VoLTE下行平均MOS（对端VoLTE）（片段）', 'datatype': 'numeric'},
        {'feild': 'msli_dl_mos_v2f_avg', 'feildName': 'VoLTE下行平均MOS（对端EPS FB）（片段）', 'datatype': 'numeric'},
        {'feild': 'msli_dl_mos_v2n_avg', 'feildName': 'VoLTE下行平均MOS（对端VoNR）（片段）', 'datatype': 'numeric'},
        {'feild': 'msli_dl_mos_v2cs_avg', 'feildName': 'VoLTE下行平均MOS（对端CS）（片段）', 'datatype': 'numeric'},
        {'feild': 'msli_dl_mos_v2all_avg', 'feildName': 'VoLTE下行平均MOS(对端ALL)（片段）', 'datatype': 'numeric'},
        {'feild': 'msli_ul_rtp_lost_rate', 'feildName': '上行RTP丢包率（片段）', 'datatype': 'numeric'},
        {'feild': 'msli_dl_rtp_lost_rate', 'feildName': '下行RTP丢包率（片段）', 'datatype': 'numeric'},
        {'feild': 'volte_sbc_net_suss', 'feildName': 'VoLTE_网络接通次数(MOC+MTC)', 'datatype': 'numeric'},
        {'feild': 'volte_sbc_net_sums', 'feildName': 'VoLTE_网络试呼次数(MOC+MTC)', 'datatype': 'numeric'},
        {'feild': 'volte_sbc_drops', 'feildName': 'VoLTE_掉话次数', 'datatype': 'numeric'},
        {'feild': 'volte_sbc_ans', 'feildName': 'VoLTE_应答复次数(掉话率)', 'datatype': 'numeric'},
        {'feild': 'volte_local_radio_single_voice_call', 'feildName': 'VoLTE_语音本端无线单通通话次数', 'datatype': 'numeric'},
        {'feild': 'volte_local_radio_dx_call', 'feildName': 'VoLTE_语音本端无线断续通话次数', 'datatype': 'numeric'},
        {'feild': 'volte_ans_voice_call', 'feildName': 'VoLTE_语音通话总次数', 'datatype': 'numeric'},
        {'feild': 'volte_local_radio_dtdx_rate', 'feildName': 'VoLTE_单通断续次数占比', 'datatype': 'numeric'},
        {'feild': 'volte_ul_tunzi_len', 'feildName': 'VoLTE_语音上行质差时长(s)', 'datatype': 'numeric'},
        {'feild': 'volte_ul_dantong_len', 'feildName': 'VoLTE_语音上行单通时长(s)', 'datatype': 'numeric'},
        {'feild': 'volte_ul_duanxu_len', 'feildName': 'VoLTE_语音上行断续时长(s)', 'datatype': 'numeric'},
        {'feild': 'volte_ul_voice_sum_len', 'feildName': 'VoLTE_语音上行总时长(s)', 'datatype': 'numeric'},
        {'feild': 'volte_dl_tunzi_len', 'feildName': 'VoLTE_语音下行质差时长(s)', 'datatype': 'numeric'},
        {'feild': 'volte_dl_dantong_len', 'feildName': 'VoLTE_语音下行单通时长(s)', 'datatype': 'numeric'},
        {'feild': 'volte_dl_duanxu_len', 'feildName': 'VoLTE_语音下行断续时长(s)', 'datatype': 'numeric'},
        {'feild': 'volte_dl_voice_sum_len', 'feildName': 'VoLTE_语音下行总时长(s)', 'datatype': 'numeric'},
        {'feild': 'micro_grid', 'feildName': '微网格标识', 'datatype': 'character varying'},
        {'feild': 'cover_scene1', 'feildName': '覆盖场景1', 'datatype': 'character varying'},
        {'feild': 'cover_scene2', 'feildName': '覆盖场景2', 'datatype': 'character varying'},
        {'feild': 'cover_scene3', 'feildName': '覆盖场景3', 'datatype': 'character varying'},
        {'feild': 'cover_scene4', 'feildName': '覆盖场景4', 'datatype': 'character varying'},
        {'feild': 'grid_road', 'feildName': '网格道路', 'datatype': 'character varying'},
        {'feild': 'marketduty', 'feildName': '市场职责', 'datatype': 'character varying'},
        {'feild': 'vendor', 'feildName': '厂商', 'datatype': 'character varying'},
        {'feild': 'state', 'feildName': '网元状态', 'datatype': 'character varying'},
        {'feild': 'coverage_type', 'feildName': '覆盖类型', 'datatype': 'character varying'},
        {'feild': 'network_type', 'feildName': '网络制式', 'datatype': 'character varying'},
        {'feild': 'freq', 'feildName': '频段_无线', 'datatype': 'character varying'},
    ]
    
    # 构建result字段（datatype需要根据字段类型设置，前5个字段用'1'，其他用'character varying'）
    # 根据浏览器HAR分析：starttime, city, cgi, grid, area 使用 datatype='1'，其他用 'character varying'
    fixed_datatype_fields = {'starttime', 'city', 'cgi', 'grid', 'area'}
    result_list = []
    for f in result_fields:
        # 字段类型：前5个字段用'1'，其他用'character varying'（与浏览器请求一致）
        field_datatype = '1' if f['feild'] in fixed_datatype_fields else 'character varying'
        result_list.append({
            'feildtype': 'VoLTE小区监控预警数据表-天',
            'table': 'csem.f_nk_volte_keykpi_cell_d',
            'tableName': 'VoLTE小区监控预警数据表-天',
            'datatype': field_datatype,
            'columntype': 1,
            'feildName': f['feildName'],
            'feild': f['feild'],
            'poly': '无',
            'anyWay': '无',
            'chart': '无',
            'chartpoly': '无'
        })
    
    payload = {
        'draw': 1,
        'start': 0,
        'length': 200,
        'total': 0,
        'geographicdimension': '小区',
        'timedimension': '天',
        'enodebField': 'enodeb_id',
        'cgiField': 'cgi',
        'timeField': 'starttime',
        'cellField': 'cell',
        'cityField': 'city',
        'columns': _build_columns_param(volte_fields),
        'order': [{'column': 0, 'dir': 'desc'}],
        'search': {'value': '', 'regex': False},
        'result': {'result': result_list, 'tableParams': {'supporteddimension': None, 'supportedtimedimension': ''}, 'columnname': ''},
        'where': [
            {'datatype': 'timestamp', 'feild': 'starttime', 'feildName': '', 'symbol': '>=', 'val': '2026-04-19 00:00:00', 'whereCon': 'and', 'query': True},
            {'datatype': 'timestamp', 'feild': 'starttime', 'feildName': '', 'symbol': '<', 'val': '2026-04-19 23:59:59', 'whereCon': 'and', 'query': True},
            {'datatype': 'character', 'feild': 'city', 'feildName': '', 'symbol': 'in', 'val': '阳江', 'whereCon': 'and', 'query': True}
        ],
        'indexcount': 0
    }
    return payload


def get_epsfb_warning_payload():
    """获取EPSFB小区监控预警payload"""
    print("[DEBUG-PAYLOAD] 生成 EPSFB小区监控预警 payload")
    
    # EPSFB表完整字段列表（153个字段）
    epsfb_fields = [
        'starttime', 'city', 'cgi', 'grid', 'area', 'nrcell_name',
        'sacs_start_moc_net_succ_rate', 'sacs_start_mtc_net_succ_rate', 'sacs_start_call_net_succ_rate',
        'sacs_start_moc_sbc_180_suss_rate', 'sacs_start_moc_sbc_suss_rate',
        'sacs_start_mtc_sbc_180_suss_rate', 'sacs_start_mtc_sbc_suss_rate',
        'sacs_start_sbc_180_suss_rate', 'sacs_start_sbc_suss_rate',
        'sacs_start_call_drop_rate', 'sacs_start_fb_succ_rate', 'sacs_start_fb_net_succ_rate',
        'sacs_start_fb_ho_succ_rate', 'sacs_start_fb_ho_net_succ_rate',
        'sacs_start_fb_rd_succ_rate', 'sacs_start_fb_rd_net_succ_rate',
        'sacs_start_fb_ho_rd_succ_rate', 'sacs_start_fb_ho_rd_net_succ_rate',
        'sacs_start_rt_succ_rate', 'sacs_start_rt_ho_succ_rate',
        'sacs_start_rt_rd_succ_rate', 'sacs_start_rt_ho_rd_succ_rate',
        'sacs_start_eps_fb_delay_avg', 'sacs_start_fb_ho_delay_avg',
        'sacs_start_fb_rd_delay_avg', 'sacs_start_fb_ho_rd_delay_avg',
        'sacs_start_eps_rt_delay_avg', 'sacs_start_rt_ho_delay_avg',
        'sacs_start_rt_rd_delay_avg', 'sacs_start_rt_ho_rd_delay_avg',
        'sacs_start_rt_25_lt_rate', 'sacs_start_rt_ho_25_lt_rate',
        'sacs_start_rt_rd_25_lt_rate', 'sacs_start_rt_ho_rd_25_lt_rate',
        'sacs_start_alert_delay_f2f_avg', 'sacs_start_alert_delay_f2v_avg',
        'sacs_start_alert_delay_f2n_avg', 'sacs_start_alert_delay_f2a_avg',
        'sacs_start_alert_delay_v2f_avg', 'sacs_start_epsfb_hold_len_avg',
        'sacs_end_moc_net_succ_rate', 'sacs_end_mtc_net_succ_rate',
        'sacs_end_call_net_succ_rate', 'sacs_end_moc_sbc_180_suss_rate',
        'sacs_end_moc_sbc_suss_rate', 'sacs_end_mtc_sbc_180_suss_rate',
        'sacs_end_mtc_sbc_suss_rate', 'sacs_end_sbc_180_suss_rate',
        'sacs_end_sbc_suss_rate', 'sacs_end_call_drop_rate',
        'sacs_end_fb_succ_rate', 'sacs_end_fb_net_succ_rate',
        'sacs_end_fb_ho_succ_rate', 'sacs_end_fb_ho_net_succ_rate',
        'sacs_end_fb_rd_succ_rate', 'sacs_end_fb_rd_net_succ_rate',
        'sacs_end_fb_ho_rd_succ_rate', 'sacs_end_fb_ho_rd_net_succ_rate',
        'sacs_end_rt_succ_rate', 'sacs_end_rt_ho_succ_rate',
        'sacs_end_rt_rd_succ_rate', 'sacs_end_rt_ho_rd_succ_rate',
        'sacs_end_eps_fb_delay_avg', 'sacs_end_fb_ho_delay_avg',
        'sacs_end_fb_rd_delay_avg', 'sacs_end_fb_ho_rd_delay_avg',
        'sacs_end_eps_rt_delay_avg', 'sacs_end_rt_ho_delay_avg',
        'sacs_end_rt_rd_delay_avg', 'sacs_end_rt_ho_rd_delay_avg',
        'sacs_end_rt_25_lt_rate', 'sacs_end_rt_ho_25_lt_rate',
        'sacs_end_rt_rd_25_lt_rate', 'sacs_end_rt_ho_rd_25_lt_rate',
        'sacs_end_alert_delay_f2f_avg', 'sacs_end_alert_delay_f2v_avg',
        'sacs_end_alert_delay_f2n_avg', 'sacs_end_alert_delay_f2a_avg',
        'sacs_end_alert_delay_v2f_avg', 'sacs_end_epsfb_hold_len_avg',
        'mconv_rtp_ul_mos_avg', 'mconv_rtp_dl_mos_avg', 'mconv_mos_300_nok_rate',
        'mconv_rtcp_ul_mos_avg', 'mconv_rtcp_dl_mos0', 'mconv_rtp_ul_pkts_lost_rate',
        'mconv_rtp_dl_pkts_lost_rate', 'mconv_rtcp_ul_pkts_lost_rate',
        'mconv_rtcp_dl_pkts_lost_rate', 'mconv_single_voice_call_rate',
        'mconv_dx_call_rate', 'mconv_rtp_ul_delay_avg', 'mconv_rtp_dl_delay_avg',
        'mconv_rtcp_ul_delay_avg', 'mconv_rtcp_dl_delay_avg', 'mconv_dl_mos_300_nok_rate',
        'msli_ul_tunzi_len_rate', 'msli_ul_duanxu_len_rate', 'msli_ul_dantong_len_rate',
        'msli_ul_mos_poor_len', 'msli_dl_tunzi_len_rate', 'msli_dl_duanxu_len_rate',
        'msli_dl_dantong_len_rate', 'msli_dl_mos_poor_len_rate',
        'msli_ul_mos_f2f_avg', 'msli_ul_mos_f2cs_avg', 'msli_ul_mos_f2all_avg',
        'msli_dl_mos_f2f_avg', 'msli_dl_mos_f2cs0', 'msli_dl_mos_f2all_avg',
        'msli_ul_rtp_lost_rate', 'msli_dl_rtp_lost_rate',
        'msli_ul_mos_f2v_avg', 'msli_ul_mos_f2n_avg',
        'msli_dl_mos_f2v_avg', 'msli_dl_mos_f2n_avg',
        'sacs_start_reg_init_succ_rate', 'sacs_start_reg_sbc_suss_rate', 'sacs_start_reg_suss_rate',
        'epsfb_sbc_net_suss', 'epsfb_sbc_net_sums', 'epsfb_sbc_drops', 'epsfb_sbc_ans',
        'epsfb_local_radio_single_voice_call', 'epsfb_local_radio_dx_call',
        'epsfb_ans_voice_call', 'epsfb_local_radio_dtdx_rate',
        'epsfb_ul_tunzi_len', 'epsfb_ul_dantong_len', 'epsfb_ul_duanxu_len',
        'epsfb_ul_voice_sum_len', 'epsfb_dl_tunzi_len', 'epsfb_dl_dantong_len',
        'epsfb_dl_duanxu_len', 'epsfb_dl_voice_sum_len',
        'micro_grid', 'cover_scene1', 'cover_scene2', 'cover_scene3', 'cover_scene4',
        'grid_road', 'marketduty', 'vendor', 'state', 'coverage_type', 'network_type', 'freq'
    ]
    
    # result中的字段定义（必须包含所有153个字段，和浏览器一致）
    result_fields = [
        {'feild': 'starttime', 'feildName': '时间', 'datatype': '1'},
        {'feild': 'city', 'feildName': '地市', 'datatype': '1'},
        {'feild': 'cgi', 'feildName': '小区', 'datatype': '1'},
        {'feild': 'grid', 'feildName': '责任网格', 'datatype': '1'},
        {'feild': 'area', 'feildName': '区县', 'datatype': '1'},
        {'feild': 'nrcell_name', 'feildName': '小区名称', 'datatype': '1'},
        {'feild': 'sacs_start_moc_net_succ_rate', 'feildName': 'EPSFB始呼网络接通率（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_mtc_net_succ_rate', 'feildName': 'EPSFB终呼网络接通率（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_call_net_succ_rate', 'feildName': 'EPSFB呼叫网络接通率（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_moc_sbc_180_suss_rate', 'feildName': 'EPSFB始呼SBC 180接通率（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_moc_sbc_suss_rate', 'feildName': 'EPSFB始呼SBC接通率（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_mtc_sbc_180_suss_rate', 'feildName': 'EPSFB终呼SBC 180接通率（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_mtc_sbc_suss_rate', 'feildName': 'EPSFB终呼SBC接通率（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_sbc_180_suss_rate', 'feildName': 'EPSFB SBC 180接通率（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_sbc_suss_rate', 'feildName': 'EPSFB SBC接通率（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_call_drop_rate', 'feildName': 'EPSFB掉话率（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_fb_succ_rate', 'feildName': 'EPSFB切换成功率（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_fb_net_succ_rate', 'feildName': 'EPSFB切换网络接通率（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_fb_ho_succ_rate', 'feildName': 'EPSFB切换成功（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_fb_ho_net_succ_rate', 'feildName': 'EPSFB切换网络接通（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_fb_rd_succ_rate', 'feildName': 'EPSFB重定向成功率（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_fb_rd_net_succ_rate', 'feildName': 'EPSFB重定向网络接通率（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_fb_ho_rd_succ_rate', 'feildName': 'EPSFB切换重定向成功率（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_fb_ho_rd_net_succ_rate', 'feildName': 'EPSFB切换重定向网络接通率（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_rt_succ_rate', 'feildName': 'EPSFB返回成功率（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_rt_ho_succ_rate', 'feildName': 'EPSFB返回切换成功率（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_rt_rd_succ_rate', 'feildName': 'EPSFB返回重定向成功率（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_rt_ho_rd_succ_rate', 'feildName': 'EPSFB返回切换重定向成功率（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_eps_fb_delay_avg', 'feildName': 'EPSFB平均时延（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_fb_ho_delay_avg', 'feildName': 'EPSFB切换平均时延（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_fb_rd_delay_avg', 'feildName': 'EPSFB重定向平均时延（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_fb_ho_rd_delay_avg', 'feildName': 'EPSFB切换重定向平均时延（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_eps_rt_delay_avg', 'feildName': 'EPSFB返回平均时延（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_rt_ho_delay_avg', 'feildName': 'EPSFB返回切换平均时延（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_rt_rd_delay_avg', 'feildName': 'EPSFB返回重定向平均时延（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_rt_ho_rd_delay_avg', 'feildName': 'EPSFB返回切换重定向平均时延（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_rt_25_lt_rate', 'feildName': 'EPSFB返回2.5G占比（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_rt_ho_25_lt_rate', 'feildName': 'EPSFB返回切换2.5G占比（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_rt_rd_25_lt_rate', 'feildName': 'EPSFB返回重定向2.5G占比（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_rt_ho_rd_25_lt_rate', 'feildName': 'EPSFB返回切换重定向2.5G占比（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_alert_delay_f2f_avg', 'feildName': '呼叫建立平均时长(F2F)（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_alert_delay_f2v_avg', 'feildName': '呼叫建立平均时长(F2V)（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_alert_delay_f2n_avg', 'feildName': '呼叫建立平均时长(F2N)（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_alert_delay_f2a_avg', 'feildName': '呼叫建立平均时长(F2ALL)（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_alert_delay_v2f_avg', 'feildName': '呼叫建立平均时长(V2F)（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_epsfb_hold_len_avg', 'feildName': 'EPSFB用户平均驻留时长（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_end_moc_net_succ_rate', 'feildName': 'EPSFB始呼网络接通率（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_mtc_net_succ_rate', 'feildName': 'EPSFB终呼网络接通率（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_call_net_succ_rate', 'feildName': 'EPSFB呼叫网络接通率（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_moc_sbc_180_suss_rate', 'feildName': 'EPSFB始呼SBC 180接通率（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_moc_sbc_suss_rate', 'feildName': 'EPSFB始呼SBC接通率（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_mtc_sbc_180_suss_rate', 'feildName': 'EPSFB终呼SBC 180接通率（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_mtc_sbc_suss_rate', 'feildName': 'EPSFB终呼SBC接通率（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_sbc_180_suss_rate', 'feildName': 'EPSFB SBC 180接通率（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_sbc_suss_rate', 'feildName': 'EPSFB SBC接通率（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_call_drop_rate', 'feildName': 'EPSFB掉话率（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_fb_succ_rate', 'feildName': 'EPSFB切换成功率（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_fb_net_succ_rate', 'feildName': 'EPSFB切换网络接通率（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_fb_ho_succ_rate', 'feildName': 'EPSFB切换成功（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_fb_ho_net_succ_rate', 'feildName': 'EPSFB切换网络接通（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_fb_rd_succ_rate', 'feildName': 'EPSFB重定向成功率（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_fb_rd_net_succ_rate', 'feildName': 'EPSFB重定向网络接通率（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_fb_ho_rd_succ_rate', 'feildName': 'EPSFB切换重定向成功率（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_fb_ho_rd_net_succ_rate', 'feildName': 'EPSFB切换重定向网络接通率（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_rt_succ_rate', 'feildName': 'EPSFB返回成功率（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_rt_ho_succ_rate', 'feildName': 'EPSFB返回切换成功率（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_rt_rd_succ_rate', 'feildName': 'EPSFB返回重定向成功率（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_rt_ho_rd_succ_rate', 'feildName': 'EPSFB返回切换重定向成功率（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_eps_fb_delay_avg', 'feildName': 'EPSFB平均时延（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_fb_ho_delay_avg', 'feildName': 'EPSFB切换平均时延（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_fb_rd_delay_avg', 'feildName': 'EPSFB重定向平均时延（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_fb_ho_rd_delay_avg', 'feildName': 'EPSFB切换重定向平均时延（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_eps_rt_delay_avg', 'feildName': 'EPSFB返回平均时延（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_rt_ho_delay_avg', 'feildName': 'EPSFB返回切换平均时延（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_rt_rd_delay_avg', 'feildName': 'EPSFB返回重定向平均时延（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_rt_ho_rd_delay_avg', 'feildName': 'EPSFB返回切换重定向平均时延（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_rt_25_lt_rate', 'feildName': 'EPSFB返回2.5G占比（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_rt_ho_25_lt_rate', 'feildName': 'EPSFB返回切换2.5G占比（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_rt_rd_25_lt_rate', 'feildName': 'EPSFB返回重定向2.5G占比（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_rt_ho_rd_25_lt_rate', 'feildName': 'EPSFB返回切换重定向2.5G占比（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_alert_delay_f2f_avg', 'feildName': '呼叫建立平均时长(F2F)（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_alert_delay_f2v_avg', 'feildName': '呼叫建立平均时长(F2V)（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_alert_delay_f2n_avg', 'feildName': '呼叫建立平均时长(F2N)（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_alert_delay_f2a_avg', 'feildName': '呼叫建立平均时长(F2ALL)（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_alert_delay_v2f_avg', 'feildName': '呼叫建立平均时长(V2F)（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_epsfb_hold_len_avg', 'feildName': 'EPSFB用户平均驻留时长（SA控制结束）', 'datatype': '1'},
        {'feild': 'mconv_rtp_ul_mos_avg', 'feildName': 'RTP上行平均MOS', 'datatype': '1'},
        {'feild': 'mconv_rtp_dl_mos_avg', 'feildName': 'RTP下行平均MOS', 'datatype': '1'},
        {'feild': 'mconv_mos_300_nok_rate', 'feildName': 'RTP上行MOS 3.0 差占比率', 'datatype': '1'},
        {'feild': 'mconv_rtcp_ul_mos_avg', 'feildName': 'RTCP上行平均MOS', 'datatype': '1'},
        {'feild': 'mconv_rtcp_dl_mos0', 'feildName': 'RTCP下行平均MOS', 'datatype': '1'},
        {'feild': 'mconv_rtp_ul_pkts_lost_rate', 'feildName': 'RTP上行丢包率', 'datatype': '1'},
        {'feild': 'mconv_rtp_dl_pkts_lost_rate', 'feildName': 'RTP下行丢包率', 'datatype': '1'},
        {'feild': 'mconv_rtcp_ul_pkts_lost_rate', 'feildName': 'RTCP上行丢包率', 'datatype': '1'},
        {'feild': 'mconv_rtcp_dl_pkts_lost_rate', 'feildName': 'RTCP下行丢包率', 'datatype': '1'},
        {'feild': 'mconv_single_voice_call_rate', 'feildName': 'EPSFB语音单通率', 'datatype': '1'},
        {'feild': 'mconv_dx_call_rate', 'feildName': 'EPSFB语音断续/掉话率', 'datatype': '1'},
        {'feild': 'mconv_rtp_ul_delay_avg', 'feildName': 'RTP上行平均时延', 'datatype': '1'},
        {'feild': 'mconv_rtp_dl_delay_avg', 'feildName': 'RTP下行平均时延', 'datatype': '1'},
        {'feild': 'mconv_rtcp_ul_delay_avg', 'feildName': 'RTCP上行平均时延', 'datatype': '1'},
        {'feild': 'mconv_rtcp_dl_delay_avg', 'feildName': 'RTCP下行平均时延', 'datatype': '1'},
        {'feild': 'mconv_dl_mos_300_nok_rate', 'feildName': 'RTP下行MOS 3.0 差占比率', 'datatype': '1'},
        {'feild': 'msli_ul_tunzi_len_rate', 'feildName': 'EPSFB语音上行质差率', 'datatype': '1'},
        {'feild': 'msli_ul_duanxu_len_rate', 'feildName': 'EPSFB语音上行断续率', 'datatype': '1'},
        {'feild': 'msli_ul_dantong_len_rate', 'feildName': 'EPSFB语音上行单通率', 'datatype': '1'},
        {'feild': 'msli_ul_mos_poor_len', 'feildName': 'EPSFB上行MOS质差时长', 'datatype': '1'},
        {'feild': 'msli_dl_tunzi_len_rate', 'feildName': 'EPSFB语音下行质差率', 'datatype': '1'},
        {'feild': 'msli_dl_duanxu_len_rate', 'feildName': 'EPSFB语音下行断续率', 'datatype': '1'},
        {'feild': 'msli_dl_dantong_len_rate', 'feildName': 'EPSFB语音下行单通率', 'datatype': '1'},
        {'feild': 'msli_dl_mos_poor_len_rate', 'feildName': 'EPSFB下行MOS质差率', 'datatype': '1'},
        {'feild': 'msli_ul_mos_f2f_avg', 'feildName': 'EPSFB上行平均MOS（对端VoLTE）', 'datatype': '1'},
        {'feild': 'msli_ul_mos_f2cs_avg', 'feildName': 'EPSFB上行平均MOS（对端CS）', 'datatype': '1'},
        {'feild': 'msli_ul_mos_f2all_avg', 'feildName': 'EPSFB上行平均MOS（对端ALL）', 'datatype': '1'},
        {'feild': 'msli_dl_mos_f2f_avg', 'feildName': 'EPSFB下行平均MOS（对端VoLTE）', 'datatype': '1'},
        {'feild': 'msli_dl_mos_f2cs0', 'feildName': 'EPSFB下行平均MOS（对端CS）', 'datatype': '1'},
        {'feild': 'msli_dl_mos_f2all_avg', 'feildName': 'EPSFB下行平均MOS（对端ALL）', 'datatype': '1'},
        {'feild': 'msli_ul_rtp_lost_rate', 'feildName': '上行RTP丢包率', 'datatype': '1'},
        {'feild': 'msli_dl_rtp_lost_rate', 'feildName': '下行RTP丢包率', 'datatype': '1'},
        {'feild': 'msli_ul_mos_f2v_avg', 'feildName': 'EPSFB上行平均MOS（对端VoNR）', 'datatype': '1'},
        {'feild': 'msli_ul_mos_f2n_avg', 'feildName': 'EPSFB上行平均MOS（对端Vo5GSA）', 'datatype': '1'},
        {'feild': 'msli_dl_mos_f2v_avg', 'feildName': 'EPSFB下行平均MOS（对端VoNR）', 'datatype': '1'},
        {'feild': 'msli_dl_mos_f2n_avg', 'feildName': 'EPSFB下行平均MOS（对端Vo5GSA）', 'datatype': '1'},
        {'feild': 'sacs_start_reg_init_succ_rate', 'feildName': 'EPS+SA+IMS+CSCF初始注册成功率（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_reg_sbc_suss_rate', 'feildName': 'EPS+SA+IMS+SBC注册成功率（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_reg_suss_rate', 'feildName': 'EPS+SA+IMS+CSCF注册成功率[含重注册]（SA控制起始）', 'datatype': '1'},
        {'feild': 'epsfb_sbc_net_suss', 'feildName': 'EPSFB_网络接通次数(MOC+MTC)', 'datatype': '1'},
        {'feild': 'epsfb_sbc_net_sums', 'feildName': 'EPSFB_网络试呼次数(MOC+MTC)', 'datatype': '1'},
        {'feild': 'epsfb_sbc_drops', 'feildName': 'EPSFB_掉话次数', 'datatype': '1'},
        {'feild': 'epsfb_sbc_ans', 'feildName': 'EPSFB_应答复次数(掉话率)', 'datatype': '1'},
        {'feild': 'epsfb_local_radio_single_voice_call', 'feildName': 'EPSFB_语音本端无线单通通话次数', 'datatype': '1'},
        {'feild': 'epsfb_local_radio_dx_call', 'feildName': 'EPSFB_语音本端无线断续通话次数', 'datatype': '1'},
        {'feild': 'epsfb_ans_voice_call', 'feildName': 'EPSFB_语音通话总次数', 'datatype': '1'},
        {'feild': 'epsfb_local_radio_dtdx_rate', 'feildName': 'EPSFB_单通断续次数占比', 'datatype': '1'},
        {'feild': 'epsfb_ul_tunzi_len', 'feildName': 'EPSFB_语音上行质差时长(s)', 'datatype': '1'},
        {'feild': 'epsfb_ul_dantong_len', 'feildName': 'EPSFB_语音上行单通时长(s)', 'datatype': '1'},
        {'feild': 'epsfb_ul_duanxu_len', 'feildName': 'EPSFB_语音上行断续时长(s)', 'datatype': '1'},
        {'feild': 'epsfb_ul_voice_sum_len', 'feildName': 'EPSFB_语音上行总时长(s)', 'datatype': '1'},
        {'feild': 'epsfb_dl_tunzi_len', 'feildName': 'EPSFB_语音下行质差时长(s)', 'datatype': '1'},
        {'feild': 'epsfb_dl_dantong_len', 'feildName': 'EPSFB_语音下行单通时长(s)', 'datatype': '1'},
        {'feild': 'epsfb_dl_duanxu_len', 'feildName': 'EPSFB_语音下行断续时长(s)', 'datatype': '1'},
        {'feild': 'epsfb_dl_voice_sum_len', 'feildName': 'EPSFB_语音下行总时长(s)', 'datatype': '1'},
        {'feild': 'micro_grid', 'feildName': '微网格标识', 'datatype': '1'},
        {'feild': 'cover_scene1', 'feildName': '覆盖场景1', 'datatype': '1'},
        {'feild': 'cover_scene2', 'feildName': '覆盖场景2', 'datatype': '1'},
        {'feild': 'cover_scene3', 'feildName': '覆盖场景3', 'datatype': '1'},
        {'feild': 'cover_scene4', 'feildName': '覆盖场景4', 'datatype': '1'},
        {'feild': 'grid_road', 'feildName': '网格道路', 'datatype': '1'},
        {'feild': 'marketduty', 'feildName': '市场职责', 'datatype': '1'},
        {'feild': 'vendor', 'feildName': '厂商', 'datatype': '1'},
        {'feild': 'state', 'feildName': '网元状态', 'datatype': '1'},
        {'feild': 'coverage_type', 'feildName': '覆盖类型', 'datatype': '1'},
        {'feild': 'network_type', 'feildName': '网络制式', 'datatype': '1'},
        {'feild': 'freq', 'feildName': '频段_无线', 'datatype': '1'},
    ]
    
    # 构建result字段（datatype需要根据字段类型设置，前5个字段用'1'，其他用'character varying'）
    # 根据浏览器HAR分析：starttime, city, cgi, grid, area 使用 datatype='1'，其他用 'character varying'
    fixed_datatype_fields = {'starttime', 'city', 'cgi', 'grid', 'area'}
    result_list = []
    for f in result_fields:
        # 字段类型：前5个字段用'1'，其他用'character varying'（与浏览器请求一致）
        field_datatype = '1' if f['feild'] in fixed_datatype_fields else 'character varying'
        result_list.append({
            'feildtype': 'EPSFB小区监控预警数据表-天',
            'table': 'csem.f_nk_epsfb_keykpi_cell_d',
            'tableName': 'EPSFB小区监控预警数据表-天',
            'datatype': field_datatype,
            'columntype': 1,
            'feildName': f['feildName'],
            'feild': f['feild'],
            'poly': '无',
            'anyWay': '无',
            'chart': '无',
            'chartpoly': '无'
        })
    
    payload = {
        'draw': 1,
        'start': 0,
        'length': 200,
        'total': 0,
        'geographicdimension': '小区',
        'timedimension': '天',
        'enodebField': '---',
        'cgiField': 'cgi',
        'timeField': 'starttime',
        'cellField': 'cell',
        'cityField': 'city',
        'columns': _build_columns_param(epsfb_fields),
        'order': [{'column': 0, 'dir': 'desc'}],
        'search': {'value': '', 'regex': False},
        'result': {'result': result_list, 'tableParams': {'supporteddimension': None, 'supportedtimedimension': ''}, 'columnname': ''},
        'where': [
            {'datatype': 'timestamp', 'feild': 'starttime', 'feildName': '', 'symbol': '>=', 'val': '2026-04-19 00:00:00', 'whereCon': 'and', 'query': True},
            {'datatype': 'timestamp', 'feild': 'starttime', 'feildName': '', 'symbol': '<', 'val': '2026-04-19 23:59:59', 'whereCon': 'and', 'query': True},
            {'datatype': 'character', 'feild': 'city', 'feildName': '', 'symbol': 'in', 'val': '阳江', 'whereCon': 'and', 'query': True}
        ],
        'indexcount': 0
    }
    return payload


def get_vonr_warning_payload():
    """获取VONR小区监控预警payload"""
    print("[DEBUG-PAYLOAD] 生成 VONR小区监控预警 payload")
    
    # VONR表完整字段列表（152个字段）
    vonr_fields = [
        'starttime', 'city', 'cgi', 'grid', 'area', 'nrcell_name',
        'sacs_start_reg_init_succ_rate', 'sacs_start_reg_suss_rate', 'sacs_start_reg_sbc_suss_rate',
        'sacs_start_vonr_moc_net_succ_rate', 'sacs_start_vonr_mtc_net_succ_rate',
        'sacs_start_vonr_call_net_succ_rate', 'sacs_start_vonr_moc_sbc_180_suss_rate',
        'sacs_start_vonr_moc_sbc_suss_rate', 'sacs_start_vonr_mtc_sbc_180_suss_rate',
        'sacs_start_vonr_mtc_sbc_suss', 'sacs_start_vonr_sbc_180_suss_rate',
        'sacs_start_vonr_sbc_suss_rate', 'sacs_start_vonr_call_drop_rate',
        'sacs_start_alert_delay_n2n_avg', 'sacs_start_alert_delay_n2v_avg',
        'sacs_start_alert_delay_n2f_avg', 'sacs_start_alert_delay_n2a_avg',
        'sacs_start_alert_delay_v2n_avg',
        'sacs_start_vo5gsa_moc_net_succ_rate', 'sacs_start_vo5gsa_mtc_net_succ_rate',
        'sacs_start_vo5gsa_call_net_succ_rate', 'sacs_start_vo5gsa_moc_sbc_180_suss_rate',
        'sacs_start_vo5gsa_moc_sbc_suss_rate', 'sacs_start_vo5gsa_mtc_sbc_180_suss_rate',
        'sacs_start_vo5gsa_mtc_sbc_suss_rate', 'sacs_start_vo5gsa_sbc_180_suss_rate',
        'sacs_start_vo5gsa_sbc_suss_rate', 'sacs_start_vo5gsa_call_drop_rate',
        'sacs_start_vo5gsa_moc_csfb_rate', 'sacs_start_vo5gsa_mtc_csfb_rate',
        'sacs_start_vo5gsa_call_csfb_rate',
        'sacs_start_vonr_hold_avg', 'sacs_start_vonr_moc_call_drop_rate',
        'sacs_start_vonr_mtc_call_drop_rate', 'sacs_start_vonr_ho_xn_succ_rate',
        'sacs_start_vonr_ho_n2_succ_rate', 'sacs_start_vonr_ho_54_succ_rate',
        'sacs_start_vonr_ho_54_c_delay_avg', 'sacs_start_vonr_ho_54_u_delay_avg',
        'csretryrate',
        'sacs_end_reg_init_succ_rate', 'sacs_end_reg_suss_rate', 'sacs_end_reg_sbc_suss_rate',
        'sacs_end_vonr_moc_net_succ_rate', 'sacs_end_vonr_mtc_net_succ_rate',
        'sacs_end_vonr_call_net_succ_rate', 'sacs_end_vonr_moc_sbc_180_suss_rate',
        'sacs_end_vonr_moc_sbc_suss_rate', 'sacs_end_vonr_mtc_sbc_180_suss_rate',
        'sacs_end_vonr_mtc_sbc_suss', 'sacs_end_vonr_sbc_180_suss_rate',
        'sacs_end_vonr_sbc_suss_rate', 'sacs_end_vonr_call_drop_rate',
        'sacs_end_alert_delay_n2n_avg', 'sacs_end_alert_delay_n2v_avg',
        'sacs_end_alert_delay_n2f_avg', 'sacs_end_alert_delay_n2a_avg',
        'sacs_end_alert_delay_v2n_avg',
        'sacs_end_vo5gsa_moc_net_succ_rate', 'sacs_end_vo5gsa_mtc_net_succ_rate',
        'sacs_end_vo5gsa_call_net_succ_rate', 'sacs_end_vo5gsa_moc_sbc_180_suss_rate',
        'sacs_end_vo5gsa_moc_sbc_suss_rate', 'sacs_end_vo5gsa_mtc_sbc_180_suss_rate',
        'sacs_end_vo5gsa_mtc_sbc_suss_rate', 'sacs_end_vo5gsa_sbc_180_suss_rate',
        'sacs_end_vo5gsa_sbc_suss_rate', 'sacs_end_vo5gsa_call_drop_rate',
        'sacs_end_vo5gsa_moc_csfb_rate', 'sacs_end_vo5gsa_mtc_csfb_rate',
        'sacs_end_vo5gsa_call_csfb_rate',
        'sacs_end_vonr_hold_avg', 'sacs_end_vonr_moc_call_drop_rate',
        'sacs_end_vonr_mtc_call_drop_rate', 'sacs_end_vonr_ho_xn_succ_rate',
        'sacs_end_vonr_ho_n2_succ_rate', 'sacs_end_vonr_ho_54_succ_rate',
        'sacs_end_vonr_ho_54_c_delay_avg', 'sacs_end_vonr_ho_54_u_delay_avg',
        'sacs_end_vonr_ho_54_attrate',
        'mconv_rtp_ul_mos_avg', 'mconv_rtp_dl_mos_avg',
        'mconv_ul_mos_300_ok_rate', 'mconv_dl_mos_300_ok_rate',
        'mconv_ul_mos_400_ok_rate', 'mconv_dl_mos_400_ok_rate',
        'mconv_rtcp_ul_mos_avg', 'mconv_rtcp_dl_mos0', 'mconv_rtp_ul_pkts_lost_rate',
        'mconv_rtp_dl_pkts_lost_rate', 'mconv_rtcp_ul_pkts_lost_rate',
        'mconv_rtcp_dl_pkts_lost_rate', 'mconv_single_voice_call_rate',
        'mconv_dx_call_rate', 'mconv_rtp_ul_delay_avg', 'mconv_rtp_dl_delay_avg',
        'mconv_rtcp_ul_delay_avg', 'mconv_rtcp_dl_delay_avg',
        'msli_ul_tunzi_len_rate', 'msli_ul_duanxu_len_rate', 'msli_ul_dantong_len_rate',
        'msli_ul_mos_poor_len_rate', 'msli_dl_tunzi_len_rate', 'msli_dl_duanxu_len_rate',
        'msli_dl_dantong_len_rate', 'msli_dl_mos_poor_len_rate',
        'msli_ul_mos_n2n_avg', 'msli_ul_mos_n2v_avg', 'msli_ul_mos_n2f_avg',
        'msli_ul_mos_n2cs_avg', 'msli_ul_mos_n2all_avg',
        'msli_dl_mos_n2n_avg', 'msli_dl_mos_n2v_avg', 'msli_dl_mos_n2f_avg',
        'msli_dl_mos_n2cs_avg', 'msli_dl_mos_n2all_avg',
        'msli_ul_rtp_lost_rate', 'msli_dl_rtp_lost_rate',
        'vonr_sbc_net_suss', 'vonr_sbc_net_sums', 'vonr_sbc_drops', 'vonr_sbc_ans',
        'vonr_local_radio_single_voice_call', 'vonr_local_radio_dx_call',
        'vonr_ans_voice_call', 'vonr_local_radio_dtdx_rate',
        'vonr_ul_tunzi_len', 'vonr_ul_dantong_len', 'vonr_ul_duanxu_len',
        'vonr_ul_voice_sum_len', 'vonr_dl_tunzi_len', 'vonr_dl_dantong_len',
        'vonr_dl_duanxu_len', 'vonr_dl_voice_sum_len',
        'micro_grid', 'cover_scene1', 'cover_scene2', 'cover_scene3', 'cover_scene4',
        'grid_road', 'marketduty', 'vendor', 'state', 'coverage_type', 'network_type', 'freq'
    ]
    
    # result中的字段定义（必须包含所有152个字段，和columns一致）
    result_fields = [
        {'feild': 'starttime', 'feildName': '时间', 'datatype': '1'},
        {'feild': 'city', 'feildName': '地市', 'datatype': '1'},
        {'feild': 'cgi', 'feildName': '小区', 'datatype': '1'},
        {'feild': 'grid', 'feildName': '责任网格', 'datatype': '1'},
        {'feild': 'area', 'feildName': '区县', 'datatype': '1'},
        {'feild': 'nrcell_name', 'feildName': '小区名称', 'datatype': '1'},
        {'feild': 'sacs_start_reg_init_succ_rate', 'feildName': '5G SA IMS CSCF初始注册成功率（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_reg_suss_rate', 'feildName': '5G SA IMS CSCF注册成功率[含重注册]（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_reg_sbc_suss_rate', 'feildName': '5G SA IMS SBC注册成功率（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_vonr_moc_net_succ_rate', 'feildName': 'VoNR始呼网络接通率（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_vonr_mtc_net_succ_rate', 'feildName': 'VoNR终呼网络接通率（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_vonr_call_net_succ_rate', 'feildName': 'VoNR呼叫网络接通率（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_vonr_moc_sbc_180_suss_rate', 'feildName': 'VoNR始呼SBC 180接通率（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_vonr_moc_sbc_suss_rate', 'feildName': 'VoNR始呼SBC接通率（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_vonr_mtc_sbc_180_suss_rate', 'feildName': 'VoNR终呼SBC 180接通率（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_vonr_mtc_sbc_suss', 'feildName': 'VoNR终呼SBC接通数（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_vonr_sbc_180_suss_rate', 'feildName': 'VoNR SBC 180接通率（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_vonr_sbc_suss_rate', 'feildName': 'VoNR SBC接通率（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_vonr_call_drop_rate', 'feildName': 'VoNR掉话率（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_alert_delay_n2n_avg', 'feildName': '呼叫建立平均时长(N2N)（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_alert_delay_n2v_avg', 'feildName': '呼叫建立平均时长(N2V)（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_alert_delay_n2f_avg', 'feildName': '呼叫建立平均时长(N2F)（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_alert_delay_n2a_avg', 'feildName': '呼叫建立平均时长(N2ALL)（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_alert_delay_v2n_avg', 'feildName': '呼叫建立平均时长(V2N)（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_vo5gsa_moc_net_succ_rate', 'feildName': 'Vo5GSA始呼网络接通率（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_vo5gsa_mtc_net_succ_rate', 'feildName': 'Vo5GSA终呼网络接通率（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_vo5gsa_call_net_succ_rate', 'feildName': 'Vo5GSA呼叫网络接通率（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_vo5gsa_moc_sbc_180_suss_rate', 'feildName': 'Vo5GSA始呼SBC 180接通率（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_vo5gsa_moc_sbc_suss_rate', 'feildName': 'Vo5GSA始呼SBC接通率（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_vo5gsa_mtc_sbc_180_suss_rate', 'feildName': 'Vo5GSA终呼SBC 180接通率（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_vo5gsa_mtc_sbc_suss_rate', 'feildName': 'Vo5GSA终呼SBC接通率（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_vo5gsa_sbc_180_suss_rate', 'feildName': 'Vo5GSA SBC 180接通率（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_vo5gsa_sbc_suss_rate', 'feildName': 'Vo5GSA SBC接通率（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_vo5gsa_call_drop_rate', 'feildName': 'Vo5GSA掉话率（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_vo5gsa_moc_csfb_rate', 'feildName': 'Vo5GSA始呼CSFB占比（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_vo5gsa_mtc_csfb_rate', 'feildName': 'Vo5GSA终呼CSFB占比（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_vo5gsa_call_csfb_rate', 'feildName': 'Vo5GSA呼叫CSFB占比（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_vonr_hold_avg', 'feildName': 'VoNR用户平均驻留时长（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_vonr_moc_call_drop_rate', 'feildName': 'VoNR始发掉话率（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_vonr_mtc_call_drop_rate', 'feildName': 'VoNR终端掉话率（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_vonr_ho_xn_succ_rate', 'feildName': 'VoNR系统间切换成功率（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_vonr_ho_n2_succ_rate', 'feildName': 'VoNR系统内切换成功率（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_vonr_ho_54_succ_rate', 'feildName': 'VoNR 5G->4G切换成功率（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_vonr_ho_54_c_delay_avg', 'feildName': 'VoNR 5G->4G切换控制面平均时延（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_start_vonr_ho_54_u_delay_avg', 'feildName': 'VoNR 5G->4G切换用户面平均时延（SA控制起始）', 'datatype': '1'},
        {'feild': 'csretryrate', 'feildName': 'Csretry占比（%）（SA控制起始）', 'datatype': '1'},
        {'feild': 'sacs_end_reg_init_succ_rate', 'feildName': '5G SA IMS CSCF初始注册成功率（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_reg_suss_rate', 'feildName': '5G SA IMS CSCF注册成功率[含重注册]（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_reg_sbc_suss_rate', 'feildName': '5G SA IMS SBC注册成功率（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_vonr_moc_net_succ_rate', 'feildName': 'VoNR始呼网络接通率（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_vonr_mtc_net_succ_rate', 'feildName': 'VoNR终呼网络接通率（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_vonr_call_net_succ_rate', 'feildName': 'VoNR呼叫网络接通率（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_vonr_moc_sbc_180_suss_rate', 'feildName': 'VoNR始呼SBC 180接通率（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_vonr_moc_sbc_suss_rate', 'feildName': 'VoNR始呼SBC接通率（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_vonr_mtc_sbc_180_suss_rate', 'feildName': 'VoNR终呼SBC 180接通率（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_vonr_mtc_sbc_suss', 'feildName': 'VoNR终呼SBC接通数（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_vonr_sbc_180_suss_rate', 'feildName': 'VoNR SBC 180接通率（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_vonr_sbc_suss_rate', 'feildName': 'VoNR SBC接通率（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_vonr_call_drop_rate', 'feildName': 'VoNR掉话率（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_alert_delay_n2n_avg', 'feildName': '呼叫建立平均时长(N2N)（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_alert_delay_n2v_avg', 'feildName': '呼叫建立平均时长(N2V)（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_alert_delay_n2f_avg', 'feildName': '呼叫建立平均时长(N2F)（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_alert_delay_n2a_avg', 'feildName': '呼叫建立平均时长(N2ALL)（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_alert_delay_v2n_avg', 'feildName': '呼叫建立平均时长(V2N)（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_vo5gsa_moc_net_succ_rate', 'feildName': 'Vo5GSA始呼网络接通率（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_vo5gsa_mtc_net_succ_rate', 'feildName': 'Vo5GSA终呼网络接通率（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_vo5gsa_call_net_succ_rate', 'feildName': 'Vo5GSA呼叫网络接通率（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_vo5gsa_moc_sbc_180_suss_rate', 'feildName': 'Vo5GSA始呼SBC 180接通率（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_vo5gsa_moc_sbc_suss_rate', 'feildName': 'Vo5GSA始呼SBC接通率（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_vo5gsa_mtc_sbc_180_suss_rate', 'feildName': 'Vo5GSA终呼SBC 180接通率（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_vo5gsa_mtc_sbc_suss_rate', 'feildName': 'Vo5GSA终呼SBC接通率（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_vo5gsa_sbc_180_suss_rate', 'feildName': 'Vo5GSA SBC 180接通率（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_vo5gsa_sbc_suss_rate', 'feildName': 'Vo5GSA SBC接通率（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_vo5gsa_call_drop_rate', 'feildName': 'Vo5GSA掉话率（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_vo5gsa_moc_csfb_rate', 'feildName': 'Vo5GSA始呼CSFB占比（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_vo5gsa_mtc_csfb_rate', 'feildName': 'Vo5GSA终呼CSFB占比（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_vo5gsa_call_csfb_rate', 'feildName': 'Vo5GSA呼叫CSFB占比（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_vonr_hold_avg', 'feildName': 'VoNR用户平均驻留时长（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_vonr_moc_call_drop_rate', 'feildName': 'VoNR始发掉话率（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_vonr_mtc_call_drop_rate', 'feildName': 'VoNR终端掉话率（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_vonr_ho_xn_succ_rate', 'feildName': 'VoNR系统间切换成功率（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_vonr_ho_n2_succ_rate', 'feildName': 'VoNR系统内切换成功率（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_vonr_ho_54_succ_rate', 'feildName': 'VoNR 5G->4G切换成功率（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_vonr_ho_54_c_delay_avg', 'feildName': 'VoNR 5G->4G切换控制面平均时延（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_vonr_ho_54_u_delay_avg', 'feildName': 'VoNR 5G->4G切换用户面平均时延（SA控制结束）', 'datatype': '1'},
        {'feild': 'sacs_end_vonr_ho_54_attrate', 'feildName': 'VoNR 5G->4G切换邻区TAM进服差占比（SA控制结束）', 'datatype': '1'},
        {'feild': 'mconv_rtp_ul_mos_avg', 'feildName': 'RTP上行平均MOS', 'datatype': '1'},
        {'feild': 'mconv_rtp_dl_mos_avg', 'feildName': 'RTP下行平均MOS', 'datatype': '1'},
        {'feild': 'mconv_ul_mos_300_ok_rate', 'feildName': 'RTP上行MOS 3.0 优占比率', 'datatype': '1'},
        {'feild': 'mconv_dl_mos_300_ok_rate', 'feildName': 'RTP下行MOS 3.0 优占比率', 'datatype': '1'},
        {'feild': 'mconv_ul_mos_400_ok_rate', 'feildName': 'RTP上行MOS 4.0 优占比率', 'datatype': '1'},
        {'feild': 'mconv_dl_mos_400_ok_rate', 'feildName': 'RTP下行MOS 4.0 优占比率', 'datatype': '1'},
        {'feild': 'mconv_rtcp_ul_mos_avg', 'feildName': 'RTCP上行平均MOS', 'datatype': '1'},
        {'feild': 'mconv_rtcp_dl_mos0', 'feildName': 'RTCP下行平均MOS', 'datatype': '1'},
        {'feild': 'mconv_rtp_ul_pkts_lost_rate', 'feildName': 'RTP上行丢包率', 'datatype': '1'},
        {'feild': 'mconv_rtp_dl_pkts_lost_rate', 'feildName': 'RTP下行丢包率', 'datatype': '1'},
        {'feild': 'mconv_rtcp_ul_pkts_lost_rate', 'feildName': 'RTCP上行丢包率', 'datatype': '1'},
        {'feild': 'mconv_rtcp_dl_pkts_lost_rate', 'feildName': 'RTCP下行丢包率', 'datatype': '1'},
        {'feild': 'mconv_single_voice_call_rate', 'feildName': 'VoNR语音单通率', 'datatype': '1'},
        {'feild': 'mconv_dx_call_rate', 'feildName': 'VoNR语音断续/掉话率', 'datatype': '1'},
        {'feild': 'mconv_rtp_ul_delay_avg', 'feildName': 'RTP上行平均时延', 'datatype': '1'},
        {'feild': 'mconv_rtp_dl_delay_avg', 'feildName': 'RTP下行平均时延', 'datatype': '1'},
        {'feild': 'mconv_rtcp_ul_delay_avg', 'feildName': 'RTCP上行平均时延', 'datatype': '1'},
        {'feild': 'mconv_rtcp_dl_delay_avg', 'feildName': 'RTCP下行平均时延', 'datatype': '1'},
        {'feild': 'msli_ul_tunzi_len_rate', 'feildName': 'VoNR语音上行质差率', 'datatype': '1'},
        {'feild': 'msli_ul_duanxu_len_rate', 'feildName': 'VoNR语音上行断续率', 'datatype': '1'},
        {'feild': 'msli_ul_dantong_len_rate', 'feildName': 'VoNR语音上行单通率', 'datatype': '1'},
        {'feild': 'msli_ul_mos_poor_len_rate', 'feildName': 'VoNR上行MOS质差率', 'datatype': '1'},
        {'feild': 'msli_dl_tunzi_len_rate', 'feildName': 'VoNR语音下行质差率', 'datatype': '1'},
        {'feild': 'msli_dl_duanxu_len_rate', 'feildName': 'VoNR语音下行断续率', 'datatype': '1'},
        {'feild': 'msli_dl_dantong_len_rate', 'feildName': 'VoNR语音下行单通率', 'datatype': '1'},
        {'feild': 'msli_dl_mos_poor_len_rate', 'feildName': 'VoNR下行MOS质差率', 'datatype': '1'},
        {'feild': 'msli_ul_mos_n2n_avg', 'feildName': 'VoNR上行平均MOS（对端VoNR）', 'datatype': '1'},
        {'feild': 'msli_ul_mos_n2v_avg', 'feildName': 'VoNR上行平均MOS（对端VoLTE）', 'datatype': '1'},
        {'feild': 'msli_ul_mos_n2f_avg', 'feildName': 'VoNR上行平均MOS（对端Vo5GSA）', 'datatype': '1'},
        {'feild': 'msli_ul_mos_n2cs_avg', 'feildName': 'VoNR上行平均MOS（对端CS）', 'datatype': '1'},
        {'feild': 'msli_ul_mos_n2all_avg', 'feildName': 'VoNR上行平均MOS（对端ALL）', 'datatype': '1'},
        {'feild': 'msli_dl_mos_n2n_avg', 'feildName': 'VoNR下行平均MOS（对端VoNR）', 'datatype': '1'},
        {'feild': 'msli_dl_mos_n2v_avg', 'feildName': 'VoNR下行平均MOS（对端VoLTE）', 'datatype': '1'},
        {'feild': 'msli_dl_mos_n2f_avg', 'feildName': 'VoNR下行平均MOS（对端Vo5GSA）', 'datatype': '1'},
        {'feild': 'msli_dl_mos_n2cs_avg', 'feildName': 'VoNR下行平均MOS（对端CS）', 'datatype': '1'},
        {'feild': 'msli_dl_mos_n2all_avg', 'feildName': 'VoNR下行平均MOS（对端ALL）', 'datatype': '1'},
        {'feild': 'msli_ul_rtp_lost_rate', 'feildName': '上行RTP丢包率', 'datatype': '1'},
        {'feild': 'msli_dl_rtp_lost_rate', 'feildName': '下行RTP丢包率', 'datatype': '1'},
        {'feild': 'vonr_sbc_net_suss', 'feildName': 'VoNR_网络接通次数(MOC+MTC)', 'datatype': '1'},
        {'feild': 'vonr_sbc_net_sums', 'feildName': 'VoNR_网络试呼次数(MOC+MTC)', 'datatype': '1'},
        {'feild': 'vonr_sbc_drops', 'feildName': 'VoNR_掉话次数', 'datatype': '1'},
        {'feild': 'vonr_sbc_ans', 'feildName': 'VoNR_应答复次数(掉话率)', 'datatype': '1'},
        {'feild': 'vonr_local_radio_single_voice_call', 'feildName': 'VoNR_语音本端无线单通通话次数', 'datatype': '1'},
        {'feild': 'vonr_local_radio_dx_call', 'feildName': 'VoNR_语音本端无线断续通话次数', 'datatype': '1'},
        {'feild': 'vonr_ans_voice_call', 'feildName': 'VoNR_语音通话总次数', 'datatype': '1'},
        {'feild': 'vonr_local_radio_dtdx_rate', 'feildName': 'VoNR_单通断续次数占比', 'datatype': '1'},
        {'feild': 'vonr_ul_tunzi_len', 'feildName': 'VoNR_语音上行质差时长(s)', 'datatype': '1'},
        {'feild': 'vonr_ul_dantong_len', 'feildName': 'VoNR_语音上行单通时长(s)', 'datatype': '1'},
        {'feild': 'vonr_ul_duanxu_len', 'feildName': 'VoNR_语音上行断续时长(s)', 'datatype': '1'},
        {'feild': 'vonr_ul_voice_sum_len', 'feildName': 'VoNR_语音上行总时长(s)', 'datatype': '1'},
        {'feild': 'vonr_dl_tunzi_len', 'feildName': 'VoNR_语音下行质差时长(s)', 'datatype': '1'},
        {'feild': 'vonr_dl_dantong_len', 'feildName': 'VoNR_语音下行单通时长(s)', 'datatype': '1'},
        {'feild': 'vonr_dl_duanxu_len', 'feildName': 'VoNR_语音下行断续时长(s)', 'datatype': '1'},
        {'feild': 'vonr_dl_voice_sum_len', 'feildName': 'VoNR_语音下行总时长(s)', 'datatype': '1'},
        {'feild': 'micro_grid', 'feildName': '微网格标识', 'datatype': '1'},
        {'feild': 'cover_scene1', 'feildName': '覆盖场景1', 'datatype': '1'},
        {'feild': 'cover_scene2', 'feildName': '覆盖场景2', 'datatype': '1'},
        {'feild': 'cover_scene3', 'feildName': '覆盖场景3', 'datatype': '1'},
        {'feild': 'cover_scene4', 'feildName': '覆盖场景4', 'datatype': '1'},
        {'feild': 'grid_road', 'feildName': '网格道路', 'datatype': '1'},
        {'feild': 'marketduty', 'feildName': '市场职责', 'datatype': '1'},
        {'feild': 'vendor', 'feildName': '厂商', 'datatype': '1'},
        {'feild': 'state', 'feildName': '网元状态', 'datatype': '1'},
        {'feild': 'coverage_type', 'feildName': '覆盖类型', 'datatype': '1'},
        {'feild': 'network_type', 'feildName': '网络制式', 'datatype': '1'},
        {'feild': 'freq', 'feildName': '频段_无线', 'datatype': '1'},
    ]
    
    # 构建result字段（datatype需要根据字段类型设置，前5个字段用'1'，其他用'character varying'）
    # 根据浏览器HAR分析：starttime, city, cgi, grid, area 使用 datatype='1'，其他用 'character varying'
    fixed_datatype_fields = {'starttime', 'city', 'cgi', 'grid', 'area'}
    result_list = []
    for f in result_fields:
        # 字段类型：前5个字段用'1'，其他用'character varying'（与浏览器请求一致）
        field_datatype = '1' if f['feild'] in fixed_datatype_fields else 'character varying'
        result_list.append({
            'feildtype': 'VONR小区监控预警数据表-天',
            'table': 'csem.f_nk_vonr_keykpi_cell_d',
            'tableName': 'VONR小区监控预警数据表-天',
            'datatype': field_datatype,
            'columntype': 1,
            'feildName': f['feildName'],
            'feild': f['feild'],
            'poly': '无',
            'anyWay': '无',
            'chart': '无',
            'chartpoly': '无'
        })
    
    payload = {
        'draw': 1,
        'start': 0,
        'length': 200,
        'total': 0,
        'geographicdimension': '小区',
        'timedimension': '天',
        'enodebField': 'gnodeb_id',
        'cgiField': 'cgi',
        'timeField': 'starttime',
        'cellField': 'cell',
        'cityField': 'city',
        'columns': _build_columns_param(vonr_fields),
        'order': [{'column': 0, 'dir': 'desc'}],
        'search': {'value': '', 'regex': False},
        'result': {'result': result_list, 'tableParams': {'supporteddimension': None, 'supportedtimedimension': ''}, 'columnname': ''},
        'where': [
            {'datatype': 'timestamp', 'feild': 'starttime', 'feildName': '', 'symbol': '>=', 'val': '2026-04-19 00:00:00', 'whereCon': 'and', 'query': True},
            {'datatype': 'timestamp', 'feild': 'starttime', 'feildName': '', 'symbol': '<', 'val': '2026-04-19 23:59:59', 'whereCon': 'and', 'query': True},
            {'datatype': 'character', 'feild': 'city', 'feildName': '', 'symbol': 'in', 'val': '阳江', 'whereCon': 'and', 'query': True}
        ],
        'indexcount': 0
    }
    return payload


def get_4g_wanchenglv_payload():
    """获取4G全程完好率报表payload"""
    print("[DEBUG-PAYLOAD] 生成 4G全程完好率报表 payload")
    
    # 4G全程完好率报表字段（基于HAR分析：appdbv3.a_common_pm_lte）
    fields = [
        'starttime', 'cgi', 'cell_name', 'city', 'branch', 'network_type', 'state', 'cover_type',
        'succconnestab', 'attconnestab', 'nbrsuccestab', 'nbrattestab',
        'succexecinc', 'ho_succ_out', 'ho_att__out',
        'hofail', 'nbrreqrelenb_normal', 'nbrreqrelenb', 'nbrleft', 'nbrhoinc'
    ]
    
    # 字段中文名映射
    field_names = {
        'starttime': '开始时间', 'cgi': 'CGI', 'cell_name': '小区名称', 'city': '所属地市',
        'branch': '人力区县分公司', 'network_type': '网络制式', 'state': '网元状态', 'cover_type': '覆盖类型',
        'succconnestab': 'RRC连接建立成功次数', 'attconnestab': 'RRC连接建立请求次数',
        'nbrsuccestab': 'E-RAB建立成功数', 'nbrattestab': 'E-RAB建立请求数',
        'succexecinc': '切换入成功次数', 'ho_succ_out': '切换成功次数', 'ho_att__out': '切换请求次数',
        'hofail': '切出失败的E-RAB数', 'nbrreqrelenb_normal': '正常的eNB请求释放的E-RAB数',
        'nbrreqrelenb': 'eNB请求释放的E-RAB数', 'nbrleft': '遗留上下文个数', 'nbrhoinc': '切换入E-RAB数'
    }
    
    # 构建result字段
    result_list = []
    # 时间维度字段需要 columntype=2，其他字段 columntype=1
    time_fields = {'starttime'}
    fixed_datatype_fields = {'starttime', 'city', 'cgi', 'branch', 'network_type', 'state', 'cover_type'}
    for f in fields:
        field_datatype = '1' if f in fixed_datatype_fields else 'character varying'
        # starttime 使用 columntype=2（时间维度）
        columntype = 2 if f in time_fields else 1
        result_list.append({
            'feildtype': '公共信息（小区级粒度）',
            'table': 'appdbv3.a_common_pm_lte',
            'tableName': '4G小区性能KPI报表',
            'datatype': field_datatype,
            'columntype': columntype,
            'feildName': field_names.get(f, f),
            'feild': f,
            'poly': '无',
            'anyWay': '无',
            'chart': '无',
            'chartpoly': '无'
        })
    
    payload = {
        'draw': 1,
        'start': 0,
        'length': 200,
        'total': 0,
        'geographicdimension': '小区，网格，地市，分公司',
        'timedimension': '小时,天,周.月,忙时,15分钟',
        'enodebField': 'enodeb_id',
        'cgiField': 'cgi',
        'timeField': 'starttime',
        'cellField': 'cell',
        'cityField': 'city',
        'columns': _build_columns_param(fields),
        'order': [{'column': 0, 'dir': 'desc'}],
        'search': {'value': '', 'regex': False},
        'result': {'result': result_list, 'tableParams': {'supporteddimension': '0', 'supportedtimedimension': '1'}, 'columnname': ''},
        'where': [
            {'datatype': 'timestamp', 'feild': 'starttime', 'feildName': '', 'symbol': '>=', 'val': '2026-04-19 00:00:00', 'whereCon': 'and', 'query': True},
            {'datatype': 'timestamp', 'feild': 'starttime', 'feildName': '', 'symbol': '<', 'val': '2026-04-19 23:59:59', 'whereCon': 'and', 'query': True},
            {'datatype': 'character', 'feild': 'city', 'feildName': '', 'symbol': 'in', 'val': '阳江', 'whereCon': 'and', 'query': True}
        ],
        'indexcount': 0
    }
    print(f"[DEBUG-PAYLOAD] 4G全程完好率报表 payload 生成完成，表名: appdbv3.a_common_pm_lte")
    return payload


def get_5g_wanchenglv_payload():
    """获取5G全程完好率报表payload"""
    print("[DEBUG-PAYLOAD] 生成 5G全程完好率报表 payload")
    
    # 5G全程完好率报表字段（基于HAR分析：appdbv3.a_common_pm_sacu）
    fields = [
        'starttime', 'ncgi', 'nrcell_name', 'branch', 'grid', 'grid_road', 'city', 'area',
        'vendor', 'network_type', 'cover_type', 'state',
        'rrc_succconnestab', 'rrc_attconnestab',
        'flow_nbrsuccestab', 'flow_nbrattestab',
        'ngsig_connestabsucc', 'ngsig_connestabatt',
        'context_attrelgnb', 'context_attrelgnb_normal',
        'context_succinitalsetup', 'context_nbrleft',
        'ho_succexecinc', 'rrc_succconnreestab_nonsrccell',
        'ho_succoutintercung', 'ho_succoutintercuxn',
        'ho_succoutintracuinterdu', 'ho_succoutintradu',
        'ho_attoutintercung', 'ho_attoutintercuxn',
        'ho_attoutintracuinterdu', 'ho_attoutcuintradu'
    ]
    
    # 字段中文名映射
    field_names = {
        'starttime': '数据时间', 'ncgi': 'NCGI', 'nrcell_name': '小区名称',
        'branch': '人力区县分公司', 'grid': '责任网格', 'grid_road': '路测网格',
        'city': '所属地市', 'area': '所属区县', 'vendor': '设备厂家',
        'network_type': '网络制式', 'cover_type': '覆盖类型', 'state': '网元状态',
        'rrc_succconnestab': 'RRC连接建立成功次数', 'rrc_attconnestab': 'RRC连接建立请求次数',
        'flow_nbrsuccestab': 'Flow建立成功数', 'flow_nbrattestab': 'Flow建立请求数',
        'ngsig_connestabsucc': 'NG接口UE相关逻辑信令连接建立成功次数',
        'ngsig_connestabatt': 'NG接口UE相关逻辑信令连接建立请求次数',
        'context_attrelgnb': 'gNB请求释放上下文数', 'context_attrelgnb_normal': '正常的gNB请求释放上下文数',
        'context_succinitalsetup': '初始上下文建立成功次数', 'context_nbrleft': '遗留上下文个数',
        'ho_succexecinc': '切换入成功次数', 'rrc_succconnreestab_nonsrccell': 'RRC连接重建成功次数(非源侧小区)',
        'ho_succoutintercung': 'gNB间NG切换出成功次数', 'ho_succoutintercuxn': 'gNB间Xn切换出成功次数',
        'ho_succoutintracuinterdu': 'CU内DU间切换出执行成功次数', 'ho_succoutintradu': 'CU内DU内切换出成功次数',
        'ho_attoutintercung': 'gNB间NG切换出准备请求次数', 'ho_attoutintercuxn': 'gNB间Xn切换出准备请求次数',
        'ho_attoutintracuinterdu': 'CU内DU间切换出执行请求次数', 'ho_attoutcuintradu': 'CU内DU内切换出执行请求次数'
    }
    
    # 构建result字段
    result_list = []
    # 时间维度字段需要 columntype=2，其他字段 columntype=1
    time_fields = {'starttime'}
    fixed_datatype_fields = {'starttime', 'city', 'ncgi', 'nrcell_name', 'branch', 'grid', 'grid_road', 'area', 'vendor', 'network_type', 'cover_type', 'state'}
    for f in fields:
        field_datatype = '1' if f in fixed_datatype_fields else 'character varying'
        # starttime 使用 columntype=2（时间维度）
        columntype = 2 if f in time_fields else 1
        result_list.append({
            'feildtype': 'SA_CU性能',  # 修正：使用与HAR一致的feildtype
            'table': 'appdbv3.a_common_pm_sacu',
            'tableName': '5GSA_CU性能报表',  # 修正：使用与HAR一致的tableName
            'datatype': field_datatype,
            'columntype': columntype,
            'feildName': field_names.get(f, f),
            'feild': f,
            'poly': '无',
            'anyWay': '无',
            'chart': '无',
            'chartpoly': '无'
        })
    
    payload = {
        'draw': 1,
        'start': 0,
        'length': 200,
        'total': 0,
        'geographicdimension': '小区，网格，地市，分公司',
        'timedimension': '小时,天,周,月',
        'enodebField': 'gnodeb_id',
        'cgiField': 'ncgi',
        'timeField': 'starttime',
        'cellField': 'nrcell',
        'cityField': 'city',
        'columns': _build_columns_param(fields),
        'order': [{'column': 0, 'dir': 'desc'}],
        'search': {'value': '', 'regex': False},
        'result': {'result': result_list, 'tableParams': {'supporteddimension': '0', 'supportedtimedimension': '1'}, 'columnname': ''},
        'where': [
            {'datatype': 'timestamp', 'feild': 'starttime', 'feildName': '', 'symbol': '>=', 'val': '2026-04-19 00:00:00', 'whereCon': 'and', 'query': True},
            {'datatype': 'timestamp', 'feild': 'starttime', 'feildName': '', 'symbol': '<', 'val': '2026-04-19 23:59:59', 'whereCon': 'and', 'query': True},
            {'datatype': 'character', 'feild': 'city', 'feildName': '', 'symbol': 'in', 'val': '阳江', 'whereCon': 'and', 'query': True}
        ],
        'indexcount': 0
    }
    print(f"[DEBUG-PAYLOAD] 5G全程完好率报表 payload 生成完成，表名: appdbv3.a_common_pm_sacu")
    return payload


def get_volte_payload():
    """获取VoLTE小区监控预警报表payload"""
    print("[DEBUG-PAYLOAD] 生成 VoLTE小区监控预警 payload")
    
    # VoLTE字段（基于HAR日志中的实际字段）
    fields = [
        'starttime', 'city', 'cgi', 'grid', 'area', 'nrcell_name',
        'volte_ul_tunzi_len', 'volte_ul_dantong_len', 'volte_ul_duanxu_len', 'volte_ul_voice_sum_len',
        'volte_ans_voice_call', 'volte_dl_tunzi_len', 'volte_dl_dantong_len', 'volte_dl_duanxu_len', 'volte_dl_voice_sum_len'
    ]
    
    # 字段中文名映射
    field_names = {
        'starttime': '时间', 'city': '地市', 'cgi': '小区', 'grid': '责任网格', 'area': '区县', 'nrcell_name': '小区名称',
        'volte_ul_tunzi_len': 'VoLTE语音上行吞字时长(s)', 'volte_ul_dantong_len': 'VoLTE语音上行单通时长(s)',
        'volte_ul_duanxu_len': 'VoLTE语音上行断续时长(s)', 'volte_ul_voice_sum_len': 'VoLTE语音上行总时长(s)',
        'volte_ans_voice_call': 'VoLTE语音通话总次数',
        'volte_dl_tunzi_len': 'VoLTE语音下行吞字时长(s)', 'volte_dl_dantong_len': 'VoLTE语音下行单通时长(s)',
        'volte_dl_duanxu_len': 'VoLTE语音下行断续时长(s)', 'volte_dl_voice_sum_len': 'VoLTE语音下行总时长(s)'
    }
    
    # 构建result字段
    result_list = []
    fixed_datatype_fields = {'starttime', 'city', 'cgi', 'grid', 'area', 'nrcell_name'}
    for f in fields:
        field_datatype = '1' if f in fixed_datatype_fields else 'character varying'
        result_list.append({
            'feildtype': 'VoLTE小区监控预警数据表-天',
            'table': 'csem.f_nk_volte_keykpi_cell_d',
            'tableName': 'VoLTE小区监控预警数据表-天',
            'datatype': field_datatype,
            'columntype': 1,
            'feildName': field_names.get(f, f),
            'feild': f,
            'poly': '无',
            'anyWay': '无',
            'chart': '无',
            'chartpoly': '无'
        })
    
    payload = {
        'draw': 1,
        'start': 0,
        'length': 200,
        'total': 0,
        'geographicdimension': '小区',
        'timedimension': '天',
        'enodebField': 'enodeb_id',
        'cgiField': 'cgi',
        'timeField': 'starttime',
        'cellField': 'cell',
        'cityField': 'city',
        'columns': _build_columns_param(fields),
        'order': [{'column': 0, 'dir': 'desc'}],
        'search': {'value': '', 'regex': False},
        'result': {'result': result_list, 'tableParams': {'supporteddimension': None, 'supportedtimedimension': ''}, 'columnname': ''},
        'where': [
            {'datatype': 'timestamp', 'feild': 'starttime', 'feildName': '', 'symbol': '>=', 'val': '2026-04-19 00:00:00', 'whereCon': 'and', 'query': True},
            {'datatype': 'timestamp', 'feild': 'starttime', 'feildName': '', 'symbol': '<', 'val': '2026-04-19 23:59:59', 'whereCon': 'and', 'query': True},
            {'datatype': 'character', 'feild': 'city', 'feildName': '', 'symbol': 'in', 'val': '阳江', 'whereCon': 'and', 'query': True}
        ],
        'indexcount': 0
    }
    print(f"[DEBUG-PAYLOAD] VoLTE小区监控预警 payload 生成完成，表名: csem.f_nk_volte_keykpi_cell_d")
    return payload


def get_epsfb_payload():
    """获取EPSFB小区监控预警报表payload"""
    print("[DEBUG-PAYLOAD] 生成 EPSFB小区监控预警 payload")
    
    # EPSFB字段（基于HAR日志中的实际字段）
    fields = [
        'starttime', 'city', 'cgi', 'grid', 'area', 'nrcell_name',
        'epsfb_ul_tunzi_len', 'epsfb_ul_dantong_len', 'epsfb_ul_duanxu_len', 'epsfb_ul_voice_sum_len',
        'epsfb_ans_voice_call', 'epsfb_dl_tunzi_len', 'epsfb_dl_dantong_len', 'epsfb_dl_duanxu_len', 'epsfb_dl_voice_sum_len'
    ]
    
    # 字段中文名映射
    field_names = {
        'starttime': '时间', 'city': '地市', 'cgi': '小区', 'grid': '责任网格', 'area': '区县', 'nrcell_name': '小区名称',
        'epsfb_ul_tunzi_len': 'EPSFB语音上行吞字时长(s)', 'epsfb_ul_dantong_len': 'EPSFB语音上行单通时长(s)',
        'epsfb_ul_duanxu_len': 'EPSFB语音上行断续时长(s)', 'epsfb_ul_voice_sum_len': 'EPSFB语音上行总时长(s)',
        'epsfb_ans_voice_call': 'EPSFB语音通话总次数',
        'epsfb_dl_tunzi_len': 'EPSFB语音下行吞字时长(s)', 'epsfb_dl_dantong_len': 'EPSFB语音下行单通时长(s)',
        'epsfb_dl_duanxu_len': 'EPSFB语音下行断续时长(s)', 'epsfb_dl_voice_sum_len': 'EPSFB语音下行总时长(s)'
    }
    
    # 构建result字段
    result_list = []
    fixed_datatype_fields = {'starttime', 'city', 'cgi', 'grid', 'area', 'nrcell_name'}
    for f in fields:
        field_datatype = '1' if f in fixed_datatype_fields else 'character varying'
        result_list.append({
            'feildtype': 'EPSFB小区监控预警数据表-天',
            'table': 'csem.f_nk_epsfb_keykpi_cell_d',
            'tableName': 'EPSFB小区监控预警数据表-天',
            'datatype': field_datatype,
            'columntype': 1,
            'feildName': field_names.get(f, f),
            'feild': f,
            'poly': '无',
            'anyWay': '无',
            'chart': '无',
            'chartpoly': '无'
        })
    
    payload = {
        'draw': 1,
        'start': 0,
        'length': 200,
        'total': 0,
        'geographicdimension': '小区',
        'timedimension': '天',
        'enodebField': '---',  # EPSFB表使用---
        'cgiField': 'cgi',
        'timeField': 'starttime',
        'cellField': 'cell',
        'cityField': 'city',
        'columns': _build_columns_param(fields),
        'order': [{'column': 0, 'dir': 'desc'}],
        'search': {'value': '', 'regex': False},
        'result': {'result': result_list, 'tableParams': {'supporteddimension': None, 'supportedtimedimension': ''}, 'columnname': ''},
        'where': [
            {'datatype': 'timestamp', 'feild': 'starttime', 'feildName': '', 'symbol': '>=', 'val': '2026-04-19 00:00:00', 'whereCon': 'and', 'query': True},
            {'datatype': 'timestamp', 'feild': 'starttime', 'feildName': '', 'symbol': '<', 'val': '2026-04-19 23:59:59', 'whereCon': 'and', 'query': True},
            {'datatype': 'character', 'feild': 'city', 'feildName': '', 'symbol': 'in', 'val': '阳江', 'whereCon': 'and', 'query': True}
        ],
        'indexcount': 0
    }
    print(f"[DEBUG-PAYLOAD] EPSFB小区监控预警 payload 生成完成，表名: csem.f_nk_epsfb_keykpi_cell_d")
    return payload


def get_4g_voice_payload():
    """获取4G语音小区报表payload（VoLTE+EPSFB联合） - 分别查询后合并"""
    print("[DEBUG-PAYLOAD] 生成 4G语音小区报表 payload (联合模式)")
    
    # 返回None，表示需要特殊处理
    # 这个函数不再直接生成payload，而是标记需要两个表分别查询
    return None


def get_5g_voice_payload():
    """获取5G语音小区报表payload（VONR）"""
    print("[DEBUG-PAYLOAD] 生成 5G语音小区报表 payload")
    
    # VONR语音相关字段
    fields = [
        'starttime', 'city', 'cgi', 'grid', 'area', 'nrcell_name',
        'vonr_ul_tunzi_len', 'vonr_ul_dantong_len', 'vonr_ul_duanxu_len', 'vonr_ul_voice_sum_len',
        'vonr_ans_voice_call'
    ]
    
    # 字段中文名映射
    field_names = {
        'starttime': '时间', 'city': '地市', 'cgi': '小区', 'grid': '责任网格', 'area': '区县', 'nrcell_name': '小区名称',
        'vonr_ul_tunzi_len': 'VoNR语音上行吞字时长(s)', 'vonr_ul_dantong_len': 'VoNR语音上行单通时长(s)',
        'vonr_ul_duanxu_len': 'VoNR语音上行断续时长(s)', 'vonr_ul_voice_sum_len': 'VoNR语音上行总时长(s)',
        'vonr_ans_voice_call': 'VoNR语音通话总次数'
    }
    
    # 构建result字段
    result_list = []
    fixed_datatype_fields = {'starttime', 'city', 'cgi', 'grid', 'area', 'nrcell_name'}
    for f in fields:
        field_datatype = '1' if f in fixed_datatype_fields else 'character varying'
        result_list.append({
            'feildtype': '5G语音小区报表',
            'table': 'csem.f_nk_vonr_keykpi_cell_d',
            'tableName': '5G语音小区报表',
            'datatype': field_datatype,
            'columntype': 1,
            'feildName': field_names.get(f, f),
            'feild': f,
            'poly': '无',
            'anyWay': '无',
            'chart': '无',
            'chartpoly': '无'
        })
    
    payload = {
        'draw': 1,
        'start': 0,
        'length': 200,
        'total': 0,
        'geographicdimension': '小区',
        'timedimension': '天',
        'enodebField': 'gnodeb_id',
        'cgiField': 'cgi',
        'timeField': 'starttime',
        'cellField': 'cell',
        'cityField': 'city',
        'columns': _build_columns_param(fields),
        'order': [{'column': 0, 'dir': 'desc'}],
        'search': {'value': '', 'regex': False},
        'result': {'result': result_list, 'tableParams': {'supporteddimension': None, 'supportedtimedimension': ''}, 'columnname': ''},
        'where': [
            {'datatype': 'timestamp', 'feild': 'starttime', 'feildName': '', 'symbol': '>=', 'val': '2026-04-19 00:00:00', 'whereCon': 'and', 'query': True},
            {'datatype': 'timestamp', 'feild': 'starttime', 'feildName': '', 'symbol': '<', 'val': '2026-04-19 23:59:59', 'whereCon': 'and', 'query': True},
            {'datatype': 'character', 'feild': 'city', 'feildName': '', 'symbol': 'in', 'val': '阳江', 'whereCon': 'and', 'query': True}
        ],
        'indexcount': 0
    }
    print(f"[DEBUG-PAYLOAD] 5G语音小区报表 payload 生成完成")
    return payload


def add_wanchenglv_columns(df, report_type='4g'):
    """添加全程完好率计算列
    
    Args:
        df: DataFrame数据
        report_type: '4g' 或 '5g'
    """
    import numpy as np
    
    def safe_value(val):
        """如果值为0或nan，返回1，否则返回原值"""
        return np.where((pd.isna(val)) | (val == 0), 1, val)
    
    def safe_divide(numerator, denominator, fill_value=np.nan):
        """安全的除法运算，防止除零错误"""
        return np.where(denominator != 0, numerator / denominator, fill_value)
    
    def debug_column_stats(col_name, series, is_denominator=False):
        """打印列的详细统计信息，帮助诊断除零问题"""
        if isinstance(series, pd.Series):
            zero_count = (series == 0).sum()
            nan_count = series.isna().sum()
            total = len(series)
            print(f"  [DEBUG-WANCHENGLV] {col_name}: 总数={total}, 零值={zero_count}, NaN={nan_count}, 有效={total-zero_count-nan_count}")
            if (zero_count > 0 or nan_count > 0) and total <= 100:  # 数据量小时显示详情
                zero_or_nan_rows = series[(series == 0) | series.isna()].index.tolist()
                print(f"  [DEBUG-WANCHENGLV]   零值/NaN行: {zero_or_nan_rows}")
    
    print(f"[DEBUG-WANCHENGLV] ========== 开始计算{'4G' if report_type == '4g' else '5G'}全程完好率 ==========")
    print(f"[DEBUG-WANCHENGLV] DataFrame形状: {df.shape}")
    print(f"[DEBUG-WANCHENGLV] DataFrame列名: {list(df.columns)}")
    
    if report_type == '4g':
        # 4G全程完好率计算
        print("[DEBUG-WANCHENGLV] === 4G全程完好率计算 ===")
        
        # 检查所需列是否存在
        required_4g_cols = [
            'RRC连接建立请求次数', 'RRC连接建立成功次数', 'E-RAB建立请求数', 'E-RAB建立成功数',
            '遗留上下文个数', '切换入E-RAB数', '切出失败的E-RAB数', '正常的eNB请求释放的E-RAB数', 'eNB请求释放的E-RAB数',
            '切换请求次数', '切换成功次数'
        ]
        missing_4g = [c for c in required_4g_cols if c not in df.columns]
        if missing_4g:
            print(f"[ERROR-WANCHENGLV] 4G计算缺少列: {missing_4g}")
            return df
        
        # 4G无线接通率(%) = (RRC连接建立成功次数/RRC连接建立请求次数) * (E-RAB建立成功数/E-RAB建立请求数) * 100
        print("[DEBUG-WANCHENGLV] >> 计算4G无线接通率...")
        debug_column_stats('RRC连接建立请求次数', df['RRC连接建立请求次数'])
        debug_column_stats('RRC连接建立成功次数', df['RRC连接建立成功次数'])
        debug_column_stats('E-RAB建立请求数', df['E-RAB建立请求数'])
        debug_column_stats('E-RAB建立成功数', df['E-RAB建立成功数'])
        rrc_denom = df['RRC连接建立请求次数'].replace(0, np.nan)
        erab_denom = df['E-RAB建立请求数'].replace(0, np.nan)
        
        # 检查分母为0的行
        rrc_zero_rows = df[df['RRC连接建立请求次数'] == 0].index.tolist()
        erab_zero_rows = df[df['E-RAB建立请求数'] == 0].index.tolist()
        if rrc_zero_rows:
            print(f"[DEBUG-WANCHENGLV]   RRC连接建立请求次数=0的行数: {len(rrc_zero_rows)}, 前10个: {rrc_zero_rows[:10]}")
        if erab_zero_rows:
            print(f"[DEBUG-WANCHENGLV]   E-RAB建立请求数=0的行数: {len(erab_zero_rows)}, 前10个: {erab_zero_rows[:10]}")
        
        df['4G无线接通率(%)'] = np.where(
            (df['RRC连接建立请求次数'] > 0) & (df['E-RAB建立请求数'] > 0),
            safe_divide(df['RRC连接建立成功次数'], rrc_denom, 0) * 
            safe_divide(df['E-RAB建立成功数'], erab_denom, 0) * 100,
            np.nan
        )
        print(f"[DEBUG-WANCHENGLV]   4G无线接通率结果: 有效={df['4G无线接通率(%)'].notna().sum()}, NaN={df['4G无线接通率(%)'].isna().sum()}")
        
        # 4GE-RAB掉线率 = (切出失败的E-RAB数 - 正常的eNB请求释放的E-RAB数 + eNB请求释放的E-RAB数) /
        #                 (遗留上下文个数 + E-RAB建立成功数 + 切换入E-RAB数) * 100
        print("[DEBUG-WANCHENGLV] >> 计算4GE-RAB掉线率...")
        debug_column_stats('遗留上下文个数', df['遗留上下文个数'])
        debug_column_stats('E-RAB建立成功数', df['E-RAB建立成功数'])
        debug_column_stats('切换入E-RAB数', df['切换入E-RAB数'])
        erab_drop_denom = df['遗留上下文个数'] + df['E-RAB建立成功数'] + df['切换入E-RAB数']
        erab_drop_denom_safe = erab_drop_denom.replace(0, np.nan)
        
        drop_denom_zero_rows = df.index[erab_drop_denom == 0].tolist()
        if drop_denom_zero_rows:
            print(f"[DEBUG-WANCHENGLV]   4GE-RAB掉线率分母=0的行数: {len(drop_denom_zero_rows)}, 前10个: {drop_denom_zero_rows[:10]}")
            # 显示这些行的分母组成
            for row_idx in drop_denom_zero_rows[:3]:
                print(f"     行{row_idx}: 遗留上下文={df.loc[row_idx, '遗留上下文个数']}, E-RAB成功={df.loc[row_idx, 'E-RAB建立成功数']}, 切换入={df.loc[row_idx, '切换入E-RAB数']}")
        
        df['4GE-RAB掉线率'] = np.where(
            erab_drop_denom > 0,
            safe_divide(
                df['切出失败的E-RAB数'] - df['正常的eNB请求释放的E-RAB数'] + df['eNB请求释放的E-RAB数'],
                erab_drop_denom_safe, 0
            ) * 100,
            np.nan
        )
        print(f"[DEBUG-WANCHENGLV]   4GE-RAB掉线率结果: 有效={df['4GE-RAB掉线率'].notna().sum()}, NaN={df['4GE-RAB掉线率'].isna().sum()}")
        
        # 4G切换成功率(%) = 切换成功次数 / 切换请求次数 * 100
        print("[DEBUG-WANCHENGLV] >> 计算4G切换成功率...")
        debug_column_stats('切换请求次数', df['切换请求次数'])
        debug_column_stats('切换成功次数', df['切换成功次数'])
        
        ho_denom = df['切换请求次数'].replace(0, np.nan)
        ho_zero_rows = df[df['切换请求次数'] == 0].index.tolist()
        if ho_zero_rows:
            print(f"[DEBUG-WANCHENGLV]   切换请求次数=0的行数: {len(ho_zero_rows)}, 前10个: {ho_zero_rows[:10]}")
        
        df['4G切换成功率(%)'] = np.where(
            df['切换请求次数'] > 0,
            safe_divide(df['切换成功次数'], ho_denom, 0) * 100,
            np.nan
        )
        print(f"[DEBUG-WANCHENGLV]   4G切换成功率结果: 有效={df['4G切换成功率(%)'].notna().sum()}, NaN={df['4G切换成功率(%)'].isna().sum()}")
        
        # 4G全程完好率 = (4G无线接通率 / 100) * (4G切换成功率 / 100) * (100 - 4GE-RAB掉线率) * 100
        # 简化为: 4G无线接通率 * 4G切换成功率 * (100 - 4GE-RAB掉线率) / 10000
        # 当任一分量为0或nan时，默认视为1，避免除0错误
        print("[DEBUG-WANCHENGLV] >> 计算4G全程完好率...")
        wanchenglv_factor = safe_value(df['4G无线接通率(%)']) * safe_value(df['4G切换成功率(%)']) * safe_value(100 - df['4GE-RAB掉线率'])
        
        # 检查是否有inf或nan
        inf_count = np.isinf(wanchenglv_factor).sum()
        nan_count_factor = np.isnan(wanchenglv_factor).sum()
        print(f"[DEBUG-WANCHENGLV]   wanchenglv_factor统计: inf={inf_count}, nan={nan_count_factor}")
        if inf_count > 0:
            inf_rows = df.index[np.isinf(wanchenglv_factor)].tolist()
            print(f"[DEBUG-WANCHENGLV]   !!! 发现inf值! 行: {inf_rows[:10]}")
        
        # 当所有中间指标都是NaN时，最终结果也应该是NaN
        all_valid = df['4G无线接通率(%)'].notna() & df['4G切换成功率(%)'].notna() & df['4GE-RAB掉线率'].notna()
        invalid_rows = df.index[~all_valid].tolist()
        if invalid_rows:
            print(f"[DEBUG-WANCHENGLV]   中间指标有NaN的行数: {len(invalid_rows)}")
        
        df['4G全程完好率'] = np.where(all_valid, wanchenglv_factor / 10000, np.nan)
        
        # 最终检查
        final_inf = np.isinf(df['4G全程完好率']).sum()
        final_nan = df['4G全程完好率'].isna().sum()
        final_valid = df['4G全程完好率'].notna().sum()
        print(f"[DEBUG-WANCHENGLV]   4G全程完好率最终结果: 有效={final_valid}, NaN={final_nan}, inf={final_inf}")
        
    else:
        # 5G全程完好率计算
        print("[DEBUG-WANCHENGLV] === 5G全程完好率计算 ===")
        
        # 检查所需列是否存在
        required_5g_cols = [
            'RRC连接建立请求次数', 'RRC连接建立成功次数',
            'Flow建立请求数', 'Flow建立成功数',
            'NG接口UE相关逻辑信令连接建立请求次数', 'NG接口UE相关逻辑信令连接建立成功次数',
            '初始上下文建立成功次数', '遗留上下文个数', '切换入成功次数', 'RRC连接重建成功次数(非源侧小区)',
            'gNB请求释放上下文数', '正常的gNB请求释放上下文数',
            'gNB间NG切换出成功次数', 'gNB间Xn切换出成功次数', 'CU内DU间切换出执行成功次数', 'CU内DU内切换出成功次数',
            'gNB间NG切换出准备请求次数', 'gNB间Xn切换出准备请求次数', 'CU内DU间切换出执行请求次数', 'CU内DU内切换出执行请求次数'
        ]
        missing_5g = [c for c in required_5g_cols if c not in df.columns]
        if missing_5g:
            print(f"[ERROR-WANCHENGLV] 5G计算缺少列: {missing_5g}")
            return df
        # SA无线接通率% = (RRC连接建立成功次数/RRC连接建立请求次数) *
        #                (Flow建立成功数/Flow建立请求数) *
        #                (NG接口UE相关逻辑信令连接建立成功次数/NG接口UE相关逻辑信令连接建立请求次数) * 100
        rrc_denom = df['RRC连接建立请求次数'].replace(0, np.nan)
        flow_denom = df['Flow建立请求数'].replace(0, np.nan)
        ng_denom = df['NG接口UE相关逻辑信令连接建立请求次数'].replace(0, np.nan)
        
        # 检查分母为0的行
        rrc_zero_rows = df[df['RRC连接建立请求次数'] == 0].index.tolist()
        flow_zero_rows = df[df['Flow建立请求数'] == 0].index.tolist()
        ng_zero_rows = df[df['NG接口UE相关逻辑信令连接建立请求次数'] == 0].index.tolist()
        if rrc_zero_rows:
            print(f"[DEBUG-WANCHENGLV]   RRC连接建立请求次数=0的行数: {len(rrc_zero_rows)}, 前10个: {rrc_zero_rows[:10]}")
        if flow_zero_rows:
            print(f"[DEBUG-WANCHENGLV]   Flow建立请求数=0的行数: {len(flow_zero_rows)}, 前10个: {flow_zero_rows[:10]}")
        if ng_zero_rows:
            print(f"[DEBUG-WANCHENGLV]   NG接口UE相关逻辑信令连接建立请求次数=0的行数: {len(ng_zero_rows)}, 前10个: {ng_zero_rows[:10]}")
        
        df['SA无线接通率%'] = np.where(
            (df['RRC连接建立请求次数'] > 0) & (df['Flow建立请求数'] > 0) & (df['NG接口UE相关逻辑信令连接建立请求次数'] > 0),
            safe_divide(df['RRC连接建立成功次数'], rrc_denom, 0) *
            safe_divide(df['Flow建立成功数'], flow_denom, 0) *
            safe_divide(df['NG接口UE相关逻辑信令连接建立成功次数'], ng_denom, 0) * 100,
            np.nan
        )
        print(f"[DEBUG-WANCHENGLV]   SA无线接通率结果: 有效={df['SA无线接通率%'].notna().sum()}, NaN={df['SA无线接通率%'].isna().sum()}")
        
        # SA无线掉线率% = (gNB请求释放上下文数 - 正常的gNB请求释放上下文数) /
        #                (初始上下文建立成功次数 + 遗留上下文个数 + 切换入成功次数 + RRC连接重建成功次数(非源侧小区)) * 100
        print("[DEBUG-WANCHENGLV] >> 计算SA无线掉线率...")
        debug_column_stats('初始上下文建立成功次数', df['初始上下文建立成功次数'])
        debug_column_stats('遗留上下文个数', df['遗留上下文个数'])
        debug_column_stats('切换入成功次数', df['切换入成功次数'])
        debug_column_stats('RRC连接重建成功次数(非源侧小区)', df['RRC连接重建成功次数(非源侧小区)'])
        debug_column_stats('gNB请求释放上下文数', df['gNB请求释放上下文数'])
        debug_column_stats('正常的gNB请求释放上下文数', df['正常的gNB请求释放上下文数'])
        
        drop_denom = (df['初始上下文建立成功次数'] + df['遗留上下文个数'] + 
                      df['切换入成功次数'] + df['RRC连接重建成功次数(非源侧小区)'])
        drop_denom_safe = drop_denom.replace(0, np.nan)
        
        drop_denom_zero_rows = df.index[drop_denom == 0].tolist()
        if drop_denom_zero_rows:
            print(f"[DEBUG-WANCHENGLV]   SA无线掉线率分母=0的行数: {len(drop_denom_zero_rows)}, 前10个: {drop_denom_zero_rows[:10]}")
            for row_idx in drop_denom_zero_rows[:3]:
                print(f"     行{row_idx}: 初始上下文={df.loc[row_idx, '初始上下文建立成功次数']}, 遗留={df.loc[row_idx, '遗留上下文个数']}, 切换入={df.loc[row_idx, '切换入成功次数']}, 重建立={df.loc[row_idx, 'RRC连接重建成功次数(非源侧小区)']}")
        
        df['SA无线掉线率%'] = np.where(
            drop_denom > 0,
            safe_divide(
                df['gNB请求释放上下文数'] - df['正常的gNB请求释放上下文数'],
                drop_denom_safe, 0
            ) * 100,
            np.nan
        )
        print(f"[DEBUG-WANCHENGLV]   SA无线掉线率结果: 有效={df['SA无线掉线率%'].notna().sum()}, NaN={df['SA无线掉线率%'].isna().sum()}")
        
        # SA切换成功率% = (gNB间NG切换出成功次数 + gNB间Xn切换出成功次数 + 
        #                  CU内DU间切换出执行成功次数 + CU内DU内切换出成功次数) /
        #                 (gNB间NG切换出准备请求次数 + gNB间Xn切换出准备请求次数 + 
        #                  CU内DU间切换出执行请求次数 + CU内DU内切换出执行请求次数) * 100
        print("[DEBUG-WANCHENGLV] >> 计算SA切换成功率...")
        debug_column_stats('gNB间NG切换出成功次数', df['gNB间NG切换出成功次数'])
        debug_column_stats('gNB间Xn切换出成功次数', df['gNB间Xn切换出成功次数'])
        debug_column_stats('CU内DU间切换出执行成功次数', df['CU内DU间切换出执行成功次数'])
        debug_column_stats('CU内DU内切换出成功次数', df['CU内DU内切换出成功次数'])
        debug_column_stats('gNB间NG切换出准备请求次数', df['gNB间NG切换出准备请求次数'])
        debug_column_stats('gNB间Xn切换出准备请求次数', df['gNB间Xn切换出准备请求次数'])
        debug_column_stats('CU内DU间切换出执行请求次数', df['CU内DU间切换出执行请求次数'])
        debug_column_stats('CU内DU内切换出执行请求次数', df['CU内DU内切换出执行请求次数'])
        
        success_sum = (df['gNB间NG切换出成功次数'] + df['gNB间Xn切换出成功次数'] + 
                      df['CU内DU间切换出执行成功次数'] + df['CU内DU内切换出成功次数'])
        attempt_sum = (df['gNB间NG切换出准备请求次数'] + df['gNB间Xn切换出准备请求次数'] + 
                      df['CU内DU间切换出执行请求次数'] + df['CU内DU内切换出执行请求次数'])
        attempt_sum_safe = attempt_sum.replace(0, np.nan)
        
        attempt_zero_rows = df.index[attempt_sum == 0].tolist()
        if attempt_zero_rows:
            print(f"[DEBUG-WANCHENGLV]   SA切换成功率分母=0的行数: {len(attempt_zero_rows)}, 前10个: {attempt_zero_rows[:10]}")
            for row_idx in attempt_zero_rows[:3]:
                print(f"     行{row_idx}: gNB间NG={df.loc[row_idx, 'gNB间NG切换出准备请求次数']}, gNB间Xn={df.loc[row_idx, 'gNB间Xn切换出准备请求次数']}, CU内DU间={df.loc[row_idx, 'CU内DU间切换出执行请求次数']}, CU内DU内={df.loc[row_idx, 'CU内DU内切换出执行请求次数']}")
        
        df['SA切换成功率%'] = np.where(
            attempt_sum > 0,
            safe_divide(success_sum, attempt_sum_safe, 0) * 100,
            np.nan
        )
        print(f"[DEBUG-WANCHENGLV]   SA切换成功率结果: 有效={df['SA切换成功率%'].notna().sum()}, NaN={df['SA切换成功率%'].isna().sum()}")
        
        # 5G全程完好率 = (SA无线接通率 / 100) * (SA切换成功率 / 100) * (100 - SA无线掉线率) * 100
        # 简化为: SA无线接通率 * SA切换成功率 * (100 - SA无线掉线率) / 10000
        # 当任一分量为0或nan时，默认视为1，避免除0错误
        print("[DEBUG-WANCHENGLV] >> 计算5G全程完好率...")
        wanchenglv_factor = safe_value(df['SA无线接通率%']) * safe_value(df['SA切换成功率%']) * safe_value(100 - df['SA无线掉线率%'])
        
        # 检查是否有inf或nan
        inf_count = np.isinf(wanchenglv_factor).sum()
        nan_count_factor = np.isnan(wanchenglv_factor).sum()
        print(f"[DEBUG-WANCHENGLV]   wanchenglv_factor统计: inf={inf_count}, nan={nan_count_factor}")
        if inf_count > 0:
            inf_rows = df.index[np.isinf(wanchenglv_factor)].tolist()
            print(f"[DEBUG-WANCHENGLV]   !!! 发现inf值! 行: {inf_rows[:10]}")
        
        # 当所有中间指标都是NaN时，最终结果也应该是NaN
        all_valid = df['SA无线接通率%'].notna() & df['SA切换成功率%'].notna() & df['SA无线掉线率%'].notna()
        invalid_rows = df.index[~all_valid].tolist()
        if invalid_rows:
            print(f"[DEBUG-WANCHENGLV]   中间指标有NaN的行数: {len(invalid_rows)}")
        
        df['5G全程完好率'] = np.where(all_valid, wanchenglv_factor / 10000, np.nan)
        
        # 最终检查
        final_inf = np.isinf(df['5G全程完好率']).sum()
        final_nan = df['5G全程完好率'].isna().sum()
        final_valid = df['5G全程完好率'].notna().sum()
        print(f"[DEBUG-WANCHENGLV]   5G全程完好率最终结果: 有效={final_valid}, NaN={final_nan}, inf={final_inf}")
    
    print("[DEBUG-WANCHENGLV] ========== 计算完成 ==========")
    return df


def add_4g_voice_columns(df):
    """添加4G语音小区计算列（VoLTE+EPSFB联合）"""
    import numpy as np
    
    # 4G语音通话质差时长比例 = 
    # (VoLTE语音上行吞字时长 + VoLTE语音上行单通时长 + VoLTE语音上行断续时长 +
    #  EPSFB语音上行吞字时长 + EPSFB语音上行单通时长 + EPSFB语音上行断续时长) /
    # (VoLTE语音上行总时长 + EPSFB语音上行总时长)
    total_bad = (df['VoLTE语音上行吞字时长(s)'].fillna(0) + df['VoLTE语音上行单通时长(s)'].fillna(0) + 
                  df['VoLTE语音上行断续时长(s)'].fillna(0) + df['EPSFB语音上行吞字时长(s)'].fillna(0) + 
                  df['EPSFB语音上行单通时长(s)'].fillna(0) + df['EPSFB语音上行断续时长(s)'].fillna(0))
    total_sum = df['VoLTE语音上行总时长(s)'].fillna(0) + df['EPSFB语音上行总时长(s)'].fillna(0)
    
    df['4G语音通话质差时长比例'] = np.where(total_sum > 0, total_bad / total_sum, np.nan)
    
    # 4G差小区判定：(VoLTE语音通话质差时长比例>2%且VoLTE通话次数>1000) 或者 (EPSFB语音通话质差时长比例>2%且EPSFB通话次数>1000)
    volte_bad_ratio = np.where(
        df['VoLTE语音上行总时长(s)'].fillna(0) > 0,
        (df['VoLTE语音上行吞字时长(s)'].fillna(0) + df['VoLTE语音上行单通时长(s)'].fillna(0) + df['VoLTE语音上行断续时长(s)'].fillna(0)) /
        df['VoLTE语音上行总时长(s)'].fillna(0),
        0
    )
    epsfb_bad_ratio = np.where(
        df['EPSFB语音上行总时长(s)'].fillna(0) > 0,
        (df['EPSFB语音上行吞字时长(s)'].fillna(0) + df['EPSFB语音上行单通时长(s)'].fillna(0) + df['EPSFB语音上行断续时长(s)'].fillna(0)) /
        df['EPSFB语音上行总时长(s)'].fillna(0),
        0
    )
    
    df['4G差小区'] = ((volte_bad_ratio > 0.02) & (df['VoLTE语音通话总次数'] > 1000)) | \
                     ((epsfb_bad_ratio > 0.02) & (df['EPSFB语音通话总次数'] > 1000))
    
    return df


def add_5g_voice_columns(df):
    """添加5G语音小区计算列（VONR）"""
    import numpy as np
    
    # 5G语音通话质差时长比例 = (VoNR语音上行吞字时长 + VoNR语音上行单通时长 + VoNR语音上行断续时长) / VoNR语音上行总时长
    total_bad = (df['VoNR语音上行吞字时长(s)'].fillna(0) + df['VoNR语音上行单通时长(s)'].fillna(0) + 
                 df['VoNR语音上行断续时长(s)'].fillna(0))
    total_sum = df['VoNR语音上行总时长(s)'].fillna(0)
    
    df['5G语音通话质差时长比例'] = np.where(total_sum > 0, total_bad / total_sum, np.nan)
    
    # 5G差小区：5G语音通话质差时长比例>2% 且 5G通话次数>1000次
    df['5G差小区'] = (df['5G语音通话质差时长比例'] > 0.02) & (df['VoNR语音通话总次数'] > 1000)
    
    return df


def get_5g_kpi_payload():
    """获取5G小区性能KPI报表payload（硬编码，字段与浏览器一致）"""
    print("[DEBUG-PAYLOAD] 生成 5G小区性能KPI报表 payload")
    
    # 5G小区KPI字段列表（与浏览器请求一致，共160个字段）
    kpi_fields = [
        'starttime', 'ncgi', 'nrcell_name', 'branch', 'grid', 'city', 'area', 'vendor', 'cover_type',
        'rrc_connmean', 'rrc_connmax', 'rrc_attconnestab', 'rrc_succconnestab', 'kpi_rrcsuccconnrate',
        'flow_nbrattestab', 'flow_nbrsuccestab', 'kpi_flowsuccconnrate',
        'ngsig_connestabatt', 'ngsig_connestabsucc', 'kpi_ngsig_succconnrate', 'kpi_wirelesssuccconnrate',
        'context_attrelgnb', 'context_attrelgnb_normal', 'context_succinitalsetup', 'context_nbrleft',
        'ho_succexecinc', 'rrc_succconnreestab_nonsrccell', 'kpi_wirelessdroprate_celllevel',
        'flow_nbrreqrelgnb', 'flow_nbrreqrelgnb_normal', 'flow_hoadmitfail', 'flow_nbrleft', 'flow_nbrhoinc',
        'kpi_flowdroprate_celllevel', 'rrc_attconnreestab', 'kpi_rrcconnreestabrate',
        'ho_attoutintercung', 'ho_succoutintercung', 'kpi_hosuccoutintergnbrate_ng',
        'ho_attoutintercuxn', 'ho_succoutintercuxn', 'kpi_hosuccoutintergnbrate_xn', 'kpi_hosuccoutintergnbrate',
        'ho_attoutintracuinterdu', 'ho_succoutintracuinterdu', 'ho_attoutcuintradu', 'ho_succoutintradu',
        'kpi_hosuccoutintragnbrate', 'kpi_hosuccoutrate',
        'ho_attoutexecintrafreq', 'ho_succoutintrafreq', 'kpi_hosuccoutrate_intrafreq',
        'ho_attoutexecinterfreq', 'ho_succoutinterfreq', 'kpi_hosuccoutrate_interfreq',
        'kpi_pdcpupoctul', 'kpi_pdcpupoctdl', 'ee_carriershutdowntime',
        'flow_nbrattestab_5qi1', 'flow_nbrsuccestab_5qi1', 'kpi_wirelesssuccconnrate_5qi1',
        'flow_nbrreqrelgnb_5qi1', 'flow_nbrreqrelgnb_normal_5qi1', 'flow_hoadmitfail_5qi1',
        'flow_nbrleft_5qi1', 'flow_nbrhoinc_5qi1', 'kpi_wirelessdroprate_celllevel_5qi1',
        'kpi_wirelessdroprate_netlevel_5qi1', 'pdcp_upoctul_5qi1', 'pdcp_upoctdl_5qi1',
        'pdcp_nbrpktlossul_5qi1', 'pdcp_nbrpktul_5qi1', 'kpi_pdcpnbrpktlossrateul_5qi1',
        'flow_nbrattestab_5qi2', 'flow_nbrsuccestab_5qi2', 'kpi_wirelesssuccconnrate_5qi2',
        'flow_nbrreqrelgnb_5qi2', 'flow_nbrreqrelgnb_normal_5qi2', 'flow_hoadmitfail_5qi2',
        'flow_nbrleft_5qi2', 'flow_nbrhoinc_5qi2', 'kpi_wirelessdroprate_celllevel_5qi2',
        'kpi_wirelessdroprate_netlevel_5qi2', 'pdcp_upoctul_5qi2', 'pdcp_upoctdl_5qi2',
        'pdcp_nbrpktlossul_5qi2', 'pdcp_nbrpktul_5qi2', 'kpi_pdcpnbrpktlossrateul_5qi2',
        'vonr_voice_traffic', 'vinr_voice_traffic',
        'iratho_succouteutran_epsfallback', 'iratho_succprepouteutran_epsfallback',
        'flow_nbrattestab_epsfb', 'iratho_attouteutran_epsfallback',
        'kpi_vonr_flowsuccconnrate', 'kpi_vinr_flowsuccconnrate',
        'kpi_vonr_flowdroprate_celllevel', 'kpi_vinr_flowdroprate_celllevel',
        'kpi_wirelessdroprate_netlevel', 'kpi_flowdroprate_netlevel', 'kpi_vonr_flowdroprate_netlevel',
        'kpi_hosuccoutrate_intersystemnrtolte', 'kpi_hosuccoutrate_intersystemltetonr',
        'kpi_esfbhosuccoutrate_intersystemnrtolte', 'rrc_redirecttolte_epsfallback',
        'iratho_succprepouteutran', 'iratho_attouteutran', 'iratho_attprepinc', 'iratho_succprepinc',
        'iratho_succouteutran', 'kpi_vonr_succconnrate', 'kpi_vinr_succconnrate',
        'kpi_slice_flowsuccconnrate', 'kpi_inactive_succconnrate',
        'kpi_hosuccoutrate_vonrtolte', 'kpi_hosuccoutrate_vinrtolte',
        'kpi_hosuccoutrate_intersystemvoltetovonr', 'kpi_hosuccoutrate_intersystemviltetovinr',
        'kpi_hosuccoutrate_vonr', 'kpi_hosuccoutrate_vinr',
        'kpi_pdcpnbrpktlossrateul', 'kpi_vonrtraffic_5qi1', 'kpi_vinrtraffic_5qi1',
        'flow_nbrsuccestab_vonr', 'flow_nbrattestab_vonr',
        'flow_nbrsuccestab_vinr', 'flow_nbrattestab_vinr',
        'flow_nbrsuccestabslice', 'flow_nbrattestabslice',
        'rrc_succconnresume', 'rrc_attconnresume',
        'iratho_succouteutran_vonr', 'iratho_attouteutran_vonr',
        'iratho_succouteutran_vinr', 'iratho_attouteutran_vinr',
        'iratho_succexecinc_voltetovonr', 'iratho_attprepinc_voltetovonr',
        'iratho_succexecinc_viltetovinr', 'iratho_attprepinc_viltetovinr',
        'ho_succoutintercung_vonr', 'ho_succoutintercuxn_vonr', 'ho_succoutintradu_vonr',
        'ho_attoutintercung_vonr', 'ho_attoutintercuxn_vonr', 'ho_attprepoutcuintradu_vonr',
        'ho_succoutintercung_vinr', 'ho_succoutintercuxn_vinr', 'ho_succoutintradu_vinr',
        'ho_attoutintercung_vinr', 'ho_attoutintercuxn_vinr', 'ho_attprepoutcuintradu_vinr',
        'pdcp_nbrpktlossul', 'pdcp_nbrpktul',
        'flow_nbrmeanestab_5qi1', 'flow_nbrmeanestab_5qi2',
        'kpi_wirelesssuccconnrate_v1_8', 'kpi_vonr_flowdroprate_netlevel_v1_8',
        'kpi_hosuccoutrate_intersystemltetonr_v1_8', 'flow_succestab_resume_vonr', 'iratho_succexecinc'
    ]
    
    # 构建result字段配置（与浏览器请求一致）
    result_fields = [
        {'feild': 'starttime', 'feildName': '数据时间', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'ncgi', 'feildName': 'NCGI', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'nrcell_name', 'feildName': '小区名称', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'branch', 'feildName': '人力区县分公司', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'grid', 'feildName': '责任网格', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'city', 'feildName': '所属地市', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'area', 'feildName': '所属区县', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'vendor', 'feildName': '设备厂家', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'cover_type', 'feildName': '覆盖类型', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'rrc_connmean', 'feildName': 'RRC连接平均数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'rrc_connmax', 'feildName': 'RRC连接最大数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'rrc_attconnestab', 'feildName': 'RRC连接建立请求次数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'rrc_succconnestab', 'feildName': 'RRC连接建立成功次数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'kpi_rrcsuccconnrate', 'feildName': 'RRC连接建立成功率', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'flow_nbrattestab', 'feildName': 'Flow建立请求数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'flow_nbrsuccestab', 'feildName': 'Flow建立成功数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'kpi_flowsuccconnrate', 'feildName': 'QoS Flow建立成功率', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'ngsig_connestabatt', 'feildName': 'NG接口UE相关逻辑信令连接建立请求次数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'ngsig_connestabsucc', 'feildName': 'NG接口UE相关逻辑信令连接建立成功次数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'kpi_ngsig_succconnrate', 'feildName': 'NG接口UE相关逻辑信令连接建立成功率', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'kpi_wirelesssuccconnrate', 'feildName': '无线接通率', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'context_attrelgnb', 'feildName': 'gNB请求释放上下文数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'context_attrelgnb_normal', 'feildName': '正常的gNB请求释放上下文数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'context_succinitalsetup', 'feildName': '初始上下文建立成功次数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'context_nbrleft', 'feildName': '遗留上下文个数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'ho_succexecinc', 'feildName': '切换入成功次数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'rrc_succconnreestab_nonsrccell', 'feildName': 'RRC连接重建成功次数(非源侧小区)', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'kpi_wirelessdroprate_celllevel', 'feildName': '无线掉线率_小区级', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'flow_nbrreqrelgnb', 'feildName': 'gNB请求释放的Flow数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'flow_nbrreqrelgnb_normal', 'feildName': '正常的GNB请求释放的Flow数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'flow_hoadmitfail', 'feildName': '切出失败的Flow数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'flow_nbrleft', 'feildName': '遗留Flow个数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'flow_nbrhoinc', 'feildName': '切换入Flow数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'kpi_flowdroprate_celllevel', 'feildName': 'Flow掉线率（小区级）', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'rrc_attconnreestab', 'feildName': 'RRC连接重建请求次数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'kpi_rrcconnreestabrate', 'feildName': 'RRC连接重建比率', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'ho_attoutintercung', 'feildName': 'gNB间NG切换出准备请求次数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'ho_succoutintercung', 'feildName': 'gNB间NG切换出成功次数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'kpi_hosuccoutintergnbrate_ng', 'feildName': 'gNB间NG切换成功率', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'ho_attoutintercuxn', 'feildName': 'gNB间Xn切换出准备请求次数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'ho_succoutintercuxn', 'feildName': 'gNB间Xn切换出成功次数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'kpi_hosuccoutintergnbrate_xn', 'feildName': 'gNB间Xn切换成功率', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'kpi_hosuccoutintergnbrate', 'feildName': 'gNB间切换成功率', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'ho_attoutintracuinterdu', 'feildName': 'CU内DU间切换出执行请求次数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'ho_succoutintracuinterdu', 'feildName': 'CU内DU间切换出执行成功次数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'ho_attoutcuintradu', 'feildName': 'CU内DU内切换出执行请求次数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'ho_succoutintradu', 'feildName': 'CU内DU内切换出成功次数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'kpi_hosuccoutintragnbrate', 'feildName': 'gNB内切换成功率', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'kpi_hosuccoutrate', 'feildName': '切换成功率', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'ho_attoutexecintrafreq', 'feildName': '同频切换出执行请求次数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'ho_succoutintrafreq', 'feildName': '同频切换出成功次数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'kpi_hosuccoutrate_intrafreq', 'feildName': '同频切换执行成功率', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'ho_attoutexecinterfreq', 'feildName': '异频切换出执行请求次数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'ho_succoutinterfreq', 'feildName': '异频切换出成功次数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'kpi_hosuccoutrate_interfreq', 'feildName': '异频切换执行成功率', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'kpi_pdcpupoctul', 'feildName': 'PDCP上行业务字节数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'kpi_pdcpupoctdl', 'feildName': 'PDCP下行业务字节数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'ee_carriershutdowntime', 'feildName': '载波关断时长', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'flow_nbrattestab_5qi1', 'feildName': 'Flow建立请求数5QI1', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'flow_nbrsuccestab_5qi1', 'feildName': 'Flow建立成功数5QI1', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'kpi_wirelesssuccconnrate_5qi1', 'feildName': 'VoNR无线接通率(5QI1)', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'flow_nbrreqrelgnb_5qi1', 'feildName': 'gNB请求释放的Flow数5QI1', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'flow_nbrreqrelgnb_normal_5qi1', 'feildName': '正常的gNB请求释放的Flow数5QI1', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'flow_hoadmitfail_5qi1', 'feildName': '切出接纳失败的Flow数5QI1', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'flow_nbrleft_5qi1', 'feildName': '遗留Flow个数5QI1', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'flow_nbrhoinc_5qi1', 'feildName': '切换入Flow数5QI1', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'kpi_wirelessdroprate_celllevel_5qi1', 'feildName': '掉线率(5QI1)(小区级)', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'kpi_wirelessdroprate_netlevel_5qi1', 'feildName': '掉线率(5QI1)(网络级)', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'pdcp_upoctul_5qi1', 'feildName': '小区用户面上行PDCP PDU字节数5QI1', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'pdcp_upoctdl_5qi1', 'feildName': '小区用户面下行PDCP PDU字节数5QI1', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'pdcp_nbrpktlossul_5qi1', 'feildName': '上行PDCP丢包数5QI1', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'pdcp_nbrpktul_5qi1', 'feildName': '上行PDCP包数5QI1', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'kpi_pdcpnbrpktlossrateul_5qi1', 'feildName': '上行PDCP SDU平均丢包率(5QI1)', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'flow_nbrattestab_5qi2', 'feildName': 'Flow建立请求数5QI2', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'flow_nbrsuccestab_5qi2', 'feildName': 'Flow建立成功数5QI2', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'kpi_wirelesssuccconnrate_5qi2', 'feildName': 'VoNR无线接通率(5QI2)', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'flow_nbrreqrelgnb_5qi2', 'feildName': 'gNB请求释放的Flow数5QI2', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'flow_nbrreqrelgnb_normal_5qi2', 'feildName': '正常的gNB请求释放的Flow数5QI2', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'flow_hoadmitfail_5qi2', 'feildName': '切出接纳失败的Flow数5QI2', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'flow_nbrleft_5qi2', 'feildName': '遗留Flow个数5QI2', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'flow_nbrhoinc_5qi2', 'feildName': '切换入Flow数5QI2', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'kpi_wirelessdroprate_celllevel_5qi2', 'feildName': '掉线率(5QI2)(小区级)', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'kpi_wirelessdroprate_netlevel_5qi2', 'feildName': '掉线率(5QI2)(网络级)', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'pdcp_upoctul_5qi2', 'feildName': '小区用户面上行PDCP PDU字节数5QI2', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'pdcp_upoctdl_5qi2', 'feildName': '小区用户面下行PDCP PDU字节数5QI2', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'pdcp_nbrpktlossul_5qi2', 'feildName': '上行PDCP丢包数5QI2', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'pdcp_nbrpktul_5qi2', 'feildName': '上行PDCP包数5QI2', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'kpi_pdcpnbrpktlossrateul_5qi2', 'feildName': '上行PDCP SDU平均丢包率（5QI2）', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'vonr_voice_traffic', 'feildName': 'VoNR语音话务量', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'vinr_voice_traffic', 'feildName': 'ViNR语音话务量', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'iratho_succouteutran_epsfallback', 'feildName': 'EpsFallBack切换至LTE成功次数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'iratho_succprepouteutran_epsfallback', 'feildName': 'EpsFallBack切换至LTE准备成功次数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'flow_nbrattestab_epsfb', 'feildName': 'EPS fallback触发的Flow建立请求数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'iratho_attouteutran_epsfallback', 'feildName': 'EpsFallBack切换至LTE准备请求次数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'kpi_vonr_flowsuccconnrate', 'feildName': 'VoNR业务Flow建立成功率(5QI1)（小区级）', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'kpi_vinr_flowsuccconnrate', 'feildName': 'ViNR业务Flow建立成功率(5QI2)（小区级）', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'kpi_vonr_flowdroprate_celllevel', 'feildName': 'VoNR业务Flow掉线率（5QI1）（小区级）', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'kpi_vinr_flowdroprate_celllevel', 'feildName': 'ViNR业务Flow掉线率（5QI2）（小区级）', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'kpi_wirelessdroprate_netlevel', 'feildName': '无线掉线率（网络级）', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'kpi_flowdroprate_netlevel', 'feildName': 'Flow掉线率（网络级）', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'kpi_vonr_flowdroprate_netlevel', 'feildName': 'VoNR业务QoS Flow掉线率（5QI1）（网络级）', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'kpi_hosuccoutrate_intersystemnrtolte', 'feildName': 'NR到LTE的系统间切换出成功率', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'kpi_hosuccoutrate_intersystemltetonr', 'feildName': 'LTE到NR的系统间切换入成功率', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'kpi_esfbhosuccoutrate_intersystemnrtolte', 'feildName': 'NR到LTE的基于切换的EPSFB成功率', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'rrc_redirecttolte_epsfallback', 'feildName': 'EPS fallback RRC 重定向到LTE次数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'iratho_succprepouteutran', 'feildName': '切换至LTE准备成功次数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'iratho_attouteutran', 'feildName': '切换至LTE准备请求次数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'iratho_attprepinc', 'feildName': 'LTE切换入准备请求次数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'iratho_succprepinc', 'feildName': 'LTE切换入准备成功次数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'iratho_succouteutran', 'feildName': '切换至LTE成功次数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'kpi_vonr_succconnrate', 'feildName': 'VoNR业务接通成功率(5QI1)', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'kpi_vinr_succconnrate', 'feildName': 'ViNR业务接通成功率(5QI2)', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'kpi_slice_flowsuccconnrate', 'feildName': '每切片QOS FLOW建立成功率', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'kpi_inactive_succconnrate', 'feildName': 'RRC Resume成功率', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'kpi_hosuccoutrate_vonrtolte', 'feildName': 'VoNR到VoLTE的系统间切换出成功率', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'kpi_hosuccoutrate_vinrtolte', 'feildName': 'ViNR到ViLTE的系统间切换出成功率', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'kpi_hosuccoutrate_intersystemvoltetovonr', 'feildName': 'VoLTE到VoNR的系统间切换入成功率', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'kpi_hosuccoutrate_intersystemviltetovinr', 'feildName': 'ViLTE到ViNR的系统间切换入成功率', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'kpi_hosuccoutrate_vonr', 'feildName': 'VoNR系统内切换成功率', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'kpi_hosuccoutrate_vinr', 'feildName': 'ViNR系统内切换成功率', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'kpi_pdcpnbrpktlossrateul', 'feildName': 'PDCP层上行丢包率', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'kpi_vonrtraffic_5qi1', 'feildName': 'VoNR语音话务量（小时级）', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'kpi_vinrtraffic_5qi1', 'feildName': 'ViNR视频话务量（小时级）', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'flow_nbrsuccestab_vonr', 'feildName': 'VoNRFlow建立成功数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'flow_nbrattestab_vonr', 'feildName': 'VoNRFlow建立请求数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'flow_nbrsuccestab_vinr', 'feildName': 'ViNRFlow建立成功数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'flow_nbrattestab_vinr', 'feildName': 'ViNRFlow建立请求数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'flow_nbrsuccestabslice', 'feildName': '每切片FLOW建立成功数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'flow_nbrattestabslice', 'feildName': '每切片FLOW建立请求数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'rrc_succconnresume', 'feildName': 'Resume成功次数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'rrc_attconnresume', 'feildName': 'resume请求次数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'iratho_succouteutran_vonr', 'feildName': 'VoNR切换至LTE成功Flow数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'iratho_attouteutran_vonr', 'feildName': 'VoNR切换至LTE准备请求Flow数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'iratho_succouteutran_vinr', 'feildName': 'ViNR切换至LTE成功Flow数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'iratho_attouteutran_vinr', 'feildName': 'ViNR切换至LTE准备请求Flow数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'iratho_succexecinc_voltetovonr', 'feildName': 'LTEVoLTEtoVoNR切换入成功flow数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'iratho_attprepinc_voltetovonr', 'feildName': 'LTEVoLTEtoVoNR切换入准备请求flow数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'iratho_succexecinc_viltetovinr', 'feildName': 'LTEViLTEtoVoiNR切换入成功flow数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'iratho_attprepinc_viltetovinr', 'feildName': 'LTEViLTEtoViNR切换入准备请求flow数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'ho_succoutintercung_vonr', 'feildName': 'VoNRgNB间NG切换出成功flow数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'ho_succoutintercuxn_vonr', 'feildName': 'VoNRgNB间Xn切换出成功flow数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'ho_succoutintradu_vonr', 'feildName': 'VoNRCU内DU内切换出成功flow数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'ho_attoutintercung_vonr', 'feildName': 'VoNRgNB间NG切换出准备请求flow数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'ho_attoutintercuxn_vonr', 'feildName': 'VoNRgNB间Xn切换出准备请求flow数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'ho_attprepoutcuintradu_vonr', 'feildName': 'VoNRCU内DU内切换出准备请求flow数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'ho_succoutintercung_vinr', 'feildName': 'ViNRgNB间NG切换出成功flow数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'ho_succoutintercuxn_vinr', 'feildName': 'ViNRgNB间Xn切换出成功flow数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'ho_succoutintradu_vinr', 'feildName': 'ViNRCU内DU内切换出成功flow数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'ho_attoutintercung_vinr', 'feildName': 'ViNRgNB间NG切换出准备请求flow数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'ho_attoutintercuxn_vinr', 'feildName': 'ViNRgNB间Xn切换出准备请求flow数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'ho_attprepoutcuintradu_vinr', 'feildName': 'ViNRCU内DU内切换出准备请求flow数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'pdcp_nbrpktlossul', 'feildName': '上行PDCP丢包数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'pdcp_nbrpktul', 'feildName': '上行PDCP包数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'flow_nbrmeanestab_5qi1', 'feildName': '5QI1的平均Flow数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'flow_nbrmeanestab_5qi2', 'feildName': '5QI2的平均Flow数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'kpi_wirelesssuccconnrate_v1_8', 'feildName': '无线接通率（1.8算法）', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'kpi_vonr_flowdroprate_netlevel_v1_8', 'feildName': 'VONR业务QOS FLOW掉线率（5QI1）（网络级）_（1.8算法）', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'kpi_hosuccoutrate_intersystemltetonr_v1_8', 'feildName': 'LTE到NR的系统间切换入成功率_（1.8算法）', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'flow_succestab_resume_vonr', 'feildName': 'VoNR Resume建立成功的Flow数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'iratho_succexecinc', 'feildName': 'LTE切换入成功次数', 'datatype': 'character varying', 'columntype': 1},
    ]
    
    # 构建result配置
    result_list = []
    for f in result_fields:
        result_list.append({
            'feildtype': 'SA_CU性能',
            'table': 'appdbv3.a_common_pm_sacu',
            'tableName': '5GSA_CU性能报表',
            'datatype': f['datatype'],
            'columntype': f['columntype'],
            'feildName': f['feildName'],
            'feild': f['feild'],
            'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'
        })
    
    payload = {
        'draw': 1, 'start': 0, 'length': 200, 'total': 0,
        'geographicdimension': '小区，网格，地市，分公司', 'timedimension': '小时,天,周,月',
        'enodebField': 'gnodeb_id', 'cgiField': 'ncgi',
        'timeField': 'starttime', 'cellField': 'nrcell', 'cityField': 'city',
        'columns': _build_columns_param(kpi_fields),
        'result': {'result': result_list, 'tableParams': {'supporteddimension': '0', 'supportedtimedimension': '1'}, 'columnname': ''},
        'order': [{'column': 0, 'dir': 'desc'}],
        'search': {'value': '', 'regex': False},
        'where': [
            {'datatype': 'timestamp', 'feild': 'starttime', 'feildName': '', 'symbol': '>=', 'val': '2026-04-19 00:00:00', 'whereCon': 'and', 'query': True},
            {'datatype': 'timestamp', 'feild': 'starttime', 'feildName': '', 'symbol': '<', 'val': '2026-04-19 23:59:59', 'whereCon': 'and', 'query': True},
            {'datatype': 'character', 'feild': 'city', 'feildName': '', 'symbol': 'in', 'val': '阳江', 'whereCon': 'and', 'query': True}
        ],
        'indexcount': 0
    }
    return payload


def get_4g_kpi_payload():
    """获取4G小区性能KPI报表payload（硬编码，字段与浏览器一致）"""
    print("[DEBUG-PAYLOAD] 生成 4G小区性能KPI报表 payload")
    
    # 4G小区KPI字段列表（与浏览器请求一致，共98个字段）
    kpi_fields = [
        'starttime', 'endtime', 'cgi', 'cell_name', 'city', 'area', 'branch', 'grid',
        'marketduty', 'marketgrid', 'network_type', 'state', 'cover_type',
        'cover_scene1', 'cover_scene2', 'cover_scene3', 'cover_scene4',
        'freq', 'vendor', 'attconnestab', 'uplastttioctdl', 'uplastttioctul',
        'rrc_succ_rate', 'rrc_restab_rate', 'nbrattestab', 'nbrsuccestab',
        'e_rab_succ_rate', 'radio_succ_rate', 'nbrattestab_1', 'e_rab_succ_rate_1',
        'nbrfailestab_rsnotavailable', 'e_rab_block_rate',
        'nbrreqrelenb', 'nbrreqrelenb_normal', 'hofail', 'nbrleft', 'nbrhoinc',
        'erab_drop_rate', 'attrelenb', 'attrelenbnormal', 'succinitalsetup',
        'nbrleft_context', 'radio_drop_rate', 'radio_drop_rate_cell',
        'attrelenb_userinactivity', 'radio_drop_rate_noui', 'radio_drop_rate_cell_noui',
        'nbrreqrelenb_userinactivity', 'erab_drop_rate_noui',
        'enbout_succ_rate_s1', 'enbinter_succ_rate_x2', 'enbinter_succ_rate',
        'enbintra_succ_rate', 'ho_succ_out', 'ho_att__out', 'enbout_succ_rate',
        'intrafreq_succ_rate', 'interfreq_succ_rate',
        'lte_gsm_succ_rate', 'gsm_lte_succ_rate',
        'lte_utran_succ_rate', 'utran_lte_succ_rate',
        'nbrpktlossul', 'nbrpktul', 'rktul_loss_rate',
        'nbrpktlossdl', 'nbrpktdl', 'rktdl_loss_rate',
        'mac_ul_reser_rate', 'mac_dl_reser_rate',
        'harq_ul_rate', 'harq_dl_rate',
        'rank2_rate', 'ul_qpsk_rate', 'dl_qpsk_rate',
        'pkt_loss_ul_1', 'pkt_loss_dl_1',
        'ul_rtb_rate_1', 'dl_rtb_rate_1',
        'upoctul', 'upoctdl', 'ul_thrp', 'dl_thrp',
        'ul_dtchprb_rate', 'ul_ctrlprb_rate',
        'dl_dtchprb_rate', 'dl_ctrlprb_rate',
        'puschprbassn', 'ul_prbuse_rate', 'puschprbtot',
        'pdschprbtot', 'pdschprbassn',
        'dl_prbuse_rate', 'prbuse_rate',
        'pagreceived', 'pagdiscarded', 'page_disc_rate', 'volte_voice_traffic'
    ]
    
    # 构建result字段配置（与浏览器请求一致）
    result_fields = [
        {'feild': 'starttime', 'feildName': '开始时间', 'datatype': 'character varying', 'columntype': 2},
        {'feild': 'endtime', 'feildName': '结束时间', 'datatype': 'character varying', 'columntype': 2},
        {'feild': 'cgi', 'feildName': 'CGI', 'datatype': 'character varying', 'columntype': 2},
        {'feild': 'cell_name', 'feildName': '小区名称', 'datatype': 'character varying', 'columntype': 2},
        {'feild': 'city', 'feildName': '所属地市', 'datatype': 'character varying', 'columntype': 2},
        {'feild': 'area', 'feildName': '所属区县', 'datatype': 'character varying', 'columntype': 2},
        {'feild': 'branch', 'feildName': '人力区县分公司', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'grid', 'feildName': '网格', 'datatype': 'character varying', 'columntype': 2},
        {'feild': 'marketduty', 'feildName': '责任田', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'marketgrid', 'feildName': '市场网格', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'network_type', 'feildName': '网络制式', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'state', 'feildName': '网元状态', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'cover_type', 'feildName': '覆盖类型', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'cover_scene1', 'feildName': '一级场景', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'cover_scene2', 'feildName': '二级场景', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'cover_scene3', 'feildName': '三级场景', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'cover_scene4', 'feildName': '四级场景', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'freq', 'feildName': '频段', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'vendor', 'feildName': '厂家', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'attconnestab', 'feildName': 'RRC连接建立请求次数', 'datatype': 'bigint', 'columntype': 1},
        {'feild': 'uplastttioctdl', 'feildName': '用户面下行尾包字节数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'uplastttioctul', 'feildName': '用户面上行尾包字节数', 'datatype': 'character varying', 'columntype': 1},
        {'feild': 'rrc_succ_rate', 'feildName': 'RRC连接建立成功率(%)', 'datatype': 'decimal', 'columntype': 1},
        {'feild': 'rrc_restab_rate', 'feildName': 'RRC连接重建比率(%)', 'datatype': 'decimal', 'columntype': 1},
        {'feild': 'nbrattestab', 'feildName': 'E-RAB建立请求数', 'datatype': 'bigint', 'columntype': 1},
        {'feild': 'nbrsuccestab', 'feildName': 'E-RAB建立成功数', 'datatype': 'bigint', 'columntype': 1},
        {'feild': 'e_rab_succ_rate', 'feildName': 'E-RAB建立成功率(%)', 'datatype': 'decimal', 'columntype': 1},
        {'feild': 'radio_succ_rate', 'feildName': '无线接通率(%)', 'datatype': 'decimal', 'columntype': 1},
        {'feild': 'nbrattestab_1', 'feildName': 'E-RAB建立请求数(QCI=1)', 'datatype': 'bigint', 'columntype': 1},
        {'feild': 'e_rab_succ_rate_1', 'feildName': 'E-RAB建立成功率(QCI=1)(%)', 'datatype': 'decimal', 'columntype': 1},
        {'feild': 'nbrfailestab_rsnotavailable', 'feildName': '无线资源不足原因导致的E-RAB建立失败数', 'datatype': 'bigint', 'columntype': 1},
        {'feild': 'e_rab_block_rate', 'feildName': 'E-RAB拥塞率(无线资源不足)(%)', 'datatype': 'decimal', 'columntype': 1},
        {'feild': 'nbrreqrelenb', 'feildName': 'eNB请求释放的E-RAB数', 'datatype': 'bigint', 'columntype': 1},
        {'feild': 'nbrreqrelenb_normal', 'feildName': '正常的eNB请求释放的E-RAB数', 'datatype': 'bigint', 'columntype': 1},
        {'feild': 'hofail', 'feildName': '切出失败的E-RAB数', 'datatype': 'bigint', 'columntype': 1},
        {'feild': 'nbrleft', 'feildName': '遗留上下文个数', 'datatype': 'bigint', 'columntype': 1},
        {'feild': 'nbrhoinc', 'feildName': '切换入E-RAB数', 'datatype': 'bigint', 'columntype': 1},
        {'feild': 'erab_drop_rate', 'feildName': 'E-RAB掉线率(%)', 'datatype': 'decimal', 'columntype': 1},
        {'feild': 'attrelenb', 'feildName': 'eNB请求释放上下文数', 'datatype': 'bigint', 'columntype': 1},
        {'feild': 'attrelenbnormal', 'feildName': '正常的eNB请求释放上下文数', 'datatype': 'bigint', 'columntype': 1},
        {'feild': 'succinitalsetup', 'feildName': '初始上下文建立成功次数', 'datatype': 'bigint', 'columntype': 1},
        {'feild': 'nbrleft_context', 'feildName': '遗留上下文个数CONTEXT', 'datatype': 'bigint', 'columntype': 1},
        {'feild': 'radio_drop_rate', 'feildName': '无线掉线率(%)', 'datatype': 'decimal', 'columntype': 1},
        {'feild': 'radio_drop_rate_cell', 'feildName': '无线掉线率(小区级)(%)', 'datatype': 'decimal', 'columntype': 1},
        {'feild': 'attrelenb_userinactivity', 'feildName': '用户不活动原因eNB请求释放上下文数', 'datatype': 'bigint', 'columntype': 1},
        {'feild': 'radio_drop_rate_noui', 'feildName': '无线掉线率(剔除UI原因)(%)', 'datatype': 'decimal', 'columntype': 1},
        {'feild': 'radio_drop_rate_cell_noui', 'feildName': '无线掉线率(剔除UI原因)(小区级)(%)', 'datatype': 'decimal', 'columntype': 1},
        {'feild': 'nbrreqrelenb_userinactivity', 'feildName': '用户不活动原因eNB请求释放的E-RAB数', 'datatype': 'bigint', 'columntype': 1},
        {'feild': 'erab_drop_rate_noui', 'feildName': 'E-RAB掉线率(剔除UI原因)(小区级)(%)', 'datatype': 'decimal', 'columntype': 1},
        {'feild': 'enbout_succ_rate_s1', 'feildName': 'eNB间S1切换成功率(%)', 'datatype': 'decimal', 'columntype': 1},
        {'feild': 'enbinter_succ_rate_x2', 'feildName': 'eNB间X2切换成功率(%)', 'datatype': 'decimal', 'columntype': 1},
        {'feild': 'enbinter_succ_rate', 'feildName': 'eNB间切换成功率(%)', 'datatype': 'decimal', 'columntype': 1},
        {'feild': 'enbintra_succ_rate', 'feildName': 'eNB内切换成功率(%)', 'datatype': 'decimal', 'columntype': 1},
        {'feild': 'ho_succ_out', 'feildName': '切换成功次数', 'datatype': 'bigint', 'columntype': 1},
        {'feild': 'ho_att__out', 'feildName': '切换请求次数', 'datatype': 'bigint', 'columntype': 1},
        {'feild': 'enbout_succ_rate', 'feildName': '切换成功率(%)', 'datatype': 'decimal', 'columntype': 1},
        {'feild': 'intrafreq_succ_rate', 'feildName': '同频切换执行成功率(%)', 'datatype': 'decimal', 'columntype': 1},
        {'feild': 'interfreq_succ_rate', 'feildName': '异频切换执行成功率(%)', 'datatype': 'decimal', 'columntype': 1},
        {'feild': 'lte_gsm_succ_rate', 'feildName': 'LTE到2G切换成功率(%)', 'datatype': 'decimal', 'columntype': 1},
        {'feild': 'gsm_lte_succ_rate', 'feildName': '2G到LTE切换成功率(%)', 'datatype': 'decimal', 'columntype': 1},
        {'feild': 'lte_utran_succ_rate', 'feildName': 'LTE到3G切换成功率(%)', 'datatype': 'decimal', 'columntype': 1},
        {'feild': 'utran_lte_succ_rate', 'feildName': '3G到LTE切换成功率(%)', 'datatype': 'decimal', 'columntype': 1},
        {'feild': 'nbrpktlossul', 'feildName': '小区上行丢包数', 'datatype': 'bigint', 'columntype': 1},
        {'feild': 'nbrpktul', 'feildName': '小区上行包数', 'datatype': 'bigint', 'columntype': 1},
        {'feild': 'rktul_loss_rate', 'feildName': '小区用户面上行丢包率(ppm)', 'datatype': 'decimal', 'columntype': 1},
        {'feild': 'nbrpktlossdl', 'feildName': '小区下行丢包数', 'datatype': 'bigint', 'columntype': 1},
        {'feild': 'nbrpktdl', 'feildName': '小区下行包数', 'datatype': 'bigint', 'columntype': 1},
        {'feild': 'rktdl_loss_rate', 'feildName': '小区用户面下行丢包率(ppm)', 'datatype': 'decimal', 'columntype': 1},
        {'feild': 'mac_ul_reser_rate', 'feildName': 'MAC层上行误块率(%)', 'datatype': 'decimal', 'columntype': 1},
        {'feild': 'mac_dl_reser_rate', 'feildName': 'MAC层下行误块率(%)', 'datatype': 'decimal', 'columntype': 1},
        {'feild': 'harq_ul_rate', 'feildName': '上行初始HARQ重传比率(%)', 'datatype': 'decimal', 'columntype': 1},
        {'feild': 'harq_dl_rate', 'feildName': '下行初始HARQ重传比率(%)', 'datatype': 'decimal', 'columntype': 1},
        {'feild': 'rank2_rate', 'feildName': '下行双流占比(%)', 'datatype': 'decimal', 'columntype': 1},
        {'feild': 'ul_qpsk_rate', 'feildName': '上行QPSK编码比例(%)', 'datatype': 'decimal', 'columntype': 1},
        {'feild': 'dl_qpsk_rate', 'feildName': '下行QPSK编码比例(%)', 'datatype': 'decimal', 'columntype': 1},
        {'feild': 'pkt_loss_ul_1', 'feildName': 'VoLTE上行丢包率(ppm)', 'datatype': 'decimal', 'columntype': 1},
        {'feild': 'pkt_loss_dl_1', 'feildName': 'VoLTE下行丢包率(ppm)', 'datatype': 'decimal', 'columntype': 1},
        {'feild': 'ul_rtb_rate_1', 'feildName': '上行半持续调度次数占比(%)', 'datatype': 'decimal', 'columntype': 1},
        {'feild': 'dl_rtb_rate_1', 'feildName': '下行半持续调度次数占比(%)', 'datatype': 'decimal', 'columntype': 1},
        {'feild': 'upoctul', 'feildName': '上行流量(KByte)', 'datatype': 'decimal', 'columntype': 1},
        {'feild': 'upoctdl', 'feildName': '下行流量(KByte)', 'datatype': 'decimal', 'columntype': 1},
        {'feild': 'ul_thrp', 'feildName': '上行用户平均速率(Mbps)', 'datatype': 'decimal', 'columntype': 1},
        {'feild': 'dl_thrp', 'feildName': '下行用户平均速率(Mbps)', 'datatype': 'decimal', 'columntype': 1},
        {'feild': 'ul_dtchprb_rate', 'feildName': '上行业务信息PRB占用率(%)', 'datatype': 'decimal', 'columntype': 1},
        {'feild': 'ul_ctrlprb_rate', 'feildName': '上行控制信息PRB占用率(%)', 'datatype': 'decimal', 'columntype': 1},
        {'feild': 'dl_dtchprb_rate', 'feildName': '下行业务信息PRB占用率(%)', 'datatype': 'decimal', 'columntype': 1},
        {'feild': 'dl_ctrlprb_rate', 'feildName': '下行控制信息PRB占用率(%)', 'datatype': 'decimal', 'columntype': 1},
        {'feild': 'puschprbassn', 'feildName': '上行PUSCHPRB占用数', 'datatype': 'decimal', 'columntype': 1},
        {'feild': 'ul_prbuse_rate', 'feildName': '上行PRB平均利用率(%)', 'datatype': 'decimal', 'columntype': 1},
        {'feild': 'puschprbtot', 'feildName': '上行PUSCHPRB可用数', 'datatype': 'decimal', 'columntype': 1},
        {'feild': 'pdschprbtot', 'feildName': '下行PDSCHPRB可用数', 'datatype': 'decimal', 'columntype': 1},
        {'feild': 'pdschprbassn', 'feildName': '下行PDSCHPRB占用数', 'datatype': 'decimal', 'columntype': 1},
        {'feild': 'dl_prbuse_rate', 'feildName': '下行PRB平均利用率(%)', 'datatype': 'decimal', 'columntype': 1},
        {'feild': 'prbuse_rate', 'feildName': '无线利用率(%)', 'datatype': 'decimal', 'columntype': 1},
        {'feild': 'pagreceived', 'feildName': '寻呼记录接收个数', 'datatype': 'bigint', 'columntype': 1},
        {'feild': 'pagdiscarded', 'feildName': '寻呼记录丢弃个数', 'datatype': 'bigint', 'columntype': 1},
        {'feild': 'page_disc_rate', 'feildName': 'eNodeB寻呼拥塞率(%)', 'datatype': 'decimal', 'columntype': 1},
        {'feild': 'volte_voice_traffic', 'feildName': 'VOLTE语音话务量', 'datatype': 'decimal', 'columntype': 1},
    ]
    
    # 构建result配置
    result_list = []
    for f in result_fields:
        result_list.append({
            'feildtype': '公共信息（小区级粒度）',
            'table': 'appdbv3.a_common_pm_lte',
            'tableName': '4G小区性能KPI报表',
            'datatype': f['datatype'],
            'columntype': f['columntype'],
            'feildName': f['feildName'],
            'feild': f['feild'],
            'poly': '无', 'anyWay': '无', 'chart': '无', 'chartpoly': '无'
        })
    
    payload = {
        'draw': 1, 'start': 0, 'length': 200, 'total': 0,
        'geographicdimension': '小区，网格，地市，分公司', 'timedimension': '小时,天,周.月,忙时,15分钟',
        'enodebField': 'enodeb_id', 'cgiField': 'cgi',
        'timeField': 'starttime', 'cellField': 'cell', 'cityField': 'city',
        'columns': _build_columns_param(kpi_fields),
        'result': {'result': result_list, 'tableParams': {'supporteddimension': '0', 'supportedtimedimension': '1'}, 'columnname': ''},
        'order': [{'column': 0, 'dir': 'desc'}],
        'search': {'value': '', 'regex': False},
        'where': [
            {'datatype': 'timestamp', 'feild': 'starttime', 'feildName': '', 'symbol': '>=', 'val': '2026-04-19 00:00:00', 'whereCon': 'and', 'query': True},
            {'datatype': 'timestamp', 'feild': 'starttime', 'feildName': '', 'symbol': '<', 'val': '2026-04-19 23:59:59', 'whereCon': 'and', 'query': True},
            {'datatype': 'character', 'feild': 'city', 'feildName': '', 'symbol': 'in', 'val': '阳江', 'whereCon': 'and', 'query': True}
        ],
        'indexcount': 0
    }
    return payload


def set_payload_time(payload, start_time, end_time):
    """设置payload的查询时间"""
    print(f"[DEBUG-PAYLOAD] 设置时间范围: {start_time} 至 {end_time}")

    # 获取 timeField 判断时间字段类型
    time_field = payload.get('timeField', 'day_id')
    print(f"[DEBUG-PAYLOAD] timeField: {time_field}")

    # 如果是 day_id（日期类型），只取日期部分
    # 如果是 starttime（timestamp类型），需要完整的时间格式
    if time_field == 'day_id':
        start_date = start_time.split(' ')[0]  # 只取日期部分
        end_date = end_time.split(' ')[0]      # 只取日期部分
        print(f"[DEBUG-PAYLOAD] day_id 类型，只取日期: {start_date} 至 {end_date}")
    elif time_field == 'starttime':
        # starttime 是 timestamp 类型，格式是 YYYY-MM-DD HH:MM:SS
        # 起始时间用 00:00:00，结束时间用 23:59:59
        start_date = start_time  # 已经是完整格式
        end_date = end_time      # 已经是完整格式
        print(f"[DEBUG-PAYLOAD] starttime 类型，使用完整时间: {start_date} 至 {end_date}")
    else:
        start_date = start_time
        end_date = end_time

    for condition in payload['where']:
        if condition['feild'] in ['day_id', 'starttime']:
            if condition['symbol'] in ['>=', '>']:
                condition['val'] = start_date
                print(f"[DEBUG-PAYLOAD] 设置 {condition['feild']} >= {condition['val']}")
            elif condition['symbol'] in ['<', '<=']:
                condition['val'] = end_date
                print(f"[DEBUG-PAYLOAD] 设置 {condition['feild']} < {condition['val']}")
    return payload


def set_payload_city(payload, city):
    """设置payload的查询城市，支持多地市（逗号分隔）"""
    print(f"[DEBUG-PAYLOAD] 设置查询城市: {city}")
    # 处理多地市情况，如 "广州,深圳,东莞" -> "广州,深圳,东莞"
    # 如果只有一个地市，也转换为列表处理
    city_list = [c.strip() for c in city.split(',') if c.strip()]
    city_value = ','.join(city_list)
    
    for condition in payload['where']:
        if condition['feild'] == 'city':
            condition['val'] = city_value
            print(f"[DEBUG-PAYLOAD] 设置 city = {city_value}")
    return payload


# ==================== 数据提取器 ====================
class CustomDataExtractor:
    """自定义数据提取器"""

    def __init__(self, username=None, password=None, parent=None):
        self.username = username or DEFAULT_USERNAME
        self.password = password or DEFAULT_PASSWORD
        self.parent = parent  # GUI 父窗口
        self.login_mgr = None
        self.jxcx = None

    def login(self):
        """执行登录"""
        self.login_mgr = LoginManager(self.username, self.password, self.parent)
        return self.login_mgr.login()
    
    def init_jxcx(self):
        """初始化即席查询"""
        if not self.login_mgr:
            print("✗ 请先执行登录")
            return False
        
        self.jxcx = JXCXQuery(self.login_mgr.sess)
        return self.jxcx.enter_jxcx(retry_times=3, timeout=60)
    
    def check_and_relogin(self):
        """检查session是否有效，如果无效则重新登录"""
        if not self.login_mgr:
            return False
        
        # 检查CASTGC cookie是否存在
        castgc = self.login_mgr.sess.cookies.get('CASTGC', domain='nqi.gmcc.net')
        if not castgc:
            castgc = self.login_mgr.sess.cookies.get('CASTGC')
        
        if not castgc:
            print("[WARNING] CASTGC cookie不存在或已过期，尝试重新登录...")
            # 尝试使用保存的cookie登录
            saved_cookie = load_cookie(self.username)
            if saved_cookie:
                self.login_mgr.sess.cookies = saved_cookie
                if self.login_mgr._check_session():
                    print("[SUCCESS] 使用保存的Cookie重新登录成功！")
                    # 重新初始化jxcx
                    self.jxcx = JXCXQuery(self.login_mgr.sess)
                    return self.jxcx.enter_jxcx(retry_times=2, timeout=60)
            
            # 保存的cookie也无效，需要重新登录
            print("[WARNING] 保存的Cookie也失效，需要重新登录")
            return False
        
        return True
    
    def extract_data(self, payload, start_date, end_date, city):
        """提取数据"""
        if not self.jxcx:
            return pd.DataFrame()
        
        payload = set_payload_time(payload, start_date + ' 00:00:00', end_date + ' 23:59:59')
        payload = set_payload_city(payload, city)
        
        return self.jxcx.get_table(payload, to_df=True)
    
    def get_payload_with_dynamic_fields(self, table_config, start_date, end_date, city):
        """获取payload，自动处理动态字段配置"""
        table_name = table_config.get('description', '')
        
        # 需要动态获取字段的报表类型（不包括4G/5G小区性能KPI报表，因为已经硬编码）
        dynamic_tables = [
            'VoLTE小区监控预警', 'EPSFB小区监控预警', 'VONR小区监控预警'
        ]
        
        if table_name in dynamic_tables:
            # 映射表名到API key、fieldtype和API类型
            # api_type: 'search' 使用adhocquery/search接口，'table' 使用adhocquery/getSelectTable接口
            # 注意：API key与HAR文件中的实际请求参数一致
            # 注意：对于'table'类型，table_key需要使用数据库表名（如appdbv3.a_common_pm_lte）
            # 而fieldtype用于过滤字段配置
            table_key_map = {
                'VoLTE小区监控预警': ('volte小区监控预警', 'VoLTE小区监控预警数据表-天', 'search'),
                'EPSFB小区监控预警': ('EPSFB', 'EPSFB小区监控预警数据表-天', 'search'),
                'VONR小区监控预警': ('vonr', 'VONR小区监控预警数据表-天', 'search'),
                '5G小区性能KPI报表': ('appdbv3.a_common_pm_sacu', 'SA_CU性能', 'table'),
                '4G小区性能KPI报表': ('appdbv3.a_common_pm_lte', '公共信息（小区级粒度）', 'table'),
            }
            
            table_key, fieldtype, api_type = table_key_map.get(table_name, (None, None, None))
            
            if table_key and fieldtype:
                # 构建where条件
                start_time = start_date + ' 00:00:00'
                end_time = end_date + ' 23:59:59'
                where_conditions = [
                    {'datatype': 'timestamp', 'feild': 'starttime', 'feildName': '', 'symbol': '>=', 'val': start_time, 'whereCon': 'and', 'query': True},
                    {'datatype': 'timestamp', 'feild': 'starttime', 'feildName': '', 'symbol': '<', 'val': end_time, 'whereCon': 'and', 'query': True},
                    {'datatype': 'character', 'feild': 'city', 'feildName': '', 'symbol': 'in', 'val': city, 'whereCon': 'and', 'query': True}
                ]
                
                # 动态构建payload
                payload = self.jxcx.build_payload_from_config(table_key, fieldtype, where_conditions, api_type)
                if payload:
                    return payload
        
        # 对于其他表，使用静态payload函数
        payload_func = table_config.get('payload_func')
        if payload_func:
            payload = payload_func()
            # 如果返回None（4G语音小区等特殊报表），返回特殊标记让调用方处理
            if payload is None:
                return {'__special__': '4G_VOICE_MERGE', 'start_date': start_date, 'end_date': end_date, 'city': city}
            payload = set_payload_time(payload, start_date + ' 00:00:00', end_date + ' 23:59:59')
            payload = set_payload_city(payload, city)
            return payload

        return None
    
    def save_to_excel(self, df, filename):
        """保存到Excel，自动分Sheet（每50万行一个Sheet）"""
        if df.empty:
            return None
        
        ensure_dirs()
        filepath = os.path.join(OUTPUT_DIR, filename)
        
        MAX_ROWS_PER_SHEET = 500000  # 每个Sheet最大行数
        
        try:
            total_rows = len(df)
            
            if total_rows <= MAX_ROWS_PER_SHEET:
                # 数据量小，直接保存
                df.to_excel(filepath, index=False, engine='openpyxl')
            else:
                # 数据量大，分多个Sheet保存
                sheet_num = (total_rows + MAX_ROWS_PER_SHEET - 1) // MAX_ROWS_PER_SHEET
                base_name = os.path.splitext(filename)[0]
                filepath = os.path.join(OUTPUT_DIR, f'{base_name}.xlsx')
                
                with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                    for i in range(sheet_num):
                        start_idx = i * MAX_ROWS_PER_SHEET
                        end_idx = min((i + 1) * MAX_ROWS_PER_SHEET, total_rows)
                        sheet_name = f'数据{i + 1}'
                        df.iloc[start_idx:end_idx].to_excel(writer, sheet_name=sheet_name, index=False)
                
                print(f"数据共 {total_rows} 行，已拆分为 {sheet_num} 个Sheet保存")
            
            return filepath
        except Exception as e:
            print(f"保存文件失败: {e}")
            return None

    def _append_df_to_excel(self, filepath, df, sheet_name):
        """追加 DataFrame 到已存在的 Excel 文件（追加模式）"""
        try:
            # 使用 append 模式打开已存在的文件
            with pd.ExcelWriter(filepath, engine='openpyxl', mode='a') as writer:
                # 删除已存在的同名 Sheet（如果存在）
                if sheet_name in writer.sheets:
                    del writer.sheets[sheet_name]
                # 写入数据
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        except Exception as e:
            print(f"追加数据到Excel失败: {e}")
            raise

    def save_4g_voice_to_excel(self, volte_df, epsfb_df, merged_df, filename):
        """保存4G语音小区数据到Excel（三个Sheet）

        Args:
            volte_df: VoLTE原始数据
            epsfb_df: EPSFB原始数据
            merged_df: 合并后的计算结果
            filename: 文件名
        """
        if volte_df.empty and epsfb_df.empty and merged_df.empty:
            return None

        ensure_dirs()
        filepath = os.path.join(OUTPUT_DIR, filename)

        try:
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Sheet1: VoLTE数据
                if not volte_df.empty:
                    volte_df.to_excel(writer, sheet_name='VoLTE数据', index=False)
                    print(f"VoLTE数据: {len(volte_df)} 行")

                # Sheet2: EPSFB数据
                if not epsfb_df.empty:
                    epsfb_df.to_excel(writer, sheet_name='EPSFB数据', index=False)
                    print(f"EPSFB数据: {len(epsfb_df)} 行")

                # Sheet3: 聚合计算结果
                if not merged_df.empty:
                    merged_df.to_excel(writer, sheet_name='聚合计算', index=False)
                    print(f"聚合计算: {len(merged_df)} 行")

            print(f"4G语音小区数据已保存到: {filepath}")
            return filepath
        except Exception as e:
            print(f"保存4G语音小区Excel失败: {e}")
            return None

    def _finalize_per_day_excel(self, filepath, expected_sheets):
        """清理按日分Sheet模式创建的临时文件"""
        try:
            if os.path.exists(filepath):
                book = load_workbook(filepath)
                
                # 删除初始化 Sheet（如果存在）
                if '_init_' in book.sheetnames:
                    del book['_init_']
                
                # 如果没有任何有效数据 Sheet，删除整个文件
                if len(book.sheetnames) == 0:
                    book.close()
                    os.remove(filepath)
                    return
                
                book.save(filepath)
                book.close()
        except Exception as e:
            print(f"清理Excel文件失败: {e}")


# ==================== 数据表配置 ====================
class TableConfig:
    """数据表配置类"""

    TABLES = {
        '5G干扰小区': {
            'payload_func': get_5g_interference_payload,
            'use_interference_filter': True,
            'description': '5G干扰小区数据提取'
        },
        '4G干扰小区': {
            'payload_func': get_4g_interference_payload,
            'use_interference_filter': True,
            'description': '4G干扰小区数据提取'
        },
        '5G小区容量报表': {
            'payload_func': get_5g_capacity_payload,
            'use_interference_filter': False,
            'description': '5G小区容量报表 - 天粒度'
        },
        '重要场景-天': {
            'payload_func': get_important_scene_payload,
            'use_interference_filter': False,
            'description': '重要场景-小区天粒度'
        },
        'VoLTE小区监控预警': {
            'payload_func': get_volte_warning_payload,
            'use_interference_filter': False,
            'description': 'VoLTE小区监控预警数据'
        },
        'EPSFB小区监控预警': {
            'payload_func': get_epsfb_warning_payload,
            'use_interference_filter': False,
            'description': 'EPSFB小区监控预警数据'
        },
        'VONR小区监控预警': {
            'payload_func': get_vonr_warning_payload,
            'use_interference_filter': False,
            'description': 'VONR小区监控预警数据'
        },
        '5G小区性能KPI报表': {
            'payload_func': get_5g_kpi_payload,
            'use_interference_filter': False,
            'description': '5G小区性能KPI报表'
        },
        '4G小区性能KPI报表': {
            'payload_func': get_4g_kpi_payload,
            'use_interference_filter': False,
            'description': '4G小区性能KPI报表'
        },
        '4G全程完好率报表': {
            'payload_func': get_4g_wanchenglv_payload,
            'use_interference_filter': False,
            'calc_columns': ['4G全程完好率'],
            'description': '4G全程完好率报表'
        },
        '5G全程完好率报表': {
            'payload_func': get_5g_wanchenglv_payload,
            'use_interference_filter': False,
            'calc_columns': ['5G全程完好率'],
            'description': '5G全程完好率报表'
        },
        '4G语音小区': {
            'payload_func': get_4g_voice_payload,
            'use_interference_filter': False,
            'calc_columns': ['4G语音小区'],
            'description': '4G语音小区（VoLTE+EPSFB联合）'
        },
        '5G语音小区': {
            'payload_func': get_5g_voice_payload,
            'use_interference_filter': False,
            'calc_columns': ['5G语音小区'],
            'description': '5G语音小区（VONR）'
        },
    }

    @classmethod
    def get_table_names(cls):
        return list(cls.TABLES.keys())

    @classmethod
    def get_table_config(cls, table_name):
        return cls.TABLES.get(table_name)


# ==================== 多选下拉组件 ====================
class MultiSelectDropdown(ttk.Frame):
    """带复选框的下拉选择组件"""
    
    GD_CITIES = ['广州', '深圳', '东莞', '佛山', '中山', '珠海', '江门', '肇庆',
                 '惠州', '汕头', '潮州', '揭阳', '汕尾', '湛江', '茂名', '阳江',
                 '云浮', '韶关', '梅州', '河源', '清远']

    def __init__(self, parent, values, width=18, select_all=False):
        super().__init__(parent)
        self.values = values
        self.var_dict = {}
        
        # 下拉框变量
        self.var = tk.StringVar(value="")
        
        # 记录选择顺序（解决显示顺序问题）
        self._selected_order = []
        
        # 创建 Entry
        self.entry = ttk.Entry(self, textvariable=self.var, width=width, state='readonly')
        self.entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # 创建下拉按钮
        self.btn = ttk.Button(self, text="▼", width=3, command=self._toggle_dropdown)
        self.btn.pack(side=tk.LEFT)
        
        # 创建下拉窗口
        self.dropdown = tk.Toplevel(self)
        self.dropdown.withdraw()
        self.dropdown.overrideredirect(True)
        self.dropdown.attributes('-topmost', True)
        
        # 复选框容器
        self.check_frame = ttk.Frame(self.dropdown)
        self.check_frame.pack(fill=tk.BOTH, expand=True, padx=2, pady=2)
        
        self.check_vars = {}
        for val in values:
            var = tk.BooleanVar(value=False)
            self.check_vars[val] = var
            cb = ttk.Checkbutton(
                self.check_frame, text=val, variable=var,
                command=lambda v=val: self._on_check_change(v)
            )
            cb.pack(anchor=tk.W, padx=5, pady=1)
        
        # 全选按钮
        btn_frame = ttk.Frame(self.dropdown)
        btn_frame.pack(fill=tk.X, padx=2, pady=2)
        ttk.Button(btn_frame, text="全选", command=self._select_all).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="取消", command=self._deselect_all).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="确定", command=self._confirm).pack(side=tk.RIGHT, padx=2)
        
        # 初始全选
        if select_all:
            self._select_all()
    
    def _toggle_dropdown(self):
        """切换下拉框显示"""
        if self.dropdown.winfo_viewable():
            self.dropdown.withdraw()
        else:
            self._show_dropdown()
    
    def _show_dropdown(self):
        """显示下拉框（自适应屏幕空间）"""
        self.dropdown.update_idletasks()

        # 获取 Entry 的位置和尺寸
        entry_x = self.entry.winfo_rootx()
        entry_y = self.entry.winfo_rooty()
        entry_h = self.entry.winfo_height()

        # 获取下拉框的所需尺寸
        dropdown_w = self.dropdown.winfo_reqwidth()
        dropdown_h = self.dropdown.winfo_reqheight()

        # 获取屏幕工作区域尺寸（排除任务栏等）
        screen_w = self.dropdown.winfo_screenwidth()
        screen_h = self.dropdown.winfo_screenheight()

        # 计算下方和上方可用的空间
        space_below = screen_h - (entry_y + entry_h)
        space_above = entry_y

        # 判断应该显示在上方还是下方
        # 优先显示在下方，如果下方空间不够但上方够用，则显示在上方
        if space_below >= dropdown_h or space_below >= space_above:
            # 显示在下方
            y = entry_y + entry_h
        else:
            # 显示在上方
            y = entry_y - dropdown_h

        # 确保不会超出屏幕左右边界
        if entry_x + dropdown_w > screen_w:
            x = screen_w - dropdown_w
        else:
            x = entry_x

        # 确保不会超出屏幕上边界
        if y < 0:
            y = 0

        self.dropdown.geometry(f"{dropdown_w}x{dropdown_h}+{x}+{y}")
        self.dropdown.deiconify()
        self.dropdown.lift()
    
    def _on_check_change(self, val=None):
        """复选框状态变化 - 记录选择顺序"""
        if val is not None:
            var = self.check_vars.get(val)
            if var:
                if var.get():
                    # 选中：添加到顺序列表（避免重复）
                    if val not in self._selected_order:
                        self._selected_order.append(val)
                else:
                    # 取消选中：从顺序列表移除
                    if val in self._selected_order:
                        self._selected_order.remove(val)
    
    def _select_all(self):
        """全选"""
        self._selected_order = list(self.values)  # 按values顺序记录
        for var in self.check_vars.values():
            var.set(True)
    
    def _deselect_all(self):
        """取消全选"""
        self._selected_order = []
        for var in self.check_vars.values():
            var.set(False)
    
    def _confirm(self):
        """确认选择"""
        # 使用记录的选择顺序返回，而不是字典遍历顺序
        selected = [val for val in self._selected_order if val in self.check_vars and self.check_vars[val].get()]
        if selected:
            self.var.set(','.join(selected))
        else:
            self.var.set("")
        self.dropdown.withdraw()
    
    def get_selected(self):
        """获取选中的值列表"""
        return [val for val, var in self.check_vars.items() if var.get()]
    
    def set_selected(self, values):
        """设置选中的值"""
        # 先清空选择顺序
        self._selected_order = []
        for val, var in self.check_vars.items():
            var.set(val in values)
            if val in values:
                self._selected_order.append(val)
        if values:
            self.var.set(','.join(values))
        else:
            self.var.set("")
    
    def get_value(self):
        """获取选中值（逗号分隔字符串）"""
        return self.var.get()
    
    def set_value(self, value):
        """设置选中值（逗号分隔字符串）"""
        if value:
            values = [v.strip() for v in value.split(',')]
            self.set_selected(values)


# ==================== GUI应用 ====================
class UniversalExtractorGUI:
    """通用数据提取工具GUI"""

    def __init__(self, root, expiry_time=None):
        self.root = root
        self.root.title("通用数据提取工具")
        self.root.geometry("950x650")
        self.root.minsize(800, 600)
        self.root.resizable(True, True)
        
        # 授权过期时间
        self.expiry_time = expiry_time

        self.log_queue = queue.Queue()
        self.extractor = None
        self.is_logged_in = False

        self.create_widgets()
        self.root.update_idletasks()
        self._center_window()

        try:
            self.root.iconbitmap('icon.ico')
        except:
            pass

        self.setup_logging()
        self.update_log()
        self.load_config()
        self._update_license_display()

    def _update_license_display(self):
        """更新授权时间显示"""
        if self.expiry_time:
            try:
                # 解析过期时间
                expiry_dt = datetime.strptime(self.expiry_time, "%Y-%m-%d")
                display_time = expiry_dt.strftime("%Y-%m-%d")
                
                # 计算剩余天数
                current_dt = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
                days_left = (expiry_dt - current_dt).days
                
                if days_left < 0:
                    self.license_label.config(text="授权已过期", fg='#fce8e6')
                elif days_left <= 7:
                    self.license_label.config(text=f"授权到期: {display_time} (剩{days_left}天)", fg='#fce8e6')
                elif days_left <= 30:
                    self.license_label.config(text=f"授权到期: {display_time} (剩{days_left}天)", fg='#fff3e0')
                else:
                    self.license_label.config(text=f"授权到期: {display_time}", fg='#e8f5e9')
            except:
                self.license_label.config(text="", fg='#e8f5e9')

    def _center_window(self):
        """窗口居中显示"""
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        if width < 100:
            width = 1100
            height = 800
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        x = (screen_width // 2) - (width // 2)
        y = (screen_height // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')

    def setup_logging(self):
        """初始化日志系统"""
        ensure_dirs()
        log_filename = datetime.now().strftime("universal_gui_log_%Y%m%d.log")
        self.log_file_path = os.path.join(LOG_DIR, log_filename)

        set_log_file(self.log_file_path)

        self.logger = logging.getLogger('UniversalExtractorGUI')
        self.logger.setLevel(logging.DEBUG)
        self.logger.handlers.clear()

        file_handler = logging.FileHandler(self.log_file_path, encoding='utf-8')
        file_handler.setLevel(logging.DEBUG)

        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s',
                                      datefmt='%Y-%m-%d %H:%M:%S')
        file_handler.setFormatter(formatter)
        self.logger.addHandler(file_handler)

        self.logger.info("=" * 50)
        self.logger.info("通用数据提取工具 GUI 启动")
        self.logger.info(f"日志文件: {self.log_file_path}")
        self.logger.info("=" * 50)

    def create_widgets(self):
        """创建界面组件 - 使用现代化设计"""
        # 顶部蓝色标题栏
        self.header = tk.Frame(self.root, bg='#165DFF', height=60)
        self.header.pack(fill=tk.X)
        self.header.pack_propagate(False)

        # 标题栏左侧 - Logo和标题
        left_frame = tk.Frame(self.header, bg='#165DFF')
        left_frame.pack(side=tk.LEFT, padx=25, pady=12)

        icon_frame = tk.Frame(left_frame, bg='#1a6ce8', width=36, height=36)
        icon_frame.pack(side=tk.LEFT, padx=(0, 12))
        icon_frame.pack_propagate(False)
        icon_label = tk.Label(icon_frame, text="📊", font=('Segoe UI Emoji', 18),
                             bg='#1a6ce8', fg='white')
        icon_label.place(relx=0.5, rely=0.5, anchor='center')

        title_frame = tk.Frame(left_frame, bg='#165DFF')
        title_frame.pack(side=tk.LEFT)

        title = tk.Label(title_frame, text="通用数据提取工具",
                        font=('Microsoft YaHei UI', 18, 'bold'),
                        bg='#165DFF', fg='white')
        title.pack(anchor='w')

        version = tk.Label(title_frame, text="v2.0",
                          font=('Microsoft YaHei UI', 9),
                          bg='#1a6ce8', fg='white',
                          padx=8, pady=2)
        version.pack(anchor='w', pady=(2, 0))

        # 标题栏右侧 - 状态和授权时间
        self.right_frame = tk.Frame(self.header, bg='#165DFF')
        self.right_frame.pack(side=tk.RIGHT, padx=25, pady=12)

        # 授权过期时间标签
        self.license_label = tk.Label(self.right_frame, text="",
                              font=('Microsoft YaHei UI', 9),
                              bg='#165DFF', fg='#e0e7ff')
        self.license_label.pack(side=tk.LEFT, padx=(0, 15))

        # 解锁按钮 - 半透明白色
        self.unlock_btn = tk.Button(self.right_frame, text="🔓 系统解锁",
                             font=('Microsoft YaHei UI', 9, 'bold'),
                             bg='#3b82f6', fg='white', bd=0,
                             cursor='hand2', padx=12, pady=4)
        self.unlock_btn.pack(side=tk.LEFT, padx=(0, 15))

        self.status_dot = tk.Label(self.right_frame, text="●", font=('Arial', 14),
                            bg='#165DFF', fg='#a5b4fc')
        self.status_dot.pack(side=tk.LEFT)
        self.status_text = tk.Label(self.right_frame, text="系统就绪",
                              font=('Microsoft YaHei UI', 10),
                              bg='#165DFF', fg='white')
        self.status_text.pack(side=tk.LEFT, padx=(6, 0))

        # 主内容区域
        self.main_container = tk.Frame(self.root, bg='#f9fafb')
        self.main_container.pack(fill=tk.BOTH, expand=True)

        content = tk.Frame(self.main_container, bg='#f9fafb')
        content.pack(fill=tk.BOTH, expand=True, padx=20, pady=15)

        # 第一行：登录配置卡片（单行紧凑）
        self._build_login_card(content)

        # 第二行：查询参数卡片
        self._build_query_card(content)

        # 第三行：提取参数卡片
        self._build_params_card(content)

        # 底部：进度和日志
        self._build_bottom_section(content)

    def _build_card(self, parent, title, **kwargs):
        """创建卡片容器"""
        card = tk.Frame(parent, bg='white', bd=0, relief='flat')

        if title:
            header = tk.Frame(card, bg='white')
            header.pack(fill=tk.X, padx=20, pady=(16, 0))

            label = tk.Label(header, text=title,
                            font=('Microsoft YaHei UI', 13, 'bold'),
                            bg='white', fg='#374151', anchor='w')
            label.pack(fill='x')

            separator = tk.Frame(card, bg='#f3f4f6', height=1)
            separator.pack(fill=tk.X, padx=20, pady=(12, 0))

        return card

    def _build_query_card(self, parent):
        """构建查询参数卡片"""
        card = self._build_card(parent, "🔍 查询参数")
        card.pack(fill=tk.BOTH, expand=True, pady=(0, 10))

        body = tk.Frame(card, bg='white')
        body.pack(fill=tk.BOTH, expand=True, padx=16, pady=12)

        top_row = tk.Frame(body, bg='white')
        top_row.pack(fill=tk.BOTH, expand=True, pady=(0, 12))

        # ========== 左侧：数据分类（带滚动条）==========
        cat_outer = tk.Frame(top_row, bg='#f8f9fa', bd=1)
        cat_outer.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 15))

        tk.Label(cat_outer, text="数据分类", font=('Microsoft YaHei UI', 10, 'bold'),
                bg='#f8f9fa', fg='#202124', padx=12, pady=8).pack(anchor='w')

        # 创建可滚动区域
        cat_scroll_frame = tk.Frame(cat_outer, bg='#f8f9fa')
        cat_scroll_frame.pack(fill=tk.BOTH, expand=True, padx=8, pady=(0, 8))

        cat_canvas = tk.Canvas(cat_scroll_frame, bg='#f8f9fa', highlightthickness=0,
                               height=150)
        cat_scrollbar = ttk.Scrollbar(cat_scroll_frame, orient=tk.VERTICAL,
                                      command=cat_canvas.yview)
        cat_canvas.configure(yscrollcommand=cat_scrollbar.set)

        cat_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        cat_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        cat_inner = tk.Frame(cat_canvas, bg='#f8f9fa')
        cat_canvas.create_window((0, 0), window=cat_inner, anchor='nw')

        def on_cat_configure(event):
            cat_canvas.configure(scrollregion=cat_canvas.bbox('all'))

        cat_inner.bind('<Configure>', on_cat_configure)

        # 鼠标滚轮支持（跨平台）
        def on_cat_mousewheel(event):
            cat_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        def on_cat_bound_to_mousewheel(event):
            cat_canvas.bind_all('<MouseWheel>', on_cat_mousewheel)

        def on_cat_unbound_to_mousewheel(event):
            cat_canvas.unbind_all('<MouseWheel>')

        cat_canvas.bind('<Enter>', on_cat_bound_to_mousewheel)
        cat_canvas.bind('<Leave>', on_cat_unbound_to_mousewheel)

        self.category_vars = {}
        categories = [
            ("干扰", False),
            ("容量", False),
            ("语音报表", False),
            ("小区性能", False),
            ("全程完好率", False),
            ("语音小区", False)
        ]

        for name, checked in categories:
            var = tk.IntVar(value=int(checked))
            self.category_vars[name] = var
            cb = tk.Checkbutton(cat_inner, text=name, variable=var,
                               font=('Microsoft YaHei UI', 10, 'bold'),
                               bg='#f8f9fa', fg='#202124',
                               selectcolor='#165DFF',
                               activebackground='#f8f9fa',
                               activeforeground='#165DFF',
                               cursor='hand2',
                               command=lambda c=name: self._on_category_changed(c))
            cb.pack(anchor='w', pady=2)

        # ========== 右侧：数据表选择（带滚动条）==========
        table_outer = tk.Frame(top_row, bg='#f8f9fa', bd=1)
        table_outer.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        tk.Label(table_outer, text="选择数据表", font=('Microsoft YaHei UI', 10, 'bold'),
                bg='#f8f9fa', fg='#202124', padx=12, pady=8).pack(anchor='w')

        # 创建可滚动区域
        table_scroll_frame = tk.Frame(table_outer, bg='#f8f9fa')
        table_scroll_frame.pack(fill=tk.BOTH, expand=True, padx=8, pady=(0, 8))

        table_canvas = tk.Canvas(table_scroll_frame, bg='#f8f9fa', highlightthickness=0,
                                 height=150)
        table_scrollbar = ttk.Scrollbar(table_scroll_frame, orient=tk.VERTICAL,
                                        command=table_canvas.yview)
        table_canvas.configure(yscrollcommand=table_scrollbar.set)

        table_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        table_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        tables_grid = tk.Frame(table_canvas, bg='#f8f9fa')
        table_canvas.create_window((0, 0), window=tables_grid, anchor='nw')

        def on_table_configure(event):
            table_canvas.configure(scrollregion=table_canvas.bbox('all'))

        tables_grid.bind('<Configure>', on_table_configure)

        # 鼠标滚轮支持（跨平台）
        def on_table_mousewheel(event):
            table_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        def on_table_bound_to_mousewheel(event):
            table_canvas.bind_all('<MouseWheel>', on_table_mousewheel)

        def on_table_unbound_to_mousewheel(event):
            table_canvas.unbind_all('<MouseWheel>')

        table_canvas.bind('<Enter>', on_table_bound_to_mousewheel)
        table_canvas.bind('<Leave>', on_table_unbound_to_mousewheel)

        self.table_vars = {}
        TABLE_CATEGORIES = {
            '干扰': ['5G干扰小区', '4G干扰小区'],
            '容量': ['5G小区容量报表', '重要场景-天'],
            '语音报表': ['VoLTE小区监控预警', 'VONR小区监控预警', 'EPSFB小区监控预警'],
            '小区性能': ['5G小区性能KPI报表', '4G小区性能KPI报表'],
            '全程完好率': ['4G全程完好率报表', '5G全程完好率报表'],
            '语音小区': ['4G语音小区', '5G语音小区'],
        }

        all_tables = []
        for tables in TABLE_CATEGORIES.values():
            all_tables.extend(tables)

        for i, name in enumerate(all_tables):
            var = tk.IntVar(value=1 if i < 2 else 0)
            self.table_vars[name] = var
            cb = tk.Checkbutton(tables_grid, text=name, variable=var,
                               font=('Microsoft YaHei UI', 10),
                               bg='#f8f9fa', fg='#202124',
                               selectcolor='#165DFF',
                               activebackground='#f8f9fa',
                               activeforeground='#165DFF',
                               cursor='hand2')
            row, col = i // 3, i % 3
            cb.grid(row=row, column=col, sticky='w', padx=(0, 25), pady=2)

        # 全选按钮
        btn_frame = tk.Frame(table_outer, bg='#f8f9fa')
        btn_frame.pack(fill=tk.X, padx=8, pady=(8, 8))
        tk.Button(btn_frame, text="全选", font=('Microsoft YaHei UI', 9, 'bold'),
                 bg='#e8eaed', fg='#202124', bd=0, padx=14, pady=4,
                 cursor='hand2', command=self._select_all_tables).pack(side=tk.LEFT, padx=(0, 5))
        tk.Button(btn_frame, text="取消", font=('Microsoft YaHei UI', 9, 'bold'),
                 bg='#e8eaed', fg='#202124', bd=0, padx=14, pady=4,
                 cursor='hand2', command=self._deselect_all_tables).pack(side=tk.LEFT)

    def _build_login_card(self, parent):
        """构建登录配置卡片（一行紧凑布局）"""
        card = self._build_card(parent, "🔐 登录配置")
        card.pack(fill=tk.X, pady=(0, 10))

        body = tk.Frame(card, bg='white')
        body.pack(fill=tk.X, padx=16, pady=10)

        # 单行布局：用户名 | 密码 | 登录状态 | 按钮
        row = tk.Frame(body, bg='white')
        row.pack(fill=tk.X)

        # 用户名
        user_frame = tk.Frame(row, bg='white')
        user_frame.pack(side=tk.LEFT, padx=(0, 12))

        tk.Label(user_frame, text="用户名", font=('Microsoft YaHei UI', 8),
                bg='white', fg='#5f6368').pack(anchor='w')
        self.username_entry = tk.Entry(user_frame, font=('Microsoft YaHei UI', 10),
                             relief='flat', bg='#f8f9fa', bd=0, width=15)
        self.username_entry.insert(0, DEFAULT_USERNAME)
        self.username_entry.pack(fill=tk.X, pady=(2, 0), ipady=4)

        # 密码
        pass_frame = tk.Frame(row, bg='white')
        pass_frame.pack(side=tk.LEFT, padx=(0, 12))

        tk.Label(pass_frame, text="密码", font=('Microsoft YaHei UI', 8),
                bg='white', fg='#5f6368').pack(anchor='w')
        self.password_entry = tk.Entry(pass_frame, font=('Microsoft YaHei UI', 10),
                             show="●", relief='flat', bg='#f8f9fa', bd=0, width=15)
        self.password_entry.insert(0, DEFAULT_PASSWORD)
        self.password_entry.pack(fill=tk.X, pady=(2, 0), ipady=4)

        # 登录状态图标和标签
        self.login_status_icon = tk.Label(row, text="○", font=('Arial', 12, 'bold'),
                              bg='white', fg='#80868b')
        self.login_status_icon.pack(side=tk.LEFT, padx=(10, 4), pady=0)

        self.login_status_lbl = tk.Label(row, text="未登录",
                             font=('Microsoft YaHei UI', 10, 'bold'),
                             bg='white', fg='#80868b')
        self.login_status_lbl.pack(side=tk.LEFT, padx=(0, 10), pady=0)

        # 登录按钮
        self.login_btn = tk.Button(row, text="登录",
                             font=('Microsoft YaHei UI', 10, 'bold'),
                             bg='#165DFF', fg='white', bd=0,
                             cursor='hand2', padx=20, pady=6,
                             command=self.login)
        self.login_btn.pack(side=tk.LEFT)

    def _build_params_card(self, parent):
        """构建提取参数卡片（紧凑布局）"""
        card = self._build_card(parent, "⚙ 提取参数")
        card.pack(fill=tk.X, pady=(0, 10))

        body = tk.Frame(card, bg='white')
        body.pack(fill=tk.X, padx=16, pady=8)

        # 第一行：地市选择 + 快捷日期 + 日期范围
        top_row = tk.Frame(body, bg='white')
        top_row.pack(fill=tk.X, pady=(0, 6))

        # 地市选择
        city_frame = tk.Frame(top_row, bg='white')
        city_frame.pack(side=tk.LEFT, padx=(0, 15))
        tk.Label(city_frame, text="地市选择", font=('Microsoft YaHei UI', 8),
                bg='white', fg='#5f6368').pack(anchor='w')

        self.city_dropdown = MultiSelectDropdown(
            city_frame,
            MultiSelectDropdown.GD_CITIES,
            width=12,
            select_all=False
        )
        self.city_dropdown.pack(pady=(2, 0))
        self.city_dropdown.set_selected(['阳江'])

        # 快捷日期
        quick_frame = tk.Frame(top_row, bg='white')
        quick_frame.pack(side=tk.LEFT, padx=(0, 15))
        tk.Label(quick_frame, text="快捷日期", font=('Microsoft YaHei UI', 8),
                bg='white', fg='#5f6368').pack(anchor='w')

        quick_inner = tk.Frame(quick_frame, bg='white')
        quick_inner.pack(pady=(2, 0))

        self.quick_date_btns = {}
        for text, days in [("昨天", 1), ("近7天", 7), ("近30天", 30)]:
            btn = tk.Button(quick_inner, text=text, font=('Microsoft YaHei UI', 8, 'bold'),
                           bg='#e8eaed', fg='#202124', bd=0, padx=10, pady=2,
                           cursor='hand2', relief='flat',
                           command=lambda d=days: self.set_quick_date(d))
            btn.pack(side=tk.LEFT, padx=(0, 3))
            self.quick_date_btns[days] = btn

        # 日期范围
        date_frame = tk.Frame(top_row, bg='white')
        date_frame.pack(side=tk.LEFT, padx=(0, 15))
        tk.Label(date_frame, text="日期范围", font=('Microsoft YaHei UI', 8),
                bg='white', fg='#5f6368').pack(anchor='w')

        date_inner = tk.Frame(date_frame, bg='white')
        date_inner.pack(pady=(2, 0))

        self.start_year_var = tk.IntVar(value=datetime.now().year)
        self.start_month_var = tk.IntVar(value=datetime.now().month)
        self.start_day_var = tk.IntVar(value=1)

        yesterday = datetime.now() - timedelta(days=1)
        self.end_year_var = tk.IntVar(value=yesterday.year)
        self.end_month_var = tk.IntVar(value=yesterday.month)
        self.end_day_var = tk.IntVar(value=yesterday.day)

        start_frame = tk.Frame(date_inner, bg='white')
        start_frame.pack(side=tk.LEFT)

        current_year = datetime.now().year
        ttk.Combobox(start_frame, textvariable=self.start_year_var,
                   values=list(range(2020, current_year + 1)),
                   width=4, state="readonly").pack(side=tk.LEFT)
        tk.Label(start_frame, text="-", font=('Microsoft YaHei UI', 8),
                bg='white', fg='#5f6368').pack(side=tk.LEFT, padx=1)
        ttk.Combobox(start_frame, textvariable=self.start_month_var,
                   values=list(range(1, 13)),
                   width=2, state="readonly").pack(side=tk.LEFT)
        tk.Label(start_frame, text="-", font=('Microsoft YaHei UI', 8),
                bg='white', fg='#5f6368').pack(side=tk.LEFT, padx=1)
        ttk.Combobox(start_frame, textvariable=self.start_day_var,
                   values=list(range(1, 32)),
                   width=2, state="readonly").pack(side=tk.LEFT)

        tk.Label(date_inner, text=" 至 ", font=('Microsoft YaHei UI', 8),
                bg='white', fg='#5f6368').pack(side=tk.LEFT, padx=3)

        end_frame = tk.Frame(date_inner, bg='white')
        end_frame.pack(side=tk.LEFT)

        ttk.Combobox(end_frame, textvariable=self.end_year_var,
                   values=list(range(2020, current_year + 1)),
                   width=4, state="readonly").pack(side=tk.LEFT)
        tk.Label(end_frame, text="-", font=('Microsoft YaHei UI', 8),
                bg='white', fg='#5f6368').pack(side=tk.LEFT, padx=1)
        ttk.Combobox(end_frame, textvariable=self.end_month_var,
                   values=list(range(1, 13)),
                   width=2, state="readonly").pack(side=tk.LEFT)
        tk.Label(end_frame, text="-", font=('Microsoft YaHei UI', 8),
                bg='white', fg='#5f6368').pack(side=tk.LEFT, padx=1)
        ttk.Combobox(end_frame, textvariable=self.end_day_var,
                   values=list(range(1, 32)),
                   width=2, state="readonly").pack(side=tk.LEFT)

        # 多日提取模式选项
        mode_frame = tk.Frame(top_row, bg='white')
        mode_frame.pack(side=tk.LEFT, padx=(0, 0))

        self.multi_day_var = tk.BooleanVar(value=False)
        multi_day_cb = tk.Checkbutton(mode_frame, text="多日模式",
                                      variable=self.multi_day_var,
                                      font=('Microsoft YaHei UI', 8),
                                      bg='white', fg='#202124',
                                      selectcolor='#e8f0fe',
                                      activebackground='white',
                                      command=self._on_multi_day_toggle)
        multi_day_cb.pack(side=tk.LEFT)

        self.multi_day_per_sheet_var = tk.BooleanVar(value=False)
        multi_day_per_sheet_cb = tk.Checkbutton(mode_frame, text="按日分Sheet",
                                               variable=self.multi_day_per_sheet_var,
                                               font=('Microsoft YaHei UI', 8),
                                               bg='white', fg='#202124',
                                               selectcolor='#e8f0fe',
                                               activebackground='white',
                                               state=tk.DISABLED,
                                               command=self._on_multi_day_per_sheet_toggle)
        self.multi_day_per_sheet_cb = multi_day_per_sheet_cb
        multi_day_per_sheet_cb.pack(side=tk.LEFT, padx=(6, 0))

        # 第二行：操作按钮
        btn_row = tk.Frame(body, bg='white')
        btn_row.pack(fill=tk.X, pady=(4, 0))

        self.extract_btn = tk.Button(btn_row, text="▶ 开始提取",
                               font=('Microsoft YaHei UI', 10, 'bold'),
                               bg='#165DFF', fg='white', bd=0,
                               cursor='hand2', padx=22, pady=5,
                               state=tk.DISABLED, command=self.start_extract)
        self.extract_btn.pack(side=tk.LEFT)

        self.stop_btn = tk.Button(btn_row, text="⏹ 停止",
                            font=('Microsoft YaHei UI', 9),
                            bg='#dc3545', fg='white', bd=0,
                            cursor='hand2', padx=14, pady=5,
                            state=tk.DISABLED, command=self.stop_extract)
        self.stop_btn.pack(side=tk.LEFT, padx=(8, 0))

        tk.Button(btn_row, text="📁 打开目录",
                 font=('Microsoft YaHei UI', 9),
                 bg='#f0f2f5', fg='#202124', bd=0,
                 cursor='hand2', padx=12, pady=5,
                 command=self.open_output_dir).pack(side=tk.RIGHT)

    def _build_bottom_section(self, parent):
        """构建底部日志区域"""
        # 创建滚动容器
        scroll_container = tk.Frame(parent)
        scroll_container.pack(fill=tk.BOTH, expand=True, pady=(0, 0))

        # 创建 Canvas 和滚动条
        bottom_canvas = tk.Canvas(scroll_container, bg='white', highlightthickness=0,
                                  height=200)
        bottom_scrollbar = ttk.Scrollbar(scroll_container, orient=tk.VERTICAL,
                                         command=bottom_canvas.yview)
        bottom_canvas.configure(yscrollcommand=bottom_scrollbar.set)

        bottom_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        bottom_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Canvas 内部的可滚动 Frame
        bottom = tk.Frame(bottom_canvas, bg='white')
        bottom_canvas.create_window((0, 0), window=bottom, anchor='nw')

        def on_bottom_configure(event):
            bottom_canvas.configure(scrollregion=bottom_canvas.bbox('all'))

        bottom.bind('<Configure>', on_bottom_configure)

        # 鼠标滚轮支持
        def on_mousewheel(event):
            bottom_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        def bound_to_mousewheel(event):
            bottom_canvas.bind_all('<MouseWheel>', on_mousewheel)

        def unbound_to_mousewheel(event):
            bottom_canvas.unbind_all('<MouseWheel>')

        bottom_canvas.bind('<Enter>', bound_to_mousewheel)
        bottom_canvas.bind('<Leave>', unbound_to_mousewheel)

        progress_area = tk.Frame(bottom, bg='white')
        progress_area.pack(fill=tk.X, padx=16, pady=(10, 6))

        progress_info = tk.Frame(progress_area, bg='white')
        progress_info.pack(fill=tk.X)

        self.progress_lbl_pct = tk.Label(progress_info, text="进度: 0%",
                          font=('Microsoft YaHei UI', 10, 'bold'),
                          bg='white', fg='#165DFF')
        self.progress_lbl_pct.pack(side=tk.LEFT)

        self.progress_lbl_detail = tk.Label(progress_info, text="就绪",
                             font=('Microsoft YaHei UI', 9),
                             bg='white', fg='#5f6368')
        self.progress_lbl_detail.pack(side=tk.RIGHT)

        # 圆角进度条
        self.progress_canvas = tk.Canvas(progress_area, height=8, bg='white', highlightthickness=0)
        self.progress_canvas.pack(fill=tk.X, pady=(6, 0))
        self.progress_bar = self.progress_canvas.create_rectangle(0, 0, 0, 8, fill='#165DFF', outline='')
        self.progress_bg = self.progress_canvas.create_rectangle(0, 0, 1000, 8, fill='#f0f2f5', outline='')
        self.progress_canvas.bind('<Configure>', lambda e: self.progress_canvas.coords(self.progress_bg, 0, 0, e.width, 8))

        log_header = tk.Frame(bottom, bg='white')
        log_header.pack(fill=tk.X, padx=16)

        tk.Label(log_header, text="📋 运行日志",
                font=('Microsoft YaHei UI', 10, 'bold'),
                bg='white', fg='#202124').pack(side=tk.LEFT)

        tk.Button(log_header, text="清空", font=('Microsoft YaHei UI', 9, 'bold'),
                 bg='#e8eaed', fg='#202124', bd=0,
                 cursor='hand2', padx=10, pady=2,
                 command=self.clear_log).pack(side=tk.RIGHT)

        # 日志框架 - 关键修改：确保填充并扩展，添加滚动条支持
        log_frame = tk.Frame(bottom, bg='#1e1e1e')
        log_frame.pack(fill=tk.BOTH, expand=True, padx=16, pady=(4, 8))

        # 创建日志文本框和滚动条
        log_scroll = ttk.Scrollbar(log_frame, orient=tk.VERTICAL)
        log_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        self.log_text = tk.Text(log_frame, wrap=tk.WORD, font=('Consolas', 10),
                          bg='#1e1e1e', fg='#4ec9b0', bd=0,
                          padx=12, pady=6, state='disabled',
                          insertbackground='white',
                          yscrollcommand=log_scroll.set)
        self.log_text.pack(fill=tk.BOTH, expand=True, side=tk.LEFT)
        log_scroll.config(command=self.log_text.yview)

        self.log_text.tag_config("#3c4043", foreground="#9aa0a6")
        self.log_text.tag_config("#34a853", foreground="#34a853")
        self.log_text.tag_config("#80868b", foreground="#80868b")
        self.log_text.tag_config("#165DFF", foreground="#6ab7ff")
        self.log_text.tag_config("#ea4335", foreground="#f28b82")

        self.status_bar = tk.Frame(bottom, bg='#f8f9fa', height=28)
        self.status_bar.pack(fill=tk.X)
        self.status_bar.pack_propagate(False)

        self.status_var = tk.StringVar(value=f"✓ 数据输出目录: {os.path.abspath(OUTPUT_DIR)}/")
        tk.Label(self.status_bar, textvariable=self.status_var,
                font=('Microsoft YaHei UI', 8, 'bold'),
                bg='#f8f9fa', fg='#202124').pack(side=tk.LEFT, padx=16, pady=6)

        tk.Label(self.status_bar, text="v2.0.1 | © 2026",
                font=('Microsoft YaHei UI', 7),
                bg='#f8f9fa', fg='#bdc1c6').pack(side=tk.RIGHT, padx=16, pady=6)

    def update_progress_bar(self, value, label=""):
        """更新进度条"""
        self.progress_lbl_pct.config(text=f"进度: {int(value)}%")
        self.progress_canvas.coords(self.progress_bar, 0, 0, max(0, self.progress_canvas.winfo_width() * value / 100), 8)
        if label:
            self.progress_lbl_detail.config(text=label)

    def _on_category_changed(self, category_name):
        """分类复选框改变时的回调"""
        TABLE_CATEGORIES = {
            '干扰': ['5G干扰小区', '4G干扰小区'],
            '容量': ['5G小区容量报表', '重要场景-天'],
            '语音报表': ['VoLTE小区监控预警', 'VONR小区监控预警', 'EPSFB小区监控预警'],
            '小区性能': ['5G小区性能KPI报表', '4G小区性能KPI报表'],
            '全程完好率': ['4G全程完好率报表', '5G全程完好率报表'],
            '语音小区': ['4G语音小区', '5G语音小区'],
        }

        is_checked = self.category_vars[category_name].get()
        tables = TABLE_CATEGORIES.get(category_name, [])

        for table_name in tables:
            if table_name in self.table_vars:
                self.table_vars[table_name].set(is_checked)

    def _select_all_tables(self):
        """全选所有数据表"""
        for var in self.table_vars.values():
            var.set(True)

    def _deselect_all_tables(self):
        """取消全选"""
        for var in self.table_vars.values():
            var.set(False)

    def _on_multi_day_toggle(self):
        """多日模式切换处理"""
        if self.multi_day_var.get():
            self.log("已启用多日模式，将按日期逐天提取数据", "INFO")
            # 启用按日分Sheet选项
            self.multi_day_per_sheet_cb.config(state=tk.NORMAL)
        else:
            self.log("已关闭多日模式", "INFO")
            # 禁用并取消按日分Sheet选项
            self.multi_day_per_sheet_var.set(False)
            self.multi_day_per_sheet_cb.config(state=tk.DISABLED)

    def _on_multi_day_per_sheet_toggle(self):
        """按日分Sheet模式切换处理"""
        if self.multi_day_per_sheet_var.get():
            self.log("已启用按日分Sheet模式，每天数据将保存在独立的Sheet中", "INFO")
        else:
            self.log("已关闭按日分Sheet模式", "INFO")

    def set_quick_date(self, days):
        """设置快捷日期"""
        yesterday = datetime.now() - timedelta(days=1)
        self.end_year_var.set(yesterday.year)
        self.end_month_var.set(yesterday.month)
        self.end_day_var.set(yesterday.day)

        start_date = yesterday - timedelta(days=days - 1)
        self.start_year_var.set(start_date.year)
        self.start_month_var.set(start_date.month)
        self.start_day_var.set(start_date.day)

        # 更新按钮选中状态
        for d, btn in self.quick_date_btns.items():
            if d == days:
                btn.config(bg='#165DFF', fg='white')
            else:
                btn.config(bg='#e8eaed', fg='#202124')

    def get_selected_tables(self):
        """获取选中的数据表列表"""
        return [name for name, var in self.table_vars.items() if var.get()]

    def validate_inputs(self):
        """验证输入"""
        selected_tables = self.get_selected_tables()
        if not selected_tables:
            messagebox.showwarning("警告", "请至少选择一个数据表")
            return False

        city = self.city_dropdown.get_value()
        if not city:
            messagebox.showwarning("警告", "请选择至少一个地市")
            return False

        start_date = self.get_date_string(self.start_year_var, self.start_month_var, self.start_day_var)
        end_date = self.get_date_string(self.end_year_var, self.end_month_var, self.end_day_var)

        if not start_date or not end_date:
            messagebox.showwarning("警告", "请输入有效的日期")
            return False

        try:
            start_dt = datetime.strptime(start_date, "%Y-%m-%d")
            end_dt = datetime.strptime(end_date, "%Y-%m-%d")
            if start_dt > end_dt:
                messagebox.showwarning("警告", "开始日期不能晚于结束日期")
                return False
        except ValueError:
            messagebox.showwarning("警告", "日期格式错误")
            return False

        return True

    def login(self):
        """执行登录"""
        if self.is_logged_in:
            messagebox.showinfo("提示", "已经登录，无需重复登录")
            return

        username = self.username_entry.get().strip()
        password = self.password_entry.get().strip()

        if not username or not password:
            messagebox.showwarning("警告", "请输入用户名和密码")
            return

        self.login_btn.config(state=tk.DISABLED, text="登录中...")
        self.update_login_status("登录中...", "pending")

        thread = threading.Thread(target=self._login_thread, args=(username, password))
        thread.daemon = True
        thread.start()

    def _login_thread(self, username, password):
        """登录线程"""
        try:
            self.extractor = CustomDataExtractor(username, password, self.root)

            import builtins
            original_print = print

            def custom_print(*args, **kwargs):
                message = ' '.join(map(str, args))
                if '[DEBUG' in message:
                    self.log(message, "DEBUG")
                elif '✓' in message or '成功' in message:
                    self.log(message, "SUCCESS")
                elif '✗' in message or '失败' in message or '错误' in message:
                    self.log(message, "ERROR")
                elif '⚠' in message or '警告' in message or 'WARNING' in message:
                    self.log(message, "WARNING")
                else:
                    self.log(message, "INFO")

            builtins.print = custom_print

            self.update_progress_bar(30, "正在连接服务器...")
            success = self.extractor.login()

            builtins.print = original_print

            if success:
                self.update_progress_bar(60, "登录成功，初始化即席查询...")

                builtins.print = custom_print
                jxcx_success = self.extractor.init_jxcx()
                builtins.print = original_print

                if jxcx_success:
                    self.is_logged_in = True
                    self.update_progress_bar(100, "登录完成")
                    self.log("登录成功！", "SUCCESS")
                    self.root.after(0, self._login_success_ui)
                else:
                    self.log("初始化即席查询失败", "ERROR")
                    self.root.after(0, self._login_failed_ui)
            else:
                self.log("登录失败", "ERROR")
                self.root.after(0, self._login_failed_ui)

        except Exception as e:
            self.log(f"登录异常: {e}", "ERROR")
            import traceback
            self.log(traceback.format_exc(), "ERROR")
            self.root.after(0, self._login_failed_ui)

    def _login_success_ui(self):
        """登录成功UI"""
        self.login_btn.config(state=tk.DISABLED, text="已登录")
        self.extract_btn.config(state=tk.NORMAL)
        self.update_login_status("已登录 ✓", "success")
        self.status_text.config(text="登录成功")
        messagebox.showinfo("成功", "登录成功！")

    def _login_failed_ui(self):
        """登录失败UI"""
        self.login_btn.config(state=tk.NORMAL, text="登录")
        self.update_login_status("登录失败", "error")
        self.update_progress_bar(0, "登录失败")
        self.status_text.config(text="登录失败")
        messagebox.showerror("失败", "登录失败，请检查账号密码或网络连接")

    def update_login_status(self, text, status="normal"):
        """更新登录状态显示"""
        colors = {
            "success": ("✓", "#e6f4ea", "#34a853"),
            "error": ("✗", "#fce8e6", "#ea4335"),
            "pending": ("...", "#fff3e0", "#f57c00"),
            "normal": ("○", "#f0f2f5", "#80868b")
        }
        icon, bg, fg = colors.get(status, colors["normal"])
        self.login_status_icon.config(text=icon, bg=bg, fg=fg)
        self.login_status_lbl.config(text=text, bg=bg, fg=fg)
        self.status_bar.config(bg=bg)
        self.login_status_icon.config(bg=bg)
        self.login_status_lbl.config(bg=bg)

    def start_extract(self):
        """开始提取"""
        if not self.is_logged_in:
            messagebox.showwarning("警告", "请先登录")
            return

        if not self.validate_inputs():
            return

        self.extract_btn.config(state=tk.DISABLED)
        self.stop_btn.config(state=tk.NORMAL)
        self.status_text.config(text="正在提取数据...")

        thread = threading.Thread(target=self._extract_thread)
        thread.daemon = True
        thread.start()

    def _extract_thread(self):
        """提取数据线程 - 支持多表批量提取和多日模式"""
        import copy
        try:
            selected_tables = self.get_selected_tables()
            city = self.city_dropdown.get_value()
            start_date = self.get_date_string(self.start_year_var, self.start_month_var, self.start_day_var)
            end_date = self.get_date_string(self.end_year_var, self.end_month_var, self.end_day_var)
            
            # 检查查询日期是否超过授权到期日期
            if self.expiry_time:
                try:
                    expiry_dt = datetime.strptime(self.expiry_time, "%Y-%m-%d")
                    start_dt = datetime.strptime(start_date, "%Y-%m-%d")
                    end_dt = datetime.strptime(end_date, "%Y-%m-%d")
                    
                    if end_dt > expiry_dt:
                        self.log(f"[ERROR] 查询日期超过授权到期日期 ({self.expiry_time})，拒绝执行", "ERROR")
                        self.root.after(0, lambda: self._show_query_blocked_dialog(start_date, end_date))
                        self.root.after(0, lambda: self._extract_failed_ui())
                        return
                except ValueError:
                    pass
            
            # 检查是否启用多日模式
            is_multi_day = self.multi_day_var.get()
            is_per_day_sheet = self.multi_day_per_sheet_var.get() if is_multi_day else False

            total_tables = len(selected_tables)
            self.log("=" * 50, "INFO")
            self.log(f"开始批量提取数据，共 {total_tables} 个数据表", "INFO")
            self.log(f"地市: {city}", "INFO")
            self.log(f"日期范围: {start_date} 至 {end_date}", "INFO")
            if is_multi_day:
                if is_per_day_sheet:
                    self.log(f"模式: 多日模式（按日分Sheet保存）", "INFO")
                else:
                    self.log(f"模式: 多日模式（按日期逐天提取，合并为一个文件）", "INFO")
            self.log("=" * 50, "INFO")

            success_count = 0
            failed_count = 0

            for idx, table_name in enumerate(selected_tables):
                self.log("", "INFO")
                self.log(f"--- [{idx + 1}/{total_tables}] 正在处理: {table_name} ---", "INFO")

                table_config = TableConfig.get_table_config(table_name)
                if not table_config:
                    self.log(f"✗ 未找到数据表配置: {table_name}", "ERROR")
                    failed_count += 1
                    continue

                # 4G语音小区不支持按日分Sheet模式，需要统一保存为三个Sheet
                is_4g_voice = table_config.get('calc_columns', []) == ['4G语音小区']
                effective_per_day_sheet = is_per_day_sheet and not is_4g_voice
                if is_4g_voice and is_per_day_sheet:
                    self.log(f"  (注: 4G语音小区不支持按日分Sheet，将合并为一个文件)", "INFO")

                if is_multi_day:
                    # 多日模式：按日期逐天提取
                    try:
                        start_dt = datetime.strptime(start_date, "%Y-%m-%d")
                        end_dt = datetime.strptime(end_date, "%Y-%m-%d")
                        current_dt = start_dt
                        
                        day_count = (end_dt - start_dt).days + 1
                        day_num = 0
                        
                        # 用于收集所有日期的数据（合并模式）
                        all_dfs = []
                        total_rows = 0
                        
                        # 用于按日分Sheet模式：直接写入Excel，无需先收集到内存
                        per_day_filepath = None
                        saved_days_count = 0
                        if effective_per_day_sheet:
                            ensure_dirs()
                            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                            filename = f'{table_name}_{city}_{start_date}_{end_date}_{timestamp}.xlsx'
                            per_day_filepath = os.path.join(OUTPUT_DIR, filename)
                            # 预创建文件（使用空的 DataFrame）
                            pd.DataFrame().to_excel(per_day_filepath, index=False, engine='openpyxl', sheet_name='_init_')
                        
                        while current_dt <= end_dt:
                            day_num += 1
                            current_date_str = current_dt.strftime("%Y-%m-%d")
                            self.log(f"  >>> 正在提取日期: {current_date_str} ({day_num}/{day_count})", "INFO")
                            
                            self.update_progress_bar(
                                int((idx / total_tables) * 100), 
                                f"正在提取 {table_name} {current_date_str}..."
                            )
                            
                            try:
                                payload = self.extractor.get_payload_with_dynamic_fields(
                                    table_config, current_date_str, current_date_str, city
                                )
                                
                                if not payload:
                                    self.log(f"  ✗ 生成payload失败: {current_date_str}", "ERROR")
                                    current_dt += timedelta(days=1)
                                    continue
                                
                                # 检查是否是特殊报表（如4G语音小区需要分别获取多个表）
                                if isinstance(payload, dict) and payload.get('__special__') == '4G_VOICE_MERGE':
                                    self.log(f"  检测到4G语音小区报表，将分别获取VoLTE和EPSFB数据...", "INFO")
                                    # 生成VoLTE和EPSFB的payload
                                    volte_payload = get_volte_payload()
                                    volte_payload = set_payload_time(volte_payload, current_date_str + ' 00:00:00', current_date_str + ' 23:59:59')
                                    volte_payload = set_payload_city(volte_payload, city)

                                    epsfb_payload = get_epsfb_payload()
                                    epsfb_payload = set_payload_time(epsfb_payload, current_date_str + ' 00:00:00', current_date_str + ' 23:59:59')
                                    epsfb_payload = set_payload_city(epsfb_payload, city)

                                    voice_data = self.extractor.jxcx._get_4g_voice_table_internal(volte_payload, epsfb_payload)
                                    volte_df = voice_data['volte']
                                    epsfb_df = voice_data['epsfb']
                                    merged_df = voice_data['merged']

                                    # 4G语音小区：收集数据用于最后统一保存
                                    if 'volte_dfs' not in locals():
                                        volte_dfs = []
                                        epsfb_dfs = []
                                        merged_dfs = []
                                    if not volte_df.empty:
                                        volte_dfs.append(volte_df)
                                    if not epsfb_df.empty:
                                        epsfb_dfs.append(epsfb_df)
                                    if not merged_df.empty:
                                        merged_dfs.append(merged_df)

                                    df = merged_df  # 用于后续判断是否为空
                                else:
                                    self.update_progress_bar(
                                        int((idx / total_tables) * 100) + 5,
                                        f"正在查询 {current_date_str}..."
                                    )
                                    df = self.extractor.jxcx.get_table(payload, to_df=True)

                                if df.empty:
                                    self.log(f"  ⚠ {current_date_str}: 未查询到数据", "WARNING")
                                else:
                                    # 处理计算列
                                    calc_columns = table_config.get('calc_columns', [])
                                    if calc_columns:
                                        try:
                                            if '4G全程完好率' in calc_columns:
                                                df = add_wanchenglv_columns(df, '4g')
                                            elif '5G全程完好率' in calc_columns:
                                                df = add_wanchenglv_columns(df, '5g')
                                            elif '5G语音小区' in calc_columns:
                                                df = add_5g_voice_columns(df)
                                        except Exception as e:
                                            import traceback
                                            error_detail = traceback.format_exc()
                                            self.log(f"  ⚠ 添加计算列异常: {e}", "WARNING")
                                            self.log(f"  ⚠ 详细错误: {error_detail[:500]}", "WARNING")
                                            # 打印DataFrame列信息帮助诊断
                                            if not df.empty:
                                                self.log(f"  ⚠ DataFrame列名: {list(df.columns)}", "WARNING")
                                            else:
                                                self.log(f"  ⚠ DataFrame为空，无法添加计算列", "WARNING")

                                    if is_per_day_sheet:
                                        # 按日分Sheet模式：查完即写，无需收集到内存
                                        sheet_name = current_date_str.replace('-', '')
                                        self.extractor._append_df_to_excel(per_day_filepath, df, sheet_name)
                                        saved_days_count += 1
                                    else:
                                        # 合并模式：收集每天的数据
                                        all_dfs.append(df)
                                    total_rows += len(df)
                                    self.log(f"  ✓ [{current_date_str}] 提取成功，共 {len(df)} 条记录", "SUCCESS")
                                    
                            except Exception as e:
                                self.log(f"  ✗ [{current_date_str}] 提取异常: {e}", "ERROR")
                            
                            current_dt += timedelta(days=1)

                        # 检查是否是4G语音小区（特殊处理）
                        is_4g_voice_table = table_config.get('calc_columns', []) == ['4G语音小区']

                        # 按日分Sheet模式：清理并完成
                        if effective_per_day_sheet and saved_days_count > 0:
                            self.extractor._finalize_per_day_excel(per_day_filepath, saved_days_count)
                            self.log(f"  ✓ 按日分Sheet保存成功！共 {total_rows} 条记录，{saved_days_count} 个Sheet", "SUCCESS")
                            self.log(f"    文件已保存: {per_day_filepath}", "SUCCESS")
                            success_count += 1
                        elif effective_per_day_sheet and saved_days_count == 0:
                            # 所有日期都无数据，删除空文件
                            if per_day_filepath and os.path.exists(per_day_filepath):
                                os.remove(per_day_filepath)
                            self.log(f"  ⚠ 所有日期均未查询到数据", "WARNING")
                        elif all_dfs:
                            # 4G语音小区：统一保存为三个Sheet
                            if is_4g_voice_table and 'volte_dfs' in locals() and volte_dfs:
                                self.update_progress_bar(
                                    int(((idx + 0.8) / total_tables) * 100),
                                    "正在合并并保存4G语音小区数据..."
                                )
                                combined_volte = pd.concat(volte_dfs, ignore_index=True) if volte_dfs else pd.DataFrame()
                                combined_epsfb = pd.concat(epsfb_dfs, ignore_index=True) if epsfb_dfs else pd.DataFrame()
                                combined_merged = pd.concat(merged_dfs, ignore_index=True) if merged_dfs else pd.DataFrame()

                                # 添加计算列
                                if not combined_merged.empty:
                                    try:
                                        combined_merged = add_4g_voice_columns(combined_merged)
                                    except Exception as e:
                                        self.log(f"  ⚠ 添加计算列异常: {e}", "WARNING")

                                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                                filename = f'{table_name}_{city}_{start_date}_{end_date}_{timestamp}.xlsx'
                                filepath = self.extractor.save_4g_voice_to_excel(combined_volte, combined_epsfb, combined_merged, filename)

                                total_rows = len(combined_volte) + len(combined_epsfb) + len(combined_merged)
                                if filepath:
                                    self.log(f"  ✓ 4G语音小区多日数据合并成功！共 {total_rows} 条记录", "SUCCESS")
                                    self.log(f"    文件已保存: {filepath}", "SUCCESS")
                                    success_count += 1
                                else:
                                    self.log(f"  ✗ 保存文件失败", "ERROR")
                                    failed_count += 1
                            else:
                                # 普通表：合并所有日期的数据并保存
                                self.update_progress_bar(
                                    int(((idx + 0.8) / total_tables) * 100),
                                    "正在合并并保存数据..."
                                )

                                combined_df = pd.concat(all_dfs, ignore_index=True)
                                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                                filename = f'{table_name}_{city}_{start_date}_{end_date}_{timestamp}.xlsx'
                                filepath = self.extractor.save_to_excel(combined_df, filename)

                                if filepath:
                                    self.log(f"  ✓ 多日数据合并成功！共 {total_rows} 条记录", "SUCCESS")
                                    self.log(f"    文件已保存: {filepath}", "SUCCESS")
                                    success_count += 1
                                else:
                                    self.log(f"  ✗ 保存文件失败", "ERROR")
                                    failed_count += 1
                        else:
                            self.log(f"  ⚠ 所有日期均未查询到数据", "WARNING")
                            
                    except Exception as e:
                        self.log(f"✗ [{table_name}] 多日模式处理异常: {e}", "ERROR")
                        failed_count += 1
                else:
                    # 普通模式：直接提取整个日期范围
                    self.update_progress_bar(int((idx / total_tables) * 100), f"正在提取 {table_name}...")

                    try:
                        payload = self.extractor.get_payload_with_dynamic_fields(
                            table_config, start_date, end_date, city
                        )

                        if not payload:
                            self.log(f"✗ 生成payload失败: {table_name}", "ERROR")
                            failed_count += 1
                            continue

                        # 检查是否是特殊报表（如4G语音小区需要分别获取多个表）
                        if isinstance(payload, dict) and payload.get('__special__') == '4G_VOICE_MERGE':
                            self.log(f"检测到4G语音小区报表，将分别获取VoLTE和EPSFB数据...", "INFO")
                            # 生成VoLTE和EPSFB的payload
                            volte_payload = get_volte_payload()
                            volte_payload = set_payload_time(volte_payload, start_date + ' 00:00:00', end_date + ' 23:59:59')
                            volte_payload = set_payload_city(volte_payload, city)

                            epsfb_payload = get_epsfb_payload()
                            epsfb_payload = set_payload_time(epsfb_payload, start_date + ' 00:00:00', end_date + ' 23:59:59')
                            epsfb_payload = set_payload_city(epsfb_payload, city)

                            self.update_progress_bar(int((idx / total_tables) * 100) + 10, "正在查询VoLTE数据...")
                            self.log(f"正在获取VoLTE数据...", "INFO")

                            # 使用新方法获取原始数据和合并数据
                            voice_data = self.extractor.jxcx._get_4g_voice_table_internal(volte_payload, epsfb_payload)
                            volte_df = voice_data['volte']
                            epsfb_df = voice_data['epsfb']
                            merged_df = voice_data['merged']

                            if merged_df.empty:
                                self.log(f"⚠ 未查询到数据: {table_name}", "WARNING")
                                continue

                            # 处理计算列（只在合并数据上计算）
                            try:
                                self.log(f"  正在添加计算列...", "INFO")
                                merged_df = add_4g_voice_columns(merged_df)
                            except Exception as e:
                                self.log(f"  ⚠ 添加计算列异常: {e}", "WARNING")

                            self.update_progress_bar(int(((idx + 0.8) / total_tables) * 100), "正在保存数据...")

                            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                            filename = f'{table_name}_{city}_{timestamp}.xlsx'
                            filepath = self.extractor.save_4g_voice_to_excel(volte_df, epsfb_df, merged_df, filename)

                            total_rows = len(volte_df) + len(epsfb_df) + len(merged_df)
                        else:
                            self.update_progress_bar(int((idx / total_tables) * 100) + 10, "正在查询数据...")
                            df = self.extractor.jxcx.get_table(payload, to_df=True)

                            if df.empty:
                                self.log(f"⚠ 未查询到数据: {table_name}", "WARNING")
                                continue

                            # 处理计算列
                            calc_columns = table_config.get('calc_columns', [])
                            if calc_columns:
                                try:
                                    self.log(f"  正在添加计算列...", "INFO")
                                    if '4G全程完好率' in calc_columns:
                                        df = add_wanchenglv_columns(df, '4g')
                                    elif '5G全程完好率' in calc_columns:
                                        df = add_wanchenglv_columns(df, '5g')
                                    elif '5G语音小区' in calc_columns:
                                        df = add_5g_voice_columns(df)
                                except Exception as e:
                                    import traceback
                                    error_detail = traceback.format_exc()
                                    self.log(f"  ⚠ 添加计算列异常: {e}", "WARNING")
                                    self.log(f"  ⚠ 详细错误: {error_detail[:500]}", "WARNING")
                                    # 打印DataFrame列信息帮助诊断
                                    if not df.empty:
                                        self.log(f"  ⚠ DataFrame列名: {list(df.columns)}", "WARNING")
                                    else:
                                        self.log(f"  ⚠ DataFrame为空，无法添加计算列", "WARNING")

                            self.update_progress_bar(int(((idx + 0.8) / total_tables) * 100), "正在保存数据...")

                            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                            filename = f'{table_name}_{city}_{timestamp}.xlsx'
                            filepath = self.extractor.save_to_excel(df, filename)
                            total_rows = len(df)

                        if filepath:
                            self.log(f"✓ [{table_name}] 提取成功！共 {total_rows} 条记录", "SUCCESS")
                            self.log(f"  文件已保存: {filepath}", "SUCCESS")
                            success_count += 1
                        else:
                            self.log(f"✗ [{table_name}] 保存文件失败", "ERROR")
                            failed_count += 1

                    except Exception as e:
                        self.log(f"✗ [{table_name}] 提取异常: {e}", "ERROR")
                        failed_count += 1
                        continue

            self.log("", "INFO")
            self.log("=" * 50, "INFO")
            self.log(f"批量提取完成！成功: {success_count}/{total_tables}", "SUCCESS" if success_count == total_tables else "INFO")
            if failed_count > 0:
                self.log(f"失败: {failed_count}/{total_tables}", "WARNING")
            self.log("=" * 50, "INFO")

            self.root.after(0, self._extract_success_ui)

        except Exception as e:
            self.log(f"提取数据异常: {e}", "ERROR")
            import traceback
            self.log(traceback.format_exc(), "ERROR")
            self.root.after(0, self._extract_failed_ui)

    def _extract_success_ui(self):
        """提取成功UI"""
        self.extract_btn.config(state=tk.NORMAL)
        self.stop_btn.config(state=tk.DISABLED)
        self.update_progress_bar(100, "提取完成")
        self.status_text.config(text="提取完成")

    def _extract_failed_ui(self):
        """提取失败UI"""
        self.extract_btn.config(state=tk.NORMAL)
        self.stop_btn.config(state=tk.DISABLED)
        self.update_progress_bar(0, "提取失败")
        self.status_text.config(text="提取失败")

    def _show_query_blocked_dialog(self, start_date, end_date):
        """显示查询被阻止的弹窗"""
        messagebox.showwarning(
            "查询被阻止",
            f"查询日期范围 ({start_date} 至 {end_date}) 超过授权到期日期 ({self.expiry_time})\n\n"
            f"请修改查询日期或联系管理员更新授权。"
        )

    def stop_extract(self):
        """停止提取"""
        messagebox.showinfo("提示", "停止功能开发中...")
        self.stop_btn.config(state=tk.DISABLED)
        self.extract_btn.config(state=tk.NORMAL)

    def clear_log(self):
        """清空日志"""
        self.log_text.config(state='normal')
        self.log_text.delete(1.0, tk.END)
        self.log_text.config(state='disabled')
        self.log("界面日志已清空", "INFO")

    def log(self, message, level="INFO"):
        """添加日志"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_queue.put((timestamp, message, level))

        color_map = {
            "INFO": "#3c4043",
            "DEBUG": "#80868b",
            "SUCCESS": "#34a853",
            "WARNING": "#f57c00",
            "ERROR": "#ea4335"
        }
        color = color_map.get(level, "#3c4043")

        self.log_text.config(state='normal')
        log_line = f"[{timestamp}] {message}\n"
        self.log_text.insert(tk.END, log_line, color)
        self.log_text.see(tk.END)
        self.log_text.config(state='disabled')

        self.logger.log(logging.INFO, message)

    def update_log(self):
        """更新日志显示（保持向后兼容）"""
        self.root.after(100, self.update_log)

    def open_output_dir(self):
        """打开输出目录"""
        output_dir = os.path.abspath(OUTPUT_DIR)
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)

        import platform
        system = platform.system()

        try:
            if system == "Windows":
                os.startfile(output_dir)
            elif system == "Darwin":
                os.system(f'open "{output_dir}"')
            else:
                os.system(f'xdg-open "{output_dir}"')
            self.log(f"已打开输出目录: {output_dir}", "SUCCESS")
        except Exception as e:
            self.log(f"打开目录失败: {e}", "ERROR")

    def get_date_string(self, year_var, month_var, day_var):
        """获取日期字符串"""
        year = year_var.get()
        month = month_var.get()
        day = day_var.get()

        try:
            last_day = calendar.monthrange(year, month)[1]
            day = min(day, last_day)
            return f"{year}-{month:02d}-{day:02d}"
        except ValueError:
            return None

    def load_config(self):
        """加载配置"""
        self.log("通用数据提取工具已就绪", "INFO")
        self.log(f"支持的数据表: {', '.join(TableConfig.get_table_names())}", "INFO")

    def on_closing(self):
        """窗口关闭"""
        self.logger.info("=" * 50)
        self.logger.info("通用数据提取工具 GUI 关闭")
        self.logger.info("=" * 50)
        self.root.destroy()


# ==================== 授权验证系统 ====================
# 新授权方式：仅使用到期日期，移除机器码绑定
# 软件超过到期日期不能打开，但修改系统时间可打开但不能查询超过到期日期的数据

def show_machine_code_dialog(machine_code):
    """显示机器码弹窗，供用户复制发给管理员"""
    import tkinter as tk
    from tkinter import ttk

    root = tk.Tk()
    root.title("机器码 - 未授权")
    root.geometry("550x320")
    root.resizable(False, False)

    # 居中显示
    root.update_idletasks()
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    x = (screen_width - 550) // 2
    y = (screen_height - 320) // 2
    root.geometry(f"550x320+{x}+{y}")

    main_frame = ttk.Frame(root, padding="20")
    main_frame.pack(fill=tk.BOTH, expand=True)

    # 标题
    ttk.Label(main_frame, text="当前机器未授权",
              font=("Microsoft YaHei UI", 14, "bold"),
              foreground="#D32F2F").pack(pady=(0, 10))

    # 提示信息
    ttk.Label(main_frame, text="您的机器码如下，请复制后发给管理员添加到授权列表：",
              font=("Microsoft YaHei UI", 10)).pack(anchor=tk.W, pady=(0, 5))

    # 机器码文本框（可复制）
    text_frame = ttk.Frame(main_frame)
    text_frame.pack(fill=tk.BOTH, expand=True, pady=5)

    machine_text = tk.Text(text_frame, height=4, font=("Courier New", 11), wrap=tk.WORD,
                           bg="#F5F5F5", fg="#333333")
    machine_text.insert("1.0", machine_code)
    machine_text.config(state=tk.DISABLED)  # 只读但可复制
    machine_scroll = ttk.Scrollbar(text_frame, orient=tk.VERTICAL, command=machine_text.yview)
    machine_text.config(yscrollcommand=machine_scroll.set)
    machine_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    machine_scroll.pack(side=tk.RIGHT, fill=tk.Y)

    # 复制按钮
    def copy_to_clipboard():
        root.clipboard_clear()
        root.clipboard_append(machine_code)
        copy_btn.config(text="已复制!", state=tk.DISABLED)
        root.after(2000, lambda: copy_btn.config(text="复制机器码", state=tk.NORMAL))

    button_frame = ttk.Frame(main_frame)
    button_frame.pack(fill=tk.X, pady=(15, 0))

    copy_btn = ttk.Button(button_frame, text="复制机器码", command=copy_to_clipboard)
    copy_btn.pack(side=tk.LEFT, padx=5)

    exit_btn = ttk.Button(button_frame, text="退出程序", command=root.destroy)
    exit_btn.pack(side=tk.LEFT, padx=5)

    root.mainloop()


def show_request_dialog(expiry_time_str, reason="授权不可用", machine_code=""):
    """显示验证请求弹窗

    Args:
        expiry_time_str: 到期日期字符串
        reason: 失败原因（"授权已过期" 或 "机器码不匹配"）
        machine_code: 当前机器码，用于显示
    """
    import tkinter as tk
    from tkinter import ttk

    root = tk.Tk()
    root.title("授权验证请求")
    root.geometry("550x420")
    root.resizable(False, False)

    # 居中显示
    root.update_idletasks()
    x = (root.winfo_screenwidth() // 2) - (275)
    y = (root.winfo_screenheight() // 2) - (210)
    root.geometry(f"550x420+{x}+{y}")

    main_frame = ttk.Frame(root, padding="20")
    main_frame.pack(fill=tk.BOTH, expand=True)

    # 警告图标
    ttk.Label(main_frame, text="⚠️", font=("Arial", 32)).pack(pady=(0, 15))

    # 标题（根据原因显示不同标题）
    title_text = "授权已过期" if reason == "授权已过期" else "授权验证失败"
    ttk.Label(main_frame, text=title_text,
              font=("Microsoft YaHei UI", 14, "bold"),
              foreground="#D32F2F").pack(pady=(0, 10))

    # 到期日期信息
    if expiry_time_str:
        ttk.Label(main_frame, text=f"到期日期: {expiry_time_str}",
                  font=("Microsoft YaHei UI", 10),
                  foreground="#666666").pack(pady=(0, 5))

    # 失败原因
    ttk.Label(main_frame, text=reason,
              font=("Microsoft YaHei UI", 10),
              foreground="#666666").pack(pady=(5, 10))

    # 机器码信息
    if machine_code:
        ttk.Label(main_frame, text="您的机器码（发给管理员）：",
                  font=("Microsoft YaHei UI", 10)).pack(anchor=tk.W, pady=(5, 5))

        text_frame = ttk.Frame(main_frame)
        text_frame.pack(fill=tk.BOTH, expand=True, pady=5)

        machine_text = tk.Text(text_frame, height=3, font=("Courier New", 10), wrap=tk.WORD,
                              bg="#F5F5F5", fg="#333333")
        machine_text.insert("1.0", machine_code)
        machine_text.config(state=tk.DISABLED)
        machine_scroll = ttk.Scrollbar(text_frame, orient=tk.VERTICAL, command=machine_text.yview)
        machine_text.config(yscrollcommand=machine_scroll.set)
        machine_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        machine_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        def copy_to_clipboard():
            root.clipboard_clear()
            root.clipboard_append(machine_code)
            copy_btn.config(text="已复制!", state=tk.DISABLED)
            root.after(2000, lambda: copy_btn.config(text="复制机器码", state=tk.NORMAL))

        copy_btn = ttk.Button(main_frame, text="复制机器码", command=copy_to_clipboard)
        copy_btn.pack(pady=(5, 0))

    # 提示信息
    ttk.Label(main_frame, text="请联系管理员获取新的授权",
              font=("Microsoft YaHei UI", 11),
              wraplength=480, justify="center").pack(pady=(10, 0))

    # 确定按钮
    ttk.Button(main_frame, text="确定", command=root.destroy).pack(pady=(20, 0))

    root.mainloop()


def verify_license():
    """验证授权，返回 (is_valid, error_msg, expiry_time, machine_code)

    逻辑：
    1. 检查 license.dat 是否存在，不存在则显示机器码弹窗并退出
    2. 读取 license.dat 中的机器码，与当前机器码比对
    3. 检查系统时间是否超过 EXPIRY_DATE
    """
    print("[DEBUG] 开始授权验证...")

    # 生成当前机器码
    current_machine_code = generate_machine_code(get_hw_info())
    print(f"[DEBUG] 当前机器码: {current_machine_code}")
    print(f"[DEBUG] 授权到期日期: {EXPIRY_DATE}")

    # 检查 license.dat 是否存在
    if not os.path.exists(LICENSE_FILE):
        print("[ERROR] 未找到 license.dat 文件")
        show_machine_code_dialog(current_machine_code)
        return False, "未找到授权文件（license.dat），请联系管理员获取", None, current_machine_code

    # 读取 license.dat 中的机器码
    try:
        with open(LICENSE_FILE, "rb") as f:
            license_data = f.read()

        # 尝试解析为有签名格式：SN长度(4字节) + SN + "|" + 签名
        if len(license_data) > 4:
            try:
                sn_len = struct.unpack(">I", license_data[:4])[0]
                if sn_len == 64 and len(license_data) > 4 + sn_len + 1:
                    stored_machine_code = license_data[4:4 + sn_len].decode("utf-8")
                else:
                    stored_machine_code = license_data.decode("utf-8").split("|")[0]
            except:
                # 简单格式：纯机器码
                stored_machine_code = license_data.decode("utf-8").strip()
        else:
            stored_machine_code = license_data.decode("utf-8").strip()

        print(f"[DEBUG] license.dat 中的机器码: {stored_machine_code}")

    except Exception as e:
        print(f"[ERROR] 读取 license.dat 失败: {e}")
        show_machine_code_dialog(current_machine_code)
        return False, "授权文件读取失败，请联系管理员", None, current_machine_code

    # 检查机器码是否匹配
    if stored_machine_code != current_machine_code:
        print("[ERROR] 机器码不匹配")
        show_request_dialog(EXPIRY_DATE, "机器码不匹配", current_machine_code)
        return False, "机器码不匹配，此授权文件不适用于当前机器", None, current_machine_code

    # 检查是否过期
    try:
        expiry_time = datetime.strptime(EXPIRY_DATE, "%Y-%m-%d")
    except ValueError:
        print("[ERROR] 授权日期格式错误")
        return False, "授权日期格式错误，请检查 EXPIRY_DATE 格式（应为 YYYY-MM-DD）", None, current_machine_code

    current_time = datetime.now()
    today = current_time.replace(hour=0, minute=0, second=0, microsecond=0)

    if today > expiry_time:
        print("[WARNING] 授权已过期，显示验证请求...")
        show_request_dialog(EXPIRY_DATE, "授权已过期", current_machine_code)
        return False, f"授权已过期（到期日期: {EXPIRY_DATE}）", EXPIRY_DATE, current_machine_code

    print("[SUCCESS] 授权验证通过")
    return True, None, EXPIRY_DATE, current_machine_code


def main():
    """主函数"""
    import tkinter as tk
    print("[MAIN] 程序启动")
    ensure_dirs()

    # 授权验证
    is_valid, error_msg, expiry_time, machine_code = verify_license()
    print(f"[MAIN] 验证完成: is_valid={is_valid}")

    if not is_valid:
        # 授权失败（机器码未授权或已过期），弹窗已在 verify_license 中显示
        print("[MAIN] 授权验证失败，退出程序")
        sys.exit(1)

    # 授权通过，打开软件
    root = tk.Tk()
    app = UniversalExtractorGUI(root, expiry_time=expiry_time)
    root.protocol("WM_DELETE_WINDOW", app.on_closing)
    root.mainloop()


if __name__ == '__main__':
    main()
