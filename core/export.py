# -*- coding: utf-8 -*-
"""
数据导出模块
负责数据导出到Excel文件
"""

import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from utils.config import OUTPUT_DIR
from utils.logger import ensure_dirs


def export_to_excel(data, filename, sheet_name='Sheet1', append=False):
    """导出数据到Excel文件

    Args:
        data: DataFrame 或 dict with 'data' key
        filename: 文件名
        sheet_name: 工作表名称
        append: 是否追加模式（多个sheet写入同一文件）

    Returns:
        str: 导出文件的完整路径
    """
    ensure_dirs()

    if isinstance(data, dict) and 'data' in data:
        df = pd.DataFrame(data['data'])
    elif isinstance(data, pd.DataFrame):
        df = data
    else:
        print(f"[ERROR-EXPORT] 不支持的数据格式: {type(data)}")
        return None

    if df.empty:
        print("[WARNING-EXPORT] 数据为空，不导出")
        return None

    filepath = os.path.join(OUTPUT_DIR, filename)

    try:
        if append and os.path.exists(filepath):
            with pd.ExcelWriter(filepath, engine='openpyxl', mode='a') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
            print(f"[SUCCESS-EXPORT] 已追加数据到 {filepath}")
        else:
            df.to_excel(filepath, sheet_name=sheet_name, index=False, engine='openpyxl')
            print(f"[SUCCESS-EXPORT] 数据已导出到 {filepath}")

        return filepath

    except Exception as e:
        print(f"[ERROR-EXPORT] 导出失败: {e}")
        import traceback
        traceback.print_exc()
        return None


def format_excel(filepath, header_color='165DFF', font_size=11):
    """格式化Excel文件

    Args:
        filepath: Excel文件路径
        header_color: 表头颜色（RGB hex）
        font_size: 字体大小
    """
    if not os.path.exists(filepath):
        print(f"[ERROR-FORMAT] 文件不存在: {filepath}")
        return

    try:
        wb = load_workbook(filepath)
        ws = wb.active

        header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=font_size)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(filepath)
        print(f"[SUCCESS-FORMAT] Excel格式化完成: {filepath}")

    except Exception as e:
        print(f"[ERROR-FORMAT] 格式化失败: {e}")
        import traceback
        traceback.print_exc()


def export_with_format(data, filename, sheet_name='Sheet1', header_color='165DFF'):
    """导出数据到Excel并格式化

    Args:
        data: DataFrame 或 dict with 'data' key
        filename: 文件名
        sheet_name: 工作表名称
        header_color: 表头颜色

    Returns:
        str: 导出文件的完整路径
    """
    filepath = export_to_excel(data, filename, sheet_name)
    if filepath:
        format_excel(filepath, header_color)
    return filepath
